import os
import subprocess
import sys
import argparse
import shutil
import time
from pathlib import Path
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent.absolute()
TRACKED_FILE = str(SCRIPT_DIR / "tracked_users.txt")
EXCEL_FILE = str(SCRIPT_DIR / "stats.xlsx")
LOCK_FILE = str(SCRIPT_DIR / "stats.xlsx.lock")

class FileLock:
    """Simple file-based lock to prevent concurrent Excel writes."""
    def __init__(self, lock_file, timeout=20, delay=0.1):
        self.lock_file = lock_file
        self.timeout = timeout
        self.delay = delay
        self._fd = None

    def __enter__(self):
        start_time = time.time()
        while True:
            try:
                # Exclusive creation of lock file
                self._fd = os.open(self.lock_file, os.O_CREAT | os.O_EXCL | os.O_RDWR)
                break
            except FileExistsError:
                # Check for stale lock (older than 60 seconds)
                try:
                    if os.path.exists(self.lock_file) and time.time() - os.stat(self.lock_file).st_mtime > 300:
                        try:
                            os.remove(self.lock_file)
                        except OSError:
                            pass
                        continue
                except OSError:
                    pass
                if time.time() - start_time >= self.timeout:
                    raise TimeoutError(f"Could not acquire lock on {self.lock_file}")
                time.sleep(self.delay)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self._fd is not None:
            os.close(self._fd)
            try:
                os.remove(self.lock_file)
            except OSError:
                pass


def safe_save_workbook(wb, filepath: str) -> bool:
    """Safely save a workbook using atomic write to prevent corruption.
    
    Writes to a temp file first, then atomically replaces the target file.
    
    Args:
        wb: The openpyxl Workbook object to save
        filepath: Path to the Excel file
        
    Returns:
        bool: True if save succeeded, False otherwise
    """
    temp_path = str(filepath) + ".tmp"
    backup_path = str(filepath) + ".backup"
    
    try:
        # 1. Save to temporary file first
        wb.save(temp_path)
        
        # 2. Create backup of existing file
        if os.path.exists(filepath):
            try:
                shutil.copy2(filepath, backup_path)
            except Exception as backup_err:
                print(f"[WARNING] Failed to create backup: {backup_err}", flush=True)
        
        # 3. Atomic replace
        os.replace(temp_path, filepath)
        print(f"[SAVE] Successfully saved: {filepath}", flush=True)
        
        # 4. Cleanup backup
        if os.path.exists(backup_path):
            try:
                os.remove(backup_path)
            except Exception:
                pass  # Not critical if backup removal fails
        
        return True
        
    except Exception as save_err:
        print(f"[ERROR] Failed to save workbook: {save_err}", flush=True)
        # Clean up temp file
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass
        
        return False
        
    finally:
        # Always try to close the workbook
        try:
            wb.close()
            print(f"[CLEANUP] Workbook closed", flush=True)
        except Exception as close_err:
            print(f"[WARNING] Error closing workbook: {close_err}", flush=True)


def load_tracked_users() -> list[str]:
    """Load tracked usernames from tracked_users.txt."""
    if not os.path.exists(TRACKED_FILE):
        return []
    with open(TRACKED_FILE, "r", encoding="utf-8") as f:
        lines = [l.strip() for l in f.readlines() if l.strip()]
    return lines


def rotate_daily_to_yesterday() -> dict:
    """Copy daily snapshot (column F) to yesterday snapshot (column H) for all tracked users.
    
    Returns:
        Dict with results: {username: success}
    """
    if not os.path.exists(EXCEL_FILE):
        print("[SKIP] Excel file not found", flush=True)
        return {}
    
    wb = None
    try:
        with FileLock(LOCK_FILE):
            # FAILSAFE: Load workbook with guaranteed cleanup
            wb = load_workbook(EXCEL_FILE)
            users = load_tracked_users()
            results = {}
            
            for username in users:
                if username not in wb.sheetnames:
                    print(f"[SKIP] {username} - sheet not found", flush=True)
                    results[username] = False
                    continue
                
                ws = wb[username]
                
                # Check if sheet has the expected layout (column F = Daily Snapshot, H = Yesterday Snapshot)
                header_f = ws.cell(row=1, column=6).value
                header_h = ws.cell(row=1, column=8).value
                
                if header_f != "Daily Snapshot" or header_h != "Yesterday Snapshot":
                    print(f"[SKIP] {username} - unexpected layout", flush=True)
                    results[username] = False
                    continue
                
                # Copy all values from column F (Daily Snapshot) to column H (Yesterday Snapshot)
                row = 2
                copied_rows = 0
                while True:
                    daily_val = ws.cell(row=row, column=6).value
                    if daily_val is None and row > 100:  # Stop if we've gone past data
                        break
                    
                    # Copy daily snapshot to yesterday snapshot
                    ws.cell(row=row, column=8, value=daily_val)
                    if daily_val is not None:
                        copied_rows += 1
                    row += 1
                
                print(f"[OK] {username} - copied {copied_rows} rows from daily to yesterday", flush=True)
                results[username] = True
            
            # Use safe save with backup and error recovery
            save_success = safe_save_workbook(wb, EXCEL_FILE)
            if not save_success:
                print("[ERROR] Failed to save Excel file after rotation", flush=True)
                return {u: False for u in users}  # Mark all as failed
            
            print(f"\n[SUMMARY] Rotated daily->yesterday for {sum(results.values())}/{len(users)} users", flush=True)
            return results
        
    except Exception as e:
        print(f"[ERROR] Exception during rotate_daily_to_yesterday: {e}", flush=True)
        return {}
        
    finally:
        # FAILSAFE: Always close workbook even if an error occurs
        if wb is not None:
            try:
                wb.close()
                print("[CLEANUP] Workbook closed", flush=True)
            except Exception as close_err:
                print(f"[WARNING] Error closing workbook: {close_err}", flush=True)


def run_api_get(username: str, api_key: str, snapshot_flags: list[str]) -> bool:
    """Run api_get.py for a user with given snapshot flags.
    
    Note: api_key parameter is ignored since api_get.py only reads from API_KEY.txt
    
    Returns True if successful, False otherwise.
    """
    try:
        cmd = [sys.executable, "api_get.py", "-ign", username]
        # api_get.py doesn't accept -key parameter, it only reads from API_KEY.txt
        cmd.extend(snapshot_flags)
        
        result = subprocess.run(cmd, cwd=str(SCRIPT_DIR), capture_output=True, text=True, timeout=30)
        if result.returncode != 0:
            print(f"[ERROR] api_get.py failed for {username}", flush=True)
            print(f"  stdout: {result.stdout}", flush=True)
            print(f"  stderr: {result.stderr}", flush=True)
        return result.returncode == 0
    except subprocess.TimeoutExpired:
        print(f"[ERROR] api_get.py timed out for {username} after 30 seconds", flush=True)
        return False
    except Exception as e:
        print(f"[ERROR] Failed to run api_get.py for {username}: {e}", flush=True)
        return False


def batch_update(schedule: str, api_key: str | None = None) -> dict:
    """Update all tracked users with appropriate snapshots.
    
    Args:
        schedule: One of 'session', 'daily', 'yesterday', 'monthly', 'all', or 'all-session'
        api_key: Optional Hypixel API key; falls back to env var or hardcoded default
    
    Returns:
        Dict with results: {username: (success, snapshots_taken)}
    """
    # Special handling for 'yesterday' schedule - rotate daily->yesterday without API calls
    if schedule == 'yesterday':
        print("[INFO] Running yesterday rotation (copying daily->yesterday snapshots)", flush=True)
        results = rotate_daily_to_yesterday()
        # Return results in the expected format
        return {username: (success, ['rotate']) for username, success in results.items()}
    
    users = load_tracked_users()
    if not users:
        print("[INFO] No tracked users found", flush=True)
        return {}
    
    if api_key is None:
        api_key = os.environ.get("HYPIXEL_API_KEY") or "0adb2317-d343-4275-aa22-e7a980eb59df"
    
    results = {}
    
    # Map schedule to snapshot types
    schedule_map = {
        'session': ['-session'],
        'daily': ['-daily'],
        'monthly': ['-monthly'],
        'all': ['-daily', '-monthly'],
        'all-session': ['-session', '-daily', '-monthly']
    }
    
    print(f"[INFO] Processing {len(users)} tracked users with schedule '{schedule}'...", flush=True)
    for idx, username in enumerate(users, 1):
        snapshots_to_take = schedule_map.get(schedule, [])
        
        if not snapshots_to_take:
            print(f"[SKIP] {username} - invalid schedule", flush=True)
            results[username] = (True, [])
            continue
        
        print(f"[RUN] [{idx}/{len(users)}] {username} - updating stats and taking snapshots: {', '.join(snapshots_to_take)}", flush=True)
        
        # Always update current stats first (column B), then take snapshots
        # This ensures the all-time stats are fresh before calculating deltas
        success = run_api_get(username, api_key, snapshots_to_take)
        
        if success:
            print(f"[OK] {username} - success", flush=True)
            results[username] = (True, snapshots_to_take)
        else:
            print(f"[ERROR] {username} - failed", flush=True)
            results[username] = (False, snapshots_to_take)
    
    print(f"\n[SUMMARY] Completed {sum(1 for s, _ in results.values() if s)}/{len(users)} users successfully", flush=True)
    return results


def main():
    parser = argparse.ArgumentParser(description="Batch update tracked users with API snapshots")
    parser.add_argument("-schedule", choices=["session", "daily", "yesterday", "monthly", "all", "all-session"], default="all",
                        help="Which snapshots to take")
    parser.add_argument("-key", "--api-key", help="Hypixel API key (optional, uses env or default)")
    args = parser.parse_args()
    
    results = batch_update(args.schedule, args.api_key)
    
    # Print summary
    successful = sum(1 for success, _ in results.values() if success)
    print(f"\n[SUMMARY] {successful}/{len(results)} users updated successfully")
    
    for username, (success, snapshots) in results.items():
        status = "[OK]" if success else "[ERROR]"
        print(f"  {status} {username}: {', '.join(snapshots) if snapshots else 'no snapshots'}")


if __name__ == "__main__":
    main()
