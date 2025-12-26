"""
Helper script to copy daily snapshot to yesterday snapshot before daily refresh.
This preserves yesterday's stats before overwriting today's daily snapshot.
"""
import os
import shutil
from pathlib import Path
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent.absolute()
EXCEL_FILE = str(SCRIPT_DIR / "stats.xlsx")
TRACKED_FILE = str(SCRIPT_DIR / "tracked_users.txt")


def safe_save_workbook(wb, filepath: str) -> bool:
    """Safely save a workbook with backup and error recovery.
    
    Creates a backup before saving. If save fails, restores from backup.
    Always ensures workbook is closed properly.
    
    Args:
        wb: The openpyxl Workbook object to save
        filepath: Path to the Excel file
        
    Returns:
        bool: True if save succeeded, False otherwise
    """
    backup_path = filepath + ".backup"
    backup_created = False
    
    try:
        # Create backup if file exists
        if os.path.exists(filepath):
            try:
                shutil.copy2(filepath, backup_path)
                backup_created = True
                print(f"[BACKUP] Created backup: {backup_path}")
            except Exception as backup_err:
                print(f"[WARNING] Failed to create backup: {backup_err}")
                # Continue anyway - we'll try to save without backup
        
        # Attempt to save
        wb.save(filepath)
        print(f"[SAVE] Successfully saved: {filepath}")
        
        # Remove backup after successful save
        if backup_created and os.path.exists(backup_path):
            try:
                os.remove(backup_path)
            except Exception:
                pass  # Not critical if backup removal fails
        
        return True
        
    except Exception as save_err:
        print(f"[ERROR] Failed to save workbook: {save_err}")
        
        # Try to restore from backup if save failed
        if backup_created and os.path.exists(backup_path):
            try:
                shutil.copy2(backup_path, filepath)
                print(f"[RESTORE] Restored from backup after save failure")
            except Exception as restore_err:
                print(f"[ERROR] Failed to restore from backup: {restore_err}")
        
        return False
        
    finally:
        # Always try to close the workbook
        try:
            wb.close()
            print(f"[CLEANUP] Workbook closed")
        except Exception as close_err:
            print(f"[WARNING] Error closing workbook: {close_err}")


def load_tracked_users() -> list[str]:
    """Load tracked usernames from tracked_users.txt."""
    if not os.path.exists(TRACKED_FILE):
        return []
    with open(TRACKED_FILE, "r", encoding="utf-8") as f:
        lines = [l.strip() for l in f.readlines() if l.strip()]
    return lines


def rotate_daily_to_yesterday():
    """Copy daily snapshot (column F) to yesterday snapshot (column H) for all users."""
    if not os.path.exists(EXCEL_FILE):
        print("[SKIP] Excel file not found")
        return
    
    wb = None
    try:
        wb = load_workbook(EXCEL_FILE)
        users = load_tracked_users()
        updated_count = 0
        
        for username in users:
            if username not in wb.sheetnames:
                print(f"[SKIP] {username} - sheet not found")
                continue
            
            ws = wb[username]
            
            # Check if sheet has the expected layout (column F = Daily Snapshot, H = Yesterday Snapshot)
            header_f = ws.cell(row=1, column=6).value
            header_h = ws.cell(row=1, column=8).value
            
            if header_f != "Daily Snapshot" or header_h != "Yesterday Snapshot":
                print(f"[SKIP] {username} - unexpected layout")
                continue
            
            # Copy all values from column F (Daily Snapshot) to column H (Yesterday Snapshot)
            row = 2
            copied_rows = 0
            while True:
                daily_val = ws.cell(row=row, column=6).value
                if daily_val is None and row > 100:  # Stop if we've gone past data
                    break
                
                # Copy daily snapshot to yesterday snapshot
                ws.cell(row=row, column=8, value=daily_val)
                if daily_val is not None:
                    copied_rows += 1
                row += 1
            
            print(f"[OK] {username} - copied {copied_rows} rows from daily to yesterday")
            updated_count += 1
        
        # Use safe save with backup and error recovery
        save_success = safe_save_workbook(wb, EXCEL_FILE)
        if not save_success:
            print("[ERROR] Failed to save Excel file after rotation")
            return 0
        
        print(f"\n[SUMMARY] Rotated daily→yesterday for {updated_count}/{len(users)} users")
        return updated_count
        
    except Exception as e:
        print(f"[ERROR] Exception during rotate_daily_to_yesterday: {e}")
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        return 0


if __name__ == "__main__":
    rotate_daily_to_yesterday()
