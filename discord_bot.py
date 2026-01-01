import discord
from discord.ext import commands
import subprocess
import sys
import openpyxl
from openpyxl import load_workbook
import os

# Database imports
from db_helper import (
    init_database,
    get_all_usernames,
    get_database_stats,
    get_user_stats_with_deltas,
    get_user_meta,
    update_user_meta,
    user_exists,
    get_discord_id,
    set_discord_link,
    get_all_user_links,
    get_default_username,
    set_default_username,
    get_all_default_users,
    get_tracked_streaks,
    update_tracked_streaks,
    get_all_tracked_streaks,
    add_tracked_user,
    remove_tracked_user,
    is_tracked_user,
    get_tracked_users,
    set_tracked_users
)
import re
import shutil
from zoneinfo import ZoneInfo
import json
from pathlib import Path
import io
import math
import requests
import asyncio
import time
from typing import Optional, Union
try:
    from PIL import Image, ImageDraw, ImageFont
except Exception:
    Image = None

# Get the directory where discord.py is located
BOT_DIR = Path(__file__).parent.absolute()

# tracked/users + creator info
# TRACKED_FILE - now in database (tracked_users table)
# JSON files are now replaced with database tables
# USER_LINKS_FILE, USER_COLORS_FILE, DEFAULT_USERS_FILE, TRACKED_STREAKS_FILE - now in database
CREATOR_NAME = "chuckegg"
# Optionally set a numeric Discord user ID for direct DM (recommended for reliability)
CREATOR_ID = "542467909549555734"
ADMIN_IDS = ["542467909549555734", "1040340714824937554"]
ADMIN_NAMES = ["chuckegg", "felix.6554"]
CREATOR_TZ = ZoneInfo("America/New_York")

START_TIME = time.time()

# Font cache to avoid repeatedly searching for fonts
_FONT_CACHE = {}

def _get_font_path(font_name: str) -> str:
    """Find the full path to a TrueType font file.
    
    Searches local fonts directory first, then common system font directories 
    on Windows, Linux, and macOS.
    
    Args:
        font_name: Name of the font file (e.g., 'DejaVuSans.ttf')
    
    Returns:
        Full path to the font file if found, otherwise returns the font_name as-is
        (will fall back to default font if not found)
    """
    if font_name in _FONT_CACHE:
        return _FONT_CACHE[font_name]
    
    # Check local fonts directory first (bundled with bot)
    local_fonts_dir = os.path.join(BOT_DIR, 'fonts')
    local_font_path = os.path.join(local_fonts_dir, font_name)
    if os.path.exists(local_font_path):
        _FONT_CACHE[font_name] = local_font_path
        return local_font_path
    
    # Common font directories by OS
    font_dirs = []
    
    if sys.platform == 'win32':
        # Windows font directories
        font_dirs = [
            os.path.expandvars(r'%WINDIR%\Fonts'),
            os.path.expandvars(r'%SystemRoot%\Fonts'),
        ]
    elif sys.platform == 'darwin':
        # macOS font directories
        font_dirs = [
            os.path.expanduser('~/Library/Fonts'),
            '/Library/Fonts',
            '/System/Library/Fonts',
        ]
    else:
        # Linux and other Unix-like systems
        font_dirs = [
            os.path.expanduser('~/.fonts'),
            '/usr/share/fonts',
            '/usr/local/share/fonts',
            '/usr/share/fonts/truetype',
            '/usr/share/fonts/truetype/dejavu',
        ]
    
    # Search for the font
    for directory in font_dirs:
        font_path = os.path.join(directory, font_name)
        if os.path.exists(font_path):
            _FONT_CACHE[font_name] = font_path
            return font_path
    
    # If not found, return the original name and let Pillow handle it
    _FONT_CACHE[font_name] = font_name
    return font_name

def _load_font(font_name: str, font_size: int):
    path = os.path.join(BOT_DIR, 'fonts', font_name)
    try:
        return ImageFont.truetype(path, font_size)
    except:
        return ImageFont.load_default()

LOCK_FILE = str(BOT_DIR / "stats.xlsx.lock")  # Kept for backward compatibility
DB_FILE = BOT_DIR / "stats.db"

class FileLock:
    """Simple file-based lock to prevent concurrent Excel writes."""
    def __init__(self, lock_file, timeout=20, delay=0.1):
        self.lock_file = lock_file
        self.timeout = timeout
        self.delay = delay
        self._fd = None

    def __enter__(self):
        start_time = time.time()
        while True:
            try:
                # Exclusive creation of lock file
                self._fd = os.open(self.lock_file, os.O_CREAT | os.O_EXCL | os.O_RDWR)
                break
            except FileExistsError:
                # Check for stale lock (older than 60 seconds)
                try:
                    if os.path.exists(self.lock_file) and time.time() - os.stat(self.lock_file).st_mtime > 300:
                        try:
                            os.remove(self.lock_file)
                        except OSError:
                            pass
                        continue
                except OSError:
                    pass
                if time.time() - start_time >= self.timeout:
                    raise TimeoutError(f"Could not acquire lock on {self.lock_file}")
                time.sleep(self.delay)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self._fd is not None:
            os.close(self._fd)
            try:
                os.remove(self.lock_file)
            except OSError:
                pass

# Global Cache System
class StatsCache:
    def __init__(self):
        self.data = {}
        self.last_mtime = 0
        self.lock = asyncio.Lock()
        self.db_path = DB_FILE

    async def get_data(self):
        """Get cached data, reloading from database if it has changed."""
        if not self.db_path.exists():
            return {}

        try:
            current_mtime = self.db_path.stat().st_mtime
            
            # Double-check locking to prevent multiple reloads
            if current_mtime > self.last_mtime:
                async with self.lock:
                    # Check again inside lock
                    if current_mtime > self.last_mtime:
                        print(f"[CACHE] Database changed. Reloading stats cache...")
                        self.data = await asyncio.to_thread(self._load_from_database)
                        self.last_mtime = current_mtime
                        print(f"[CACHE] Reload complete. Cached {len(self.data)} users.")
            
            return self.data
        except Exception as e:
            print(f"[CACHE] Error accessing cache: {e}")
            return self.data

    def _load_from_database(self):
        """Load all user data from SQLite database."""
        cache = {}

        try:
            # Get all usernames
            usernames = get_all_usernames()
            
            for username in usernames:
                try:
                    # Get stats with deltas
                    stats = get_user_stats_with_deltas(username)
                    
                    # Get metadata
                    meta_db = get_user_meta(username)
                    
                    if not stats:
                        continue
                    
                    user_cache = {"stats": stats, "meta": {}}
                    
                    # Calculate level from stats if available
                    level = int(stats.get('level', {}).get('lifetime', 0))
                    if level == 0 and 'experience' in stats:
                        level = int(stats['experience']['lifetime'] / 5000)
                    
                    # Get user color info from database
                    ign_color = meta_db.get('ign_color', '#FFFFFF') if meta_db else '#FFFFFF'
                    g_tag = meta_db.get('guild_tag') if meta_db else None
                    g_hex = meta_db.get('guild_hex', '#AAAAAA') if meta_db else '#AAAAAA'
                    
                    user_cache["meta"] = {
                        "level": level,
                        "icon": get_prestige_icon(level),
                        "ign_color": ign_color,
                        "guild_tag": g_tag,
                        "guild_hex": g_hex,
                        "username": username
                    }
                    
                    cache[username] = user_cache
                    
                except Exception as e:
                    print(f"[CACHE] Error loading user {username}: {e}")
                    continue
            
            return cache
            
        except Exception as e:
            print(f"[CACHE] Error loading from database: {e}")
            return {}

    async def update_cache_entry(self, username: str, processed_stats: dict):
        """Update a single user's cache entry from api_get.py output without reloading database."""
        async with self.lock:
            # Ensure data dict exists
            if not self.data:
                self.data = {}
            
            # Get meta from database to ensure colors/guilds are preserved
            meta_db = get_user_meta(username)
            
            # Calculate meta
            level = int(processed_stats.get('level', {}).get('lifetime', 0))
            if level == 0 and 'experience' in processed_stats:
                level = int(processed_stats['experience']['lifetime'] / 5000)
            
            ign_color = meta_db.get('ign_color', '#FFFFFF') if meta_db else '#FFFFFF'
            g_tag = meta_db.get('guild_tag') if meta_db else None
            g_hex = meta_db.get('guild_hex', '#AAAAAA') if meta_db else '#AAAAAA'

            # Update cache
            user_cache = {"stats": processed_stats, "meta": {"level": level, "icon": get_prestige_icon(level), "ign_color": ign_color, "guild_tag": g_tag, "guild_hex": g_hex, "username": username}}
            self.data[username] = user_cache

            # Update streaks if applicable
            try:
                update_streaks_from_stats(username, processed_stats)
            except Exception as e:
                print(f"[STREAK] Failed to update streaks for {username}: {e}")
            
            # Update mtime to prevent the next get_data call from reloading
            if self.db_path.exists():
                self.last_mtime = self.db_path.stat().st_mtime
            
            return user_cache

    async def refresh(self):
        """Force reload of cache from database."""
        async with self.lock:
            print("[CACHE] Forcing cache refresh...")
            self.data = await asyncio.to_thread(self._load_from_database)
            if self.db_path.exists():
                self.last_mtime = self.db_path.stat().st_mtime
            print(f"[CACHE] Refresh complete. Cached {len(self.data)} users.")

STATS_CACHE = StatsCache()

def safe_save_workbook(wb, filepath: str) -> bool:
    """Safely save a workbook using atomic write to prevent corruption.
    
    Writes to a temp file first, then atomically replaces the target file.
    
    Args:
        wb: The openpyxl Workbook object to save
        filepath: Path to the Excel file
        
    Returns:
        bool: True if save succeeded, False otherwise
    """
    temp_path = str(filepath) + ".tmp"
    backup_path = str(filepath) + ".backup"
    
    try:
        # 1. Save to temporary file first
        wb.save(temp_path)
        
        # 2. Create backup of existing file
        if os.path.exists(str(filepath)):
            try:
                shutil.copy2(str(filepath), backup_path)
            except Exception as backup_err:
                print(f"[WARNING] Failed to create backup: {backup_err}")
        
        # 3. Atomic replace
        os.replace(temp_path, str(filepath))
        print(f"[SAVE] Successfully saved: {filepath}")
        
        # 4. Cleanup backup
        if os.path.exists(backup_path):
            try:
                os.remove(backup_path)
            except Exception:
                pass  # Not critical if backup removal fails
        
        return True
        
    except Exception as save_err:
        print(f"[ERROR] Failed to save workbook: {save_err}")
        # Clean up temp file
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass
        
        return False
        
    finally:
        # Always try to close the workbook
        try:
            wb.close()
            print(f"[CLEANUP] Workbook closed")
        except Exception as close_err:
            print(f"[WARNING] Error closing workbook: {close_err}")

# sanitize output for Discord (remove problematic unicode/control chars)
def sanitize_output(text: str) -> str:
    if text is None:
        return ""
    # Replace a few common emoji with ASCII labels
    replacements = {
        '✅': '[OK]',
        '❌': '[ERROR]',
        '⚠️': '[WARNING]',
        '📊': '[DATA]',
        '📋': '[INFO]',
        '⏭️': '[SKIP]',
    }
    for k, v in replacements.items():
        text = text.replace(k, v)

    # Remove C0 control chars except newline and tab
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', str(text))
    # Collapse very long whitespace
    text = re.sub(r"\s{3,}", ' ', text)
    return text


def validate_and_normalize_ign(ign: str):
    s = str(ign or '').strip()
    if not re.fullmatch(r'^[A-Za-z0-9_]{3,16}$', s):
        return False, None
    try:
        r = requests.get(f'https://api.mojang.com/users/profiles/minecraft/{s}', timeout=5)
        if r.status_code == 200:
            data = r.json()
            name = data.get('name')
            if isinstance(name, str) and re.fullmatch(r'^[A-Za-z0-9_]{3,16}$', name):
                return True, name
            return True, s
        if r.status_code in (204, 404):
            return False, None
        return True, s
    except Exception:
        return True, s


def _to_number(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return val
    s = str(val).replace(".", "").replace(",", "").strip()
    try: return float(s)
    except: return 0

# Helper function to run scripts with proper working directory
def run_script(script_name, args, timeout=30):
    """Run a Python script in the bot directory with proper working directory"""
    return subprocess.run(
        [sys.executable, script_name, *args],
        cwd=str(BOT_DIR),
        capture_output=True,
        text=True,
        encoding='utf-8',
        errors='replace',
        timeout=timeout
    )

def run_script_batch(script_name, args):
    """Run a batch script with extended timeout (5 minutes for large user lists)"""
    return subprocess.run(
        [sys.executable, script_name, *args],
        cwd=str(BOT_DIR),
        capture_output=True,
        text=True,
        encoding='utf-8',
        errors='replace',
        timeout=300  # 5 minutes for batch operations
    )

async def ensure_user_cached(ign: str, timeout: int = 60) -> tuple[bool, str]:
    """
    Ensure a user's data is cached. If not in cache, fetch it.
    Returns (success, actual_ign) tuple.
    """
    cache_data = await STATS_CACHE.get_data()
    key = ign.casefold()
    
    # Check if already cached
    for name, data in cache_data.items():
        if name.casefold() == key:
            return True, name
    
    # Not cached, fetch it
    print(f"[CACHE] User {ign} not in cache, fetching now...")
    try:
        result = await asyncio.to_thread(run_script, "api_get.py", ["-ign", ign], timeout=timeout)
        if result.returncode != 0:
            print(f"[CACHE] Failed to fetch {ign}: {result.stderr}")
            return False, ign
        
        # Verify it's now cached
        cache_data = await STATS_CACHE.get_data()
        for name, data in cache_data.items():
            if name.casefold() == key:
                return True, name
        
        print(f"[CACHE] Fetched {ign} but still not in cache")
        return False, ign
    except Exception as e:
        print(f"[CACHE] Exception fetching {ign}: {e}")
        return False, ign

# additional imports for background tasks
import asyncio
import datetime
import random

def format_playtime(seconds: int) -> str:
    if not isinstance(seconds, (int, float)) or seconds <= 0:
        return "0s"
    seconds = int(seconds)
    days = seconds // 86400
    rem = seconds % 86400
    hours = rem // 3600
    minutes = (rem % 3600) // 60
    if days > 0: return f"{days}d {hours}h"
    if hours > 0: return f"{hours}h {minutes}m"
    return f"{minutes}m"


# Prestige icons per 100 levels (index 0 = levels 0-99)
PRESTIGE_ICONS = [
    "❤", "✙", "✫", "✈", "✠", "♙", "⚡", "☢", "✏", "☯",
    "☃️", "۞", "✤", "♫", "♚", "❉", "Σ", "￡", "✖", "❁",
    "✚", "✯", "✆", "❥", "☾⋆⁺", "⚜", "✦", "⚝", "✉", "ツ",
    "❣", "✮", "✿", "✲", "❂", "ƒ", "$", "⋚⋚", "Φ", "✌",
]

# Prestige colors (RGB tuples for Discord embed colors)
# Levels: 0, 100, 200, 300, 400, 500, 600, 700, 800, 900, 1000+
PRESTIGE_COLORS = {
    0: (119, 119, 119),      # GRAY (§7)
    100: (255, 255, 255),    # WHITE (§f)
    200: (255, 85, 85),      # RED (§c)
    300: (255, 170, 0),      # GOLD (§6)
    400: (255, 255, 85),     # YELLOW (§e)
    500: (85, 255, 85),      # LIGHT_GREEN (§a)
    600: (0, 170, 170),      # DARK_AQUA (§3)
    700: (170, 0, 170),      # DARK_PURPLE (§5)
    800: (255, 85, 255),     # LIGHT_PURPLE (§d)
    900: None,               # Rainbow (special handling)
    1000: (255, 255, 255),   # WHITE (§f)
    1100: (170, 170, 170),   # &7 -> GRAY
    1200: (255, 85, 85),     # &c -> RED
    1300: (255, 170, 0),     # &6 -> GOLD
    1400: (255, 255, 85),    # &e -> YELLOW
    1500: (85, 255, 85),     # &a -> GREEN
    1600: (0, 170, 170),     # &3 -> DARK_AQUA
    1700: (255, 85, 255),    # &d -> LIGHT_PURPLE
    1800: (170, 0, 170),     # &5 -> DARK_PURPLE
    1900: None,              # Rainbow
    2000: (0, 170, 0),       # &2 -> DARK_GREEN
    2100: (170, 170, 170),   # &7 -> GRAY
    2200: (255, 255, 85),    # &e -> YELLOW
    2300: (255, 255, 85),    # &e -> YELLOW
    2400: (85, 255, 255),    # &b -> AQUA
    2500: (85, 255, 85),     # &a -> GREEN
    2600: (85, 255, 255),    # &b -> AQUA
    2700: (255, 85, 255),    # &d -> LIGHT_PURPLE
    2800: (170, 0, 170),     # &5 -> DARK_PURPLE
    2900: None,              # Rainbow
    3000: (0, 0, 0),         # &0 -> BLACK
    3100: (255, 255, 255),   # &f -> WHITE
    3200: (255, 85, 85),     # &c -> RED
    3300: (255, 170, 0),     # &6 -> GOLD
    3400: (255, 255, 85),    # &e -> YELLOW
    3500: (85, 255, 85),     # &a -> GREEN
    3600: (0, 170, 170),     # &3 -> DARK_AQUA
    3700: (255, 85, 255),    # &d -> LIGHT_PURPLE
    3800: (170, 0, 170),     # &5 -> DARK_PURPLE
    3900: None,              # Rainbow
    4000: (85, 85, 85),      # &8 -> DARK_GRAY
    4100: (255, 255, 255),   # &f -> WHITE
    4200: (255, 85, 85),     # &c -> RED
    4300: (255, 170, 0),     # &6 -> GOLD
    4400: (255, 255, 85),    # &e -> YELLOW
    4500: (85, 255, 85),     # &a -> GREEN
    4600: (0, 170, 170),     # &3 -> DARK_AQUA
    4700: (255, 85, 255),    # &d -> LIGHT_PURPLE
    4800: (170, 0, 170),     # &5 -> DARK_PURPLE
    4900: None,              # Rainbow
    5000: None,              # Rainbow (High level default)
}


def get_prestige_icon(level: int) -> str:
    try:
        lvl = int(level)
    except Exception:
        lvl = 0
    base = (lvl // 100) * 100
    # If a raw pattern exists and contains an icon, extract it (strip color codes)
    raw = PRESTIGE_RAW_PATTERNS.get(base)
    if raw:
        stripped = re.sub(r'&[0-9a-fA-F]', '', raw)
        # Look for content inside brackets
        m = re.search(r"\[(.*?)\]", stripped)
        if m:
            inner = m.group(1)
            # remove leading digits (the level number) to get icon
            icon = re.sub(r'^[0-9]+', '', inner).strip()
            if icon:
                return icon

    # Fallback to PRESTIGE_ICONS list
    idx = max(0, lvl // 100)
    if idx >= len(PRESTIGE_ICONS):
        idx = len(PRESTIGE_ICONS) - 1
    return PRESTIGE_ICONS[idx]

def get_prestige_color(level: int) -> tuple:
    """Get RGB color tuple for a given prestige level.
    Supports levels 0-1000. Returns default dark gray for levels outside this range.
    """
    try:
        lvl = int(level)
    except Exception:
        lvl = 0

    base = (lvl // 100) * 100

    # If a raw pattern exists for this prestige base, prefer its first color code
    raw = PRESTIGE_RAW_PATTERNS.get(base)
    if raw:
        m = re.search(r'&([0-9a-fA-F])', raw)
        if m:
            code = m.group(1).lower()
            hexcol = MINECRAFT_CODE_TO_HEX.get(code)
            if hexcol:
                return hex_to_rgb(hexcol)

    # Otherwise fall back to explicit PRESTIGE_COLORS mapping
    for prestige_level in sorted(PRESTIGE_COLORS.keys(), reverse=True):
        if lvl >= prestige_level:
            color = PRESTIGE_COLORS[prestige_level]
            # Handle Rainbow (None) by returning a default color or cycling
            if color is None:
                return (255, 100, 200)
            return color

    # Fallback to gray if below 0
    return (119, 119, 119)

def get_ansi_color_code(level: int) -> str:
    """Get ANSI color code for a given prestige level."""
    color = get_prestige_color(level)
    
    # Map RGB to closest basic ANSI color for Discord compatibility
    r, g, b = color
    
    # Determine which basic ANSI color is closest
    if r > 200 and g > 200 and b > 200:
        return "\u001b[0;37m"  # White
    elif r < 100 and g < 100 and b < 100:
        return "\u001b[0;30m"  # Gray
    elif r > 200 and g < 100 and b < 100:
        return "\u001b[0;31m"  # Red
    elif r > 200 and g > 150 and b < 100:
        return "\u001b[0;33m"  # Yellow/Gold
    elif r < 100 and g > 200 and b < 100:
        return "\u001b[0;32m"  # Green
    elif r < 100 and g > 150 and b > 150:
        return "\u001b[0;36m"  # Cyan
    elif r > 150 and g < 100 and b > 150:
        return "\u001b[0;35m"  # Magenta/Pink
    elif r > 200 and g > 200 and b < 100:
        return "\u001b[0;33m"  # Yellow
    else:
        return "\u001b[0;37m"  # Default White

def make_bold_ansi(code: str) -> str:
    """Convert a basic ANSI color code to bold variant.
    Expects codes like "\u001b[0;33m" and returns "\u001b[1;33m".
    """
    if not code:
        return code
    # If already contains bold flag, return as-is
    if "1;" in code or "\u001b[1m" in code:
        return code
    # If code already contains bold or is empty, return it
    if not code:
        return code
    if "1;" in code or "\u001b[1m" in code:
        return code
    # For any CSI like '\x1b[...m', insert '1;' after the '[' if not present
    m = re.match(r"^(\x1b\[)(?!1;)(.*)m$", code)
    if m:
        return f"{m.group(1)}1;{m.group(2)}m"
    return code


# Mapping of Minecraft color codes (§) to approximate ANSI codes for inline coloring
# Official Minecraft-ish main hex colors for § codes (main hex values)
MINECRAFT_CODE_TO_HEX = {
    '0': '#000000',
    '1': '#0000AA',
    '2': '#00AA00',
    '3': '#00AAAA',
    '4': '#AA0000',
    '5': '#AA00AA',
    '6': '#FFAA00',
    '7': '#AAAAAA',
    '8': '#555555',
    '9': '#5555FF',
    'a': '#55FF55',
    'b': '#55FFFF',
    'c': '#FF5555',
    'd': '#FF55FF',
    'e': '#FFFF55',
    'f': '#FFFFFF',
}

# Minecraft color name to hex (from Hypixel API)
MINECRAFT_NAME_TO_HEX = {
    'BLACK': '#000000',
    'DARK_BLUE': '#0000AA',
    'DARK_GREEN': '#00AA00',
    'DARK_AQUA': '#00AAAA',
    'DARK_RED': '#AA0000',
    'DARK_PURPLE': '#AA00AA',
    'GOLD': '#FFAA00',
    'GRAY': '#AAAAAA',
    'DARK_GRAY': '#555555',
    'BLUE': '#5555FF',
    'GREEN': '#55FF55',
    'AQUA': '#55FFFF',
    'RED': '#FF5555',
    'LIGHT_PURPLE': '#FF55FF',
    'YELLOW': '#FFFF55',
    'WHITE': '#FFFFFF',
}

def hex_to_rgb(h: str) -> tuple:
    h = h.lstrip('#')
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def hex_to_ansi(h: str, background: bool = False) -> str:
    r, g, b = hex_to_rgb(h)
    if background:
        return f"\u001b[48;2;{r};{g};{b}m"
    return f"\u001b[38;2;{r};{g};{b}m"

def rgb_to_ansi256_index(r: int, g: int, b: int) -> int:
    """Convert RGB 0-255 to xterm-256 color index."""
    # Grayscale approximation
    if r == g == b:
        if r < 8:
            return 16
        if r > 248:
            return 231
        return 232 + int((r - 8) / 247 * 24)

    # 6x6x6 color cube
    ri = int(round((r / 255) * 5))
    gi = int(round((g / 255) * 5))
    bi = int(round((b / 255) * 5))
    return 16 + (36 * ri) + (6 * gi) + bi

def rgb_to_ansi256_escape(r: int, g: int, b: int, background: bool = False) -> str:
    idx = rgb_to_ansi256_index(r, g, b)
    if background:
        return f"\u001b[48;5;{idx}m"
    return f"\u001b[38;5;{idx}m"

def hex_to_ansi256(h: str, background: bool = False) -> str:
    r, g, b = hex_to_rgb(h)
    return rgb_to_ansi256_escape(r, g, b, background=background)

# Map Minecraft color codes to chosen xterm-256 indices for clearer, distinct colors
# These indices were selected to improve visual separation between gold/yellow/green
_MINECRAFT_256_INDEX = {
    '0': 16,   # black
    '1': 19,   # dark_blue
    '2': 28,   # dark_green
    '3': 37,   # dark_aqua
    '4': 124,  # dark_red
    '5': 127,  # dark_purple
    '6': 214,  # gold/orange
    '7': 248,  # gray
    '8': 239,  # dark_gray
    '9': 75,   # blue
    'a': 46,   # bright green
    'b': 51,   # aqua
    'c': 203,  # red
    'd': 201,  # pink
    'e': 227,  # yellow
    'f': 15,   # white
}

MINECRAFT_CODE_TO_ANSI_SGR = {k: f"\u001b[38;5;{idx}m" for k, idx in _MINECRAFT_256_INDEX.items()}

# Keep the 24-bit hex map for embed accent colors
MINECRAFT_CODE_TO_ANSI = {k: hex_to_ansi(v) for k, v in MINECRAFT_CODE_TO_HEX.items()}

# Patterns for multi-colored prestige prefixes. Key = prestige base (e.g. 1900),
# For flexibility we store raw Minecraft-style color sequences per prestige.
# Each string uses '&' followed by hex code, e.g. '&c[&61&e9&a0&30&5✖&d]'.
# The runtime parser below converts those into (code, text) pieces.
PRESTIGE_RAW_PATTERNS = {
    0: "&7[0❤]",
    100: "&f[100✙]",
    200: "&c[200✫]",
    300: "&6[300✈]",
    400: "&e[400✠]",
    500: "&a[500♙]",
    600: "&3[600⚡]",
    700: "&5[700✠]",
    800: "&d[800✏]",
    900: "&c[&69&e0&a0&b✏&d]",
    1000: "&0[&f1000☯&0]",
    1100: "&0[&81100☃️&0]",
    1200: "&0[&c1200۞&0]",
    1300: "&0[&61300✤&0]",
    1400: "&0[&e1400♫&0]",
    1500: "&0[&a1500♚&0]",
    1600: "&0[&31600❉&0]",
    1700: "&0[&51700Σ&0]",
    1800: "&0[&d1800✖&0]",
    1900: "&c[&61&e9&a0&30&5✖&d]",
    2000: "&0[2&80&700&f❁]",
    2100: "&f[2&710&80&0✚]",
    2200: "&f[2&e20&60&c✯]",
    2300: "&c[2&630&e0&a✆]",
    2400: "&b[2&340&50&d❥]",
    2500: "&f[2&a500&2☾⋆⁺]",
    2600: "&f[2&b60&30&9⚜&1]",
    2700: "&f[2&d700&5✦]",
    2800: "&c[2&480&50&d✉]",
    2900: "&d[&52&39&a0&e0&6✉&c]",
    3000: "&f[&03&80&00&80&0ツ&f]",
    3100: "&0[&f3&71&f0&70&f❣&0]",
    3200: "&0[&c3&42&c0&40&c✮&0]",
    3300: "&0[&63&c3&60&c0&6✿&0]",
    3400: "&0[&e3&64&e0&60&e✲&0]",
    3500: "&0[&a3&25&a0&20&a❂&0]",
    3600: "&0[&33&16&30&10&3ƒ&0]",
    3700: "&0[&d3&57&d0&50&d$&0]",
    3800: "&0[&53&48&50&40&5⋚⋚&0]",
    3900: "&4[&63&e9&20&10&5Φ&d]",
    4000: "&0[4&80&70&80&0✌]",
    4900: "&4[&64&e9&a0&30&5✖&d]",
}

def _parse_raw_pattern(raw: str) -> list:
    """Parse a raw pattern into list of (code, text) pieces."""
    parts = []
    cur_code = None
    buf = ''
    i = 0
    while i < len(raw):
        ch = raw[i]
        if ch == '&' and i + 1 < len(raw):
            if buf:
                parts.append((cur_code or 'f', buf))
                buf = ''
            cur_code = raw[i+1].lower()
            i += 2
            continue
        else:
            buf += ch
            i += 1
    if buf:
        parts.append((cur_code or 'f', buf))
    return parts

def get_prestige_segments(level: int, icon: str) -> list:
    """Generate colored text segments for a prestige level."""
    base = (level // 100) * 100
    raw = PRESTIGE_RAW_PATTERNS.get(base)
    segments = []

    # Check if this should be rainbow (explicitly defined as None in colors, or ends in 900)
    # Also treat anything >= 5000 as rainbow based on user feedback
    is_rainbow = (base in PRESTIGE_COLORS and PRESTIGE_COLORS[base] is None) or (base % 1000 == 900) or (base >= 5000)
    
    if raw:
        parts = _parse_raw_pattern(raw)
        concat = ''.join(t for (_, t) in parts)
        m = re.search(r"\d+", concat)
        
        if m:
            num_start, num_end = m.start(), m.end()
            pos = 0
            replaced = False
            
            for code, text in parts:
                part_start = pos
                part_end = pos + len(text)
                pos = part_end
                hexcol = MINECRAFT_CODE_TO_HEX.get(code.lower(), '#FFFFFF')
                
                if part_end <= num_start or part_start >= num_end:
                    segments.append((hexcol, text))
                    continue
                
                # Prefix before number
                prefix_len = max(0, num_start - part_start)
                if prefix_len > 0:
                    segments.append((hexcol, text[:prefix_len]))
                
                # Replace with actual level
                if not replaced:
                    if is_rainbow:
                        colors_in_span = []
                        pos2 = 0
                        for c_code, c_text in parts:
                            part_s = pos2
                            part_e = pos2 + len(c_text)
                            pos2 = part_e
                            overlap_s = max(part_s, num_start)
                            overlap_e = min(part_e, num_end)
                            if overlap_e > overlap_s:
                                hexcol_span = MINECRAFT_CODE_TO_HEX.get(c_code.lower(), '#FFFFFF')
                                for _ in range(overlap_e - overlap_s):
                                    colors_in_span.append(hexcol_span)
                        
                        if not colors_in_span:
                            RAINBOW_CODES = ['c', '6', 'e', 'a', 'b', 'd', '9', '3']
                            colors_in_span = [MINECRAFT_CODE_TO_HEX.get(c, '#FFFFFF') for c in RAINBOW_CODES]
                        
                        for i, ch in enumerate(str(level)):
                            col = colors_in_span[i % len(colors_in_span)]
                            segments.append((col, ch))
                    else:
                        segments.append((hexcol, str(level)))
                    replaced = True
                
                # Suffix after number
                suffix_start_in_part = max(0, num_end - part_start)
                if suffix_start_in_part < len(text):
                    segments.append((hexcol, text[suffix_start_in_part:]))
        else:
            segments = [(MINECRAFT_CODE_TO_HEX.get(code, '#FFFFFF'), text) for code, text in parts]
    else:
        if is_rainbow:
            # Default rainbow behavior for undefined high levels (e.g. 4900)
            bracket_col = MINECRAFT_CODE_TO_HEX.get('8', '#555555') # Dark Gray
            segments.append((bracket_col, "["))
            
            RAINBOW_CODES = ['c', '6', 'e', 'a', 'b', 'd', '9', '3']
            rainbow_hexes = [MINECRAFT_CODE_TO_HEX.get(c, '#FFFFFF') for c in RAINBOW_CODES]
            
            for i, ch in enumerate(str(level)):
                col = rainbow_hexes[i % len(rainbow_hexes)]
                segments.append((col, ch))
                
            segments.append((bracket_col, f"{icon}]"))
        else:
            color = get_prestige_color(level)
            hexcol = '#{:02x}{:02x}{:02x}'.format(*color)
            segments = [(hexcol, f"[{level}{icon}]")]
    
    return segments

def _safe_guild_tag(guild_tag: str) -> Optional[str]:
    """Try to return guild tag, but return None if it contains problematic unicode."""
    if not guild_tag:
        return None
    # Only allow ASCII characters to prevent rendering issues
    try:
        guild_tag.encode('ascii')
        return guild_tag
    except UnicodeEncodeError:
        # Filter out non-ASCII
        cleaned = "".join(c for c in guild_tag if ord(c) < 128)
        return cleaned if cleaned else None

def render_prestige_with_text(level: int, icon: str, ign: str, suffix: str, ign_color: str = None, guild_tag: str = None, guild_color: str = None, two_line: bool = False) -> io.BytesIO:
    """Render a prestige prefix with IGN, optional guild tag, and optional suffix text.
    
    Returns a BytesIO containing the rendered PNG image.
    If Pillow is not available, raises RuntimeError.
    ign_color: Hex color code for the IGN (e.g., '#FF5555')
    guild_tag: Guild tag to display after username (e.g., 'QUEBEC')
    guild_color: Color name from Hypixel API (e.g., 'DARK_AQUA')
    two_line: If True, formats as [level icon] username [guild] on first line, suffix on second line
    """
    if Image is None:
        raise RuntimeError("Pillow not available")
    
    segments = get_prestige_segments(level, icon)
    
    # Add IGN with custom color if specified
    ign_hex = ign_color if ign_color else MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF')
    segments.append((ign_hex, f" {ign}"))
    
    # Add guild tag if provided (with safety check)
    safe_tag = _safe_guild_tag(guild_tag)
    if safe_tag and guild_color:
        if guild_color.startswith('#'):
            guild_hex = guild_color
        else:
            guild_hex = MINECRAFT_NAME_TO_HEX.get(guild_color.upper(), '#FFFFFF')
        segments.append((guild_hex, f" [{safe_tag}]"))
    elif safe_tag:
        segments.append((MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF'), f" [{safe_tag}]"))
    
    if two_line and suffix:
        # Two-line format: first line ends after guild tag, second line is the suffix
        return _render_text_segments_to_image_multiline([segments, [(MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF'), suffix)]])
    elif suffix:
        # Single line format: append suffix with " - " prefix
        segments.append((MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF'), suffix))
    
    return _render_text_segments_to_image(segments)


def _render_text_segments_to_image(segments: list, font=None, padding=(8,6)) -> io.BytesIO:
    """Render colored text segments to a PNG and return a BytesIO."""
    if Image is None:
        raise RuntimeError("Pillow not available")
    if font is None:
        font = _load_font("DejaVuSans.ttf", 18)

    # Measure total size and vertical bounds
    total_w = 0
    min_y = float('inf')
    max_y = float('-inf')
    
    draw_dummy = ImageDraw.Draw(Image.new('RGBA', (1,1)))
    
    # Handle empty segments case
    if not segments:
        min_y, max_y = 0, 0

    for color_hex, text in segments:
        bbox = draw_dummy.textbbox((0, 0), text, font=font)
        # bbox is (left, top, right, bottom)
        w = bbox[2] - bbox[0]
        total_w += w
        
        if bbox[1] < min_y: min_y = bbox[1]
        if bbox[3] > max_y: max_y = bbox[3]

    if min_y == float('inf'): min_y = 0
    if max_y == float('-inf'): max_y = 0
    
    content_h = max_y - min_y
    # Ensure minimal height
    if content_h <= 0: content_h = 10

    img_w = int(total_w + padding[0]*2)
    img_h = int(content_h + padding[1]*2)
    
    img = Image.new('RGBA', (img_w, img_h), (0,0,0,0))
    draw = ImageDraw.Draw(img)

    x = padding[0]
    # Shift drawing so the top of the ink (min_y) aligns with padding[1]
    y_draw = padding[1] - min_y
    
    for color_hex, text in segments:
        try:
            color = tuple(int(color_hex.lstrip('#')[i:i+2], 16) for i in (0,2,4))
        except Exception:
            color = (255,255,255)
        draw.text((x, y_draw), text, font=font, fill=color)
        bbox = draw.textbbox((x, y_draw), text, font=font)
        w = bbox[2] - bbox[0]
        x += w

    out = io.BytesIO()
    img.save(out, format='PNG')
    out.seek(0)
    return out


def _render_text_segments_to_image_multiline(lines: list, font=None, padding=(8,6), line_spacing=2) -> io.BytesIO:
    """Render multiple lines of colored text segments to a PNG.
    
    Args:
        lines: List of segment lists, where each segment list is [(color_hex, text), ...]
        font: Font to use
        padding: Horizontal and vertical padding
        line_spacing: Additional vertical space between lines
    """
    if Image is None:
        raise RuntimeError("Pillow not available")
    if font is None:
        font = _load_font("DejaVuSans.ttf", 26)

    draw_dummy = ImageDraw.Draw(Image.new('RGBA', (1,1)))
    
    # Measure each line
    line_widths = []
    line_heights = []
    for segments in lines:
        line_w = 0
        line_h = 0
        for color_hex, text in segments:
            bbox = draw_dummy.textbbox((0, 0), text, font=font)
            w = bbox[2] - bbox[0]
            h = bbox[3] - bbox[1]
            line_w += w
            line_h = max(line_h, h)
        line_widths.append(line_w)
        line_heights.append(line_h)
    
    # Calculate total image size
    max_w = max(line_widths) if line_widths else 0
    total_h = sum(line_heights) + (len(lines) - 1) * line_spacing if lines else 0
    
    img_w = max_w + padding[0] * 2
    img_h = total_h + padding[1] * 2
    img = Image.new('RGBA', (img_w, img_h), (0,0,0,0))
    draw = ImageDraw.Draw(img)
    
    # Draw each line (center each line horizontally)
    y = padding[1]
    for line_idx, segments in enumerate(lines):
        # Calculate starting x position to center this line
        line_width = line_widths[line_idx]
        x = (img_w - line_width) // 2
        
        for color_hex, text in segments:
            try:
                color = tuple(int(color_hex.lstrip('#')[i:i+2], 16) for i in (0,2,4))
            except Exception:
                color = (255,255,255)
            draw.text((x, y), text, font=font, fill=color)
            bbox = draw.textbbox((x, y), text, font=font)
            w = bbox[2] - bbox[0]
            x += w
        y += line_heights[line_idx] + line_spacing
    
    out = io.BytesIO()
    img.save(out, format='PNG')
    out.seek(0)
    return out


def render_stat_box(label: str, value: str, width: int = 200, height: int = 80):
    """Render a single stat box with label and value using modern card style."""
    if Image is None:
        raise RuntimeError("Pillow not available")
    
    # Determine color based on label content for consistency with sheepwars command
    color = (255, 255, 255)
    l = label.lower()
    if "win" in l or "wlr" in l or "kdr" in l:
        color = (85, 255, 85)
    elif "loss" in l:
        color = (255, 85, 85)
    elif "playtime" in l:
        color = (255, 85, 255)
        
    return render_modern_card(label, value, width, height, color=color)


def create_stats_composite_image(level, icon, ign, tab_name, wins, losses, wl_ratio, kills, deaths, kd_ratio, 
                                  ign_color=None, guild_tag=None, guild_hex=None, playtime_seconds=0,
                                  status_text="Online", status_color=(85, 255, 85), skin_image=None):
    canvas_w, canvas_h = 1200, 650
    margin, spacing = 40, 15
    composite = Image.new('RGBA', (canvas_w, canvas_h), (18, 18, 20, 255))
    
    formatted_playtime = format_playtime(playtime_seconds)
    skin_w, skin_h = 240, 285
    header_card_w = (canvas_w - (margin * 2) - skin_w - (spacing * 2)) // 2
    
    skin_card = Image.new('RGBA', (skin_w, skin_h), (0, 0, 0, 0))
    ImageDraw.Draw(skin_card).rounded_rectangle([0, 0, skin_w-1, skin_h-1], radius=15, fill=(35, 30, 45, 240))
    if skin_image:
        skin = skin_image
    else:
        skin = get_player_body(ign)
    if skin:
        skin.thumbnail((220, 260), Image.Resampling.LANCZOS)
        skin_card.paste(skin, ((skin_w - skin.width)//2, (skin_h - skin.height)//2), skin)
    composite.paste(skin_card, (margin, margin), skin_card)

    col1_x = margin + skin_w + spacing
    col2_x = col1_x + header_card_w + spacing
    
    ign_rgb = (85, 255, 255)
    if ign_color:
        try:
            ign_rgb = tuple(int(str(ign_color).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            pass

    c1 = render_modern_card("IGN", ign, header_card_w, 85, is_header=True, color=ign_rgb)
    
    # Render Level card with multi-color support
    c2 = render_modern_card("Prestige", "", header_card_w, 85)
    segs = get_prestige_segments(level, icon)
    font_lvl = _load_font("DejaVuSans-Bold.ttf", 24)
    txt_io = _render_text_segments_to_image(segs, font=font_lvl, padding=(0,0))
    txt_img = Image.open(txt_io).convert("RGBA")
    # Align vertically with other cards (which draw text centered at height * 0.6)
    c2.paste(txt_img, ((c2.width - txt_img.width) // 2, int(c2.height * 0.6 - txt_img.height / 2) + 4), txt_img)

    c3 = render_modern_card("Mode", tab_name.upper(), header_card_w, 85)
    c4 = render_modern_card("Playtime", formatted_playtime, header_card_w, 85, is_header=True, color=(255, 85, 255))
    
    g_rgb = (255, 255, 255)
    if guild_hex:
        # Handle Minecraft color names (e.g. "DARK_AQUA")
        if str(guild_hex).upper() in MINECRAFT_NAME_TO_HEX:
            guild_hex = MINECRAFT_NAME_TO_HEX[str(guild_hex).upper()]
            
        try:
            g_rgb = tuple(int(str(guild_hex).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            g_rgb = (170, 170, 170)
    c5 = render_modern_card("Guild", f"{str(guild_tag) if guild_tag else 'None'}", header_card_w, 85, color=g_rgb)
    c6 = render_modern_card("Status", status_text, header_card_w, 85, color=status_color)

    for i, card in enumerate([c1, c2, c3]):
        composite.paste(card, (col1_x, margin + i*(85+spacing)), card)
    for i, card in enumerate([c4, c5, c6]):
        composite.paste(card, (col2_x, margin + i*(85+spacing)), card)

    line_y = margin + skin_h + 25
    ImageDraw.Draw(composite).line([margin, line_y, canvas_w - margin, line_y], fill=(60, 60, 80), width=2)
    
    grid_y = line_y + 25
    cols = 3
    grid_card_w = (canvas_w - (margin * 2) - (spacing * (cols - 1))) // cols
    grid_card_h = 110
    
    stats_data = [
        ("Wins", f"{int(wins):,}", (85, 255, 85)), ("Losses", f"{int(losses):,}", (255, 85, 85)), ("WLR", f"{wl_ratio:.2f}", (85, 255, 85)),
        ("Kills", f"{int(kills):,}", (255, 255, 255)), ("Deaths", f"{int(deaths):,}", (255, 255, 255)), ("KDR", f"{kd_ratio:.2f}", (85, 255, 85))
    ]

    for i, (label, val, color) in enumerate(stats_data):
        row, col = divmod(i, cols)
        card = render_modern_card(label, val, grid_card_w, grid_card_h, color=color)
        composite.paste(card, (int(margin + col * (grid_card_w + spacing)), int(grid_y + row * (grid_card_h + spacing))), card)

    # Footer
    draw = ImageDraw.Draw(composite)
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((canvas_w - text_w) // 2, canvas_h - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    composite.convert("RGB").save(out, format='PNG')
    out.seek(0)
    return out


def create_full_stats_image(ign: str, tab_name: str, level: int, icon: str, stats: dict,
                             ign_color: str = None, guild_tag: str = None, guild_color: str = None) -> io.BytesIO:
    """Render the full stats layout defined in Template.xlsx.

    Layout rules:
    - First line: 5 boxes (Wins/Hour, EXP/Hour, Playtime, EXP/Game, Kills/Hour)
    - Remaining lines: 5 boxes each
    """
    if Image is None:
        raise RuntimeError("Pillow not available")

    # Title image with prestige icon and tab name
    title_io = render_prestige_with_text(level, icon, ign, f"{tab_name.title()} Stats", ign_color, guild_tag, guild_color, two_line=True)
    title_img = Image.open(title_io)

    box_width = 200
    box_height = 80
    spacing = 10
    max_boxes = 5
    line_width_max = box_width * max_boxes + spacing * (max_boxes - 1)

    # Build lines from the template-driven order
    lines = [
        [
            ("Wins/Hour", stats.get("wins_per_hour", "0")),
            ("Exp/Hour", stats.get("exp_per_hour", "0")),
            ("Playtime", stats.get("playtime", "0")),
            ("Exp/Game", stats.get("exp_per_game", "0")),
            ("Kills/Hour", stats.get("kills_per_hour", "0")),
        ],
        [
            ("Wins", stats.get("wins", "0")),
            ("Losses", stats.get("losses", "0")),
            ("WLR", stats.get("wlr", "0")),
            ("Damage/Game", stats.get("damage_per_game", "0")),
            ("Coins", stats.get("coins", "0")),
        ],
        [
            ("Kills", stats.get("kills", "0")),
            ("Deaths", stats.get("deaths", "0")),
            ("KDR", stats.get("kdr", "0")),
            ("Kill/Game", stats.get("kills_per_game", "0")),
            ("Kill/Win", stats.get("kills_per_win", "0")),
        ],
        [
            ("Damage dealt", stats.get("damage", "0")),
            ("Damage/Game", stats.get("damage_per_game", "0")),
            ("Void kills", stats.get("void_kills", "0")),
            ("Void deaths", stats.get("void_deaths", "0")),
            ("Void KDR", stats.get("void_kdr", "0")),
        ],
        [
            ("Magic wools", stats.get("magic_wools", "0")),
            ("Wools/Game", stats.get("wools_per_game", "0")),
            ("Explosive kills", stats.get("explosive_kills", "0")),
            ("Explosive deaths", stats.get("explosive_deaths", "0")),
            ("Explosive KDR", stats.get("explosive_kdr", "0")),
        ],
        [
            ("Sheeps thrown", stats.get("sheeps_thrown", "0")),
            ("Sheeps thrown/Game", stats.get("sheeps_per_game", "0")),
            ("Bow kills", stats.get("bow_kills", "0")),
            ("Bow deaths", stats.get("bow_deaths", "0")),
            ("Bow KDR", stats.get("bow_kdr", "0")),
        ],
        [
            ("Games Played", stats.get("games_played", "0")),
            ("Damage/Sheep", stats.get("damage_per_sheep", "0")),
            ("Meelee kills", stats.get("melee_kills", "0")),
            ("Meelee Deaths", stats.get("melee_deaths", "0")),
            ("Meelee KDR", stats.get("melee_kdr", "0")),
        ],
    ]

    # Render all boxes
    rendered_lines = []
    for line_idx, line in enumerate(lines):
        rendered = []
        for col_idx, (label, value) in enumerate(line):
            try:
                rendered.append(render_stat_box(label, str(value), width=box_width, height=box_height))
            except Exception as e:
                print(f"[WARNING] Failed to render box {label}: {e}")
        rendered_lines.append(rendered)

    # Compute overall dimensions
    line_heights = []
    line_widths = []
    for line in rendered_lines:
        line_height = box_height
        # Calculate width
        line_width = 0
        for i, box in enumerate(line):
            line_width += box.width
            if i < len(line) - 1:
                line_width += spacing
        line_heights.append(line_height)
        line_widths.append(line_width)

    grid_height = sum(line_heights) + spacing * (len(rendered_lines) - 1)
    grid_width = line_width_max

    margin_x = 40

    # Scale title if too wide
    title_width = title_img.width
    title_height = title_img.height
    if title_width > grid_width:
        scale_factor = grid_width / title_width
        title_width = grid_width
        title_height = int(title_img.height * scale_factor)
        title_img = title_img.resize((title_width, title_height), Image.LANCZOS)

    composite_width = grid_width + (margin_x * 2)
    title_x_offset = (composite_width - title_width) // 2
    bottom_padding = 40
    composite_height = title_height + spacing + grid_height + bottom_padding

    composite = Image.new('RGBA', (composite_width, composite_height), (18, 18, 20, 255))
    composite.paste(title_img, (title_x_offset, 0), title_img if title_img.mode == 'RGBA' else None)

    # Paste lines centered horizontally
    y_offset = title_height + spacing
    for idx, line in enumerate(rendered_lines):
        line_width = line_widths[idx]
        x_start = margin_x + (grid_width - line_width) // 2 if line_width > 0 else margin_x
        x = x_start
        for box in line:
            composite.paste(box, (x, y_offset), box)
            x += box.width + spacing
        y_offset += line_heights[idx] + spacing

    # Footer
    draw = ImageDraw.Draw(composite)
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((composite_width - text_w) // 2, composite_height - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    composite.save(out, format='PNG')
    out.seek(0)
    return out



def create_streaks_image(ign: str, level: int, icon: str, ign_color: str, guild_tag: str, guild_color: str, winstreak: int, killstreak: int) -> io.BytesIO:
    if Image is None:
        raise RuntimeError("Pillow not available")

    title_io = render_prestige_with_text(level, icon, ign, "", ign_color, guild_tag, guild_color, two_line=False)
    title_img = Image.open(title_io)

    box_width = 300
    box_height = 120
    spacing = 20

    boxes = [
        render_stat_box("Current Winstreak", f"{int(winstreak):,}", width=box_width, height=box_height),
        render_stat_box("Current Killstreak", f"{int(killstreak):,}", width=box_width, height=box_height),
    ]

    line_width = boxes[0].width + boxes[1].width + spacing
    grid_height = box_height
    margin_x = 40
    margin_y = 40

    title_width = title_img.width
    title_height = title_img.height
    # Enforce minimum width and even width for symmetry
    content_width = max(title_width, line_width, 800)

    composite_width = content_width + margin_x * 2
    if composite_width % 2 != 0:
        composite_width += 1
    
    # Adjust bottom margin to match visual top margin of text (title image has ~6px top padding)
    visual_top_margin = margin_y + 6
    margin_bottom = visual_top_margin
    composite_height = title_height + spacing + grid_height + margin_y + margin_bottom

    composite = Image.new('RGBA', (composite_width, composite_height), (18, 18, 20, 255))
    composite.paste(title_img, ((composite_width - title_width) // 2, margin_y), title_img if title_img.mode == 'RGBA' else None)

    y_offset = margin_y + title_height + spacing
    x_start = (composite_width - line_width) // 2
    x = x_start
    for box in boxes:
        composite.paste(box, (x, y_offset), box)
        x += box.width + spacing

    draw = ImageDraw.Draw(composite)
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((composite_width - text_w) // 2, composite_height - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    composite.save(out, format='PNG')
    out.seek(0)
    return out



def create_leaderboard_image(tab_name: str, metric_label: str, leaderboard_data: list, page: int = 0, total_pages: int = 1) -> io.BytesIO:
    # Design constants matching sheepwars command
    canvas_w = 1200
    margin = 40
    spacing = 10
    row_height = 60
    header_height = 80
    
    content_height = header_height + spacing + (len(leaderboard_data) * (row_height + spacing))
    canvas_h = margin + content_height + margin
    
    img = Image.new('RGBA', (canvas_w, canvas_h), (18, 18, 20, 255))
    draw = ImageDraw.Draw(img)
    
    font_header = _load_font("DejaVuSans-Bold.ttf", 32)
    font_rank = _load_font("DejaVuSans-Bold.ttf", 24)
    font_name = _load_font("DejaVuSans-Bold.ttf", 24)
    font_val = _load_font("DejaVuSans-Bold.ttf", 24)
    font_small = _load_font("DejaVuSans-Bold.ttf", 16)
    
    # Header Card
    draw.rounded_rectangle([margin, margin, canvas_w - margin, margin + header_height], radius=15, fill=(35, 30, 45, 240))
    
    title_text = f"{tab_name} {metric_label} Leaderboard"
    page_text = f"Page {page + 1}/{total_pages}"
    
    bbox = draw.textbbox((0, 0), title_text, font=font_header)
    draw.text((margin + (canvas_w - margin*2 - (bbox[2]-bbox[0]))//2, margin + (header_height - (bbox[3]-bbox[1]))//2 - 5), title_text, font=font_header, fill=(255, 255, 255))
    
    bbox_p = draw.textbbox((0, 0), page_text, font=font_small)
    draw.text((canvas_w - margin - (bbox_p[2]-bbox_p[0]) - 20, margin + (header_height - (bbox_p[3]-bbox_p[1]))//2), page_text, font=font_small, fill=(180, 180, 200))

    y = margin + header_height + spacing
    
    for entry in leaderboard_data:
        rank, player, level, icon, p_hex, g_tag, g_hex, value, is_playtime = entry
        player = str(player)
        
        # Row Card
        draw.rounded_rectangle([margin, y, canvas_w - margin, y + row_height], radius=15, fill=(35, 30, 45, 240))
        
        # Rank color
        r_col = (180, 180, 200)
        if rank == 1: r_col = (255, 215, 0)
        elif rank == 2: r_col = (192, 192, 192)
        elif rank == 3: r_col = (205, 127, 50)
        
        draw.text((margin + 20, y + 15), f"#{rank}", font=font_rank, fill=r_col)
        
        # Prestige
        rank_w = draw.textbbox((0,0), f"#{rank}", font=font_rank)[2] - draw.textbbox((0,0), f"#{rank}", font=font_rank)[0]
        p_x = margin + 20 + rank_w + 15
        
        segments = get_prestige_segments(level, icon)
        current_x = p_x
        for hex_color, text in segments:
            try:
                rgb = tuple(int(hex_color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
            except:
                rgb = (255, 255, 255)
            draw.text((current_x, y + 15), text, font=font_name, fill=rgb)
            seg_w = draw.textbbox((0,0), text, font=font_name)[2] - draw.textbbox((0,0), text, font=font_name)[0]
            current_x += seg_w
        
        # Name
        n_x = current_x + 10
        try:
            p_rgb = tuple(int(str(p_hex).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            p_rgb = (255, 255, 255)
        draw.text((n_x, y + 15), player, font=font_name, fill=p_rgb)
        
        # Guild
        if g_tag:
            n_w = draw.textbbox((0,0), player, font=font_name)[2] - draw.textbbox((0,0), player, font=font_name)[0]
            g_x = n_x + n_w + 10
            try:
                g_rgb = tuple(int(str(g_hex).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
            except:
                g_rgb = (170, 170, 170)
            draw.text((g_x, y + 15), f"[{str(g_tag)}]", font=font_name, fill=g_rgb)
        
        # Value
        if is_playtime:
            val_str = format_playtime(int(value))
        elif "Ratio" in metric_label or "/" in metric_label or "Per" in metric_label:
            val_str = f"{float(value):,.2f}"
        else:
            val_str = f"{int(value):,}"
        v_w = draw.textbbox((0,0), val_str, font=font_val)[2] - draw.textbbox((0,0), val_str, font=font_val)[0]
        draw.text((canvas_w - margin - 20 - v_w, y + 15), val_str, font=font_val, fill=(85, 255, 255))
        
        y += row_height + spacing

    # Footer
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((canvas_w - text_w) // 2, canvas_h - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    img.save(out, format='PNG')
    out.seek(0)
    return out


def create_distribution_pie(title: str, slices: list) -> io.BytesIO:
    """Render a pie chart with a subtle 3D tilt and legend."""
    if Image is None:
        raise RuntimeError("Pillow not available")

    total = sum(v for _, v, _ in slices) # Calculate total value for percentages
    if total <= 0:
        total = 1

    width, height = 1032, 672
    padding = 45
    legend_height = 220
    pie_top = 85
    depth = 45  # vertical extrusion to fake 3D
    usable_height = height - legend_height - padding - pie_top
    pie_height = max(160, usable_height - depth)

    img = Image.new("RGBA", (width, height), (18, 18, 20, 255))
    draw = ImageDraw.Draw(img)

    try:
        title_font = _load_font("DejaVuSans-Bold.ttf", 26)
        legend_font = _load_font("DejaVuSans.ttf", 17)
    except Exception:
        title_font = ImageFont.load_default()
        legend_font = ImageFont.load_default()

    bbox = draw.textbbox((0, 0), title, font=title_font)
    title_width = bbox[2] - bbox[0]
    draw.text(((width - title_width) // 2, 20), title, font=title_font, fill=(230, 230, 230))

    top_bbox = (padding, pie_top, width - padding, pie_top + pie_height)
    outline_dark = (18, 18, 24)

    def _shade(color, factor: float):
        return tuple(max(0, min(255, int(channel * factor))) for channel in color)

    # Precompute slice angles so we can reuse them for the depth and top faces
    slice_angles = []
    start_angle = 90  # start at 90 degrees (middle-right position)
    for _, value, color in slices:
        extent = 360 * (value / total)
        end_angle = start_angle + extent
        if extent > 0:
            slice_angles.append((start_angle, end_angle, color))
        start_angle = end_angle

    # Draw depth layers from back to front, one z-level at a time
    # This ensures all slices are visible at each depth level
    for z in range(depth, -1, -1):  # Include z=0 to eliminate gap
        for start_angle, end_angle, color in slice_angles:
            # Shade the sides to be slightly darker
            side_color = _shade(color, 0.8)
            offset_bbox = (
                top_bbox[0],
                top_bbox[1] + z,
                top_bbox[2],
                top_bbox[3] + z,
            )
            # Use side_color for both fill and outline to eliminate any gaps between layers
            draw.pieslice(offset_bbox, start=start_angle, end=end_angle, fill=side_color, outline=side_color, width=2)

    # Draw separator lines on the top face only (no fill, just outline)
    separator_color = (20, 20, 25)  # Dark separator between slices

    # Draw vertical separators for the visible sides (front face)
    cx = (top_bbox[0] + top_bbox[2]) / 2
    cy = (top_bbox[1] + top_bbox[3]) / 2
    rx = (top_bbox[2] - top_bbox[0]) / 2
    ry = (top_bbox[3] - top_bbox[1]) / 2

    boundaries = set()
    for s, e, _ in slice_angles:
        boundaries.add(s % 360)
        boundaries.add(e % 360)

    for angle in boundaries:
        # Only draw separators on the front face (0 to 180 degrees)
        if 0 <= angle <= 180:
            rad = math.radians(angle)
            x = cx + rx * math.cos(rad)
            y = cy + ry * math.sin(rad)
            draw.line([(x, y), (x, y + depth)], fill=separator_color, width=2)

    for start_angle, end_angle, color in slice_angles:
        draw.pieslice(top_bbox, start=start_angle, end=end_angle, fill=None, outline=separator_color, width=2)

    legend_x = padding + 10
    legend_y = top_bbox[3] + depth + 24
    box_size = 20
    line_spacing = 28
    for idx, (label, value, color) in enumerate(slices):
        percent = (value / total * 100) if total else 0
        y = legend_y + idx * line_spacing
        draw.rectangle([legend_x, y, legend_x + box_size, y + box_size], fill=color, outline=(240, 240, 240))
        text = f"{label}: {value} ({percent:.2f}%)"
        draw.text((legend_x + box_size + 10, y - 2), text, font=legend_font, fill=(220, 220, 220))

    # Footer
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((width - text_w) // 2, height - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def render_prestige_range_image(base: int, end_display: int) -> io.BytesIO:
    """Render an image showing the colored start and end prestige brackets from raw pattern."""
    raw = PRESTIGE_RAW_PATTERNS.get(base)
    if not raw:
        # Fallback to simple text
        parts = [(MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF'), f'[{base}] - [{end_display}]')]
        return _render_text_segments_to_image(parts)

    parts = _parse_raw_pattern(raw)

    def _build_replaced_segments(parts, replacement_str, rainbow=False):
        """Replace the first numeric span in the concatenated parts with replacement_str once, preserving color segments.

        If `rainbow` is True, expand the replacement into per-character colored segments cycling a rainbow palette.
        """
        concat = ''.join(t for (_, t) in parts)
        m = re.search(r"\d+", concat)
        if not m:
            return [(MINECRAFT_CODE_TO_HEX.get(code.lower(), '#FFFFFF'), text) for code, text in parts]

        num_start, num_end = m.start(), m.end()
        out_parts = []
        pos = 0
        replaced = False
        for code, text in parts:
            part_start = pos
            part_end = pos + len(text)
            pos = part_end
            hexcol = MINECRAFT_CODE_TO_HEX.get(code.lower(), '#FFFFFF')

            if part_end <= num_start or part_start >= num_end:
                out_parts.append((hexcol, text))
                continue

            # prefix
            prefix_len = max(0, num_start - part_start)
            if prefix_len > 0:
                prefix = text[:prefix_len]
                out_parts.append((hexcol, prefix))

            # replacement
            if not replaced:
                if rainbow:
                    # Build the original color sequence that covered the numeric span
                    colors_in_span = []
                    span_pos = 0
                    # Re-iterate to collect per-char colors within the numeric span
                    pos2 = 0
                    for c_code, c_text in parts:
                        part_s = pos2
                        part_e = pos2 + len(c_text)
                        pos2 = part_e
                        overlap_s = max(part_s, num_start)
                        overlap_e = min(part_e, num_end)
                        if overlap_e > overlap_s:
                            hex_here = MINECRAFT_CODE_TO_HEX.get(c_code.lower(), '#FFFFFF')
                            # number of covered chars in original
                            count = overlap_e - overlap_s
                            colors_in_span.extend([hex_here] * count)

                    if not colors_in_span:
                        # fallback rainbow cycle
                        RAINBOW_CODES = ['c', '6', 'e', 'a', 'b', 'd', '9', '3']
                        colors_in_span = [MINECRAFT_CODE_TO_HEX.get(code, '#FFFFFF') for code in RAINBOW_CODES]

                    # Apply colors across the replacement string, repeating as needed
                    repl = str(replacement_str)
                    for i, ch in enumerate(repl):
                        col = colors_in_span[i % len(colors_in_span)]
                        out_parts.append((col, ch))
                else:
                    out_parts.append((hexcol, replacement_str))
                replaced = True

            # suffix
            suffix_start_in_part = max(0, num_end - part_start)
            if suffix_start_in_part < len(text):
                suffix = text[suffix_start_in_part:]
                out_parts.append((hexcol, suffix))

        return out_parts

    # Choose fallback icons for bases where emoji fonts may be missing
    BAD_ICON_BASES = {800, 1200, 1800, 2800, 3800}

    # Determine if this prestige base should be rainbow (PRESTIGE_COLORS maps to None)
    rainbow_bases = {k for k, v in PRESTIGE_COLORS.items() if v is None}

    start_segments = _build_replaced_segments(parts, str(base), rainbow=(base in rainbow_bases))
    end_segments = _build_replaced_segments(parts, str(end_display), rainbow=(end_display in rainbow_bases))

    # If problematic base, replace any non-ascii icon with fallback from PRESTIGE_ICONS
    if base in BAD_ICON_BASES or end_display in BAD_ICON_BASES:
        def _replace_bad_icons(segments, base_val):
            res = []
            for col, txt in segments:
                # replace any non-basic symbol at end inside brackets with fallback
                newtxt = re.sub(r"\[(\s*\d+)([^\d\]]+)\]", lambda m: f"[{m.group(1)}{PRESTIGE_ICONS[(base_val//100) % len(PRESTIGE_ICONS)]}]", txt)
                res.append((col, newtxt))
            return res
        start_segments = _replace_bad_icons(start_segments, base)
        end_segments = _replace_bad_icons(end_segments, base)

    combined = []
    combined.extend(start_segments)
    combined.append((MINECRAFT_CODE_TO_HEX.get('7', '#AAAAAA'), ' -> '))
    combined.extend(end_segments)

    return _render_text_segments_to_image(combined)


def render_all_prestiges_combined(spacing: int = 20) -> io.BytesIO:
    """Render all prestiges as individual images and combine them vertically into one PNG."""
    if Image is None:
        raise RuntimeError("Pillow not available")

    # Build a 4-column layout where columns are offsets [0,1000,2000,3000]
    offsets = [0, 1000, 2000, 3000]

    # Rows are the base mods 0,100,...,900 (we limit to prestiges up to 4000)
    base_mods = [i * 100 for i in range(0, 10)]

    # Prepare grid of images (rows x cols). Use placeholder transparent images for missing cells.
    grid = []
    for base_mod in base_mods:
        row_imgs = []
        for off in offsets:
            key = base_mod + off
            if key in PRESTIGE_RAW_PATTERNS:
                try:
                    imgio = render_prestige_range_image(key, key + 99)
                    imgio.seek(0)
                    im = Image.open(imgio).convert('RGBA')
                except Exception:
                    im = Image.new('RGBA', (200, 40), (0,0,0,0))
            else:
                im = Image.new('RGBA', (200, 40), (0,0,0,0))
            row_imgs.append(im)
        grid.append(row_imgs)

    # Compute uniform cell size
    max_w = max((im.width for row in grid for im in row), default=200) + 30
    max_h = max((im.height for row in grid for im in row), default=40) + 20

    # Optional title at the top
    title_text = "Wool Games Prestiges"
    try:
        title_font = _load_font("DejaVuSans-Bold.ttf", 32)
    except Exception:
        title_font = ImageFont.load_default()
    
    draw_dummy = ImageDraw.Draw(Image.new('RGBA', (1,1)))
    tb = draw_dummy.textbbox((0,0), title_text, font=title_font)
    title_h = tb[3] - tb[1] + 40

    cols = len(offsets)
    rows = len(grid)

    margin = 40
    total_w = margin * 2 + cols * max_w + spacing * (cols - 1)
    total_h = margin * 2 + title_h + rows * max_h + spacing * (rows - 1) + 40

    combined = Image.new('RGBA', (total_w, total_h), (18, 18, 20, 255))
    draw = ImageDraw.Draw(combined)

    # Draw title centered
    title_x = total_w // 2
    title_y = margin + (title_h // 2) - 10
    draw.text((title_x, title_y), title_text, font=title_font, fill=(255, 255, 255), anchor='mm')

    start_y = margin + title_h
    for r, row in enumerate(grid):
        y = start_y + r * (max_h + spacing)
        base_mod = base_mods[r]
        for c, im in enumerate(row):
            x = margin + c * (max_w + spacing)
            offset = offsets[c]
            level = base_mod + offset
            
            # Determine background color based on text brightness
            p_color = get_prestige_color(level)
            #lum = (0.299 * p_color[0] + 0.587 * p_color[1] + 0.114 * p_color[2])
            bg_color = (35, 30, 45, 255) #if lum < 90 else (35, 30, 45, 255)

            # Draw card background
            draw.rounded_rectangle([x, y, x + max_w, y + max_h], radius=8, fill=bg_color)
            
            # center each image within its cell
            paste_x = x + (max_w - im.width) // 2
            paste_y = y + (max_h - im.height) // 2
            combined.paste(im, (paste_x, paste_y), im)

    # Footer
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((total_w - text_w) // 2, total_h - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    combined.save(out, format='PNG')
    out.seek(0)
    return out



def format_prestige_ansi(level: int, icon: str) -> str:
    """Return an ANSI-colored prestige bracket+level+icon string.

    If a multi-color pattern exists for the prestige base (e.g. 1900), use it;
    otherwise color the whole bracket using the single prestige color.
    """
    reset = "\u001b[0m"
    try:
        lvl = int(level)
    except Exception:
        lvl = 0
    base = (lvl // 100) * 100
    # If a raw pattern exists, parse it into (code, text) pieces
    if base in PRESTIGE_RAW_PATTERNS:
        raw = PRESTIGE_RAW_PATTERNS[base]
        parts = []
        cur_code = None
        buf = ''
        i = 0
        while i < len(raw):
            ch = raw[i]
            if ch == '&' and i + 1 < len(raw):
                # flush buf
                if buf:
                    parts.append((cur_code or 'f', buf))
                    buf = ''
                cur_code = raw[i+1].lower()
                i += 2
                continue
            else:
                buf += ch
                i += 1
        if buf:
            parts.append((cur_code or 'f', buf))

        out = []
        for code, text in parts:
            # Use chosen xterm-256 SGR for inline/code-block rendering
            sgr = MINECRAFT_CODE_TO_ANSI_SGR.get(code.lower(), "\u001b[37m")
            out.append(make_bold_ansi(sgr) + text)

        joined = ''.join(out) + reset
        # When a raw pattern exists we trust it includes the correct icon and colors.
        return joined

    # Fallback: color whole bracket with single color for the level
    ansi = get_ansi_color_code(level)
    bold = make_bold_ansi(ansi)
    return f"{bold}[{level}{icon}]{reset}"


async def _send_paged_ansi_followups(interaction: discord.Interaction, lines: list[str], block: str = 'ansi'):
    """Send potentially-long ANSI lines as one or more followup messages, each <= 2000 chars.

    Splits `lines` into code-block chunks and sends them via `interaction.followup.send`.
    Falls back to sanitized plain text if sending fails.
    """
    wrapper_open = f"```{block}\n"
    wrapper_close = "\n```"
    max_len = 2000

    chunks = []
    cur_lines = []
    # start with the wrapper overhead
    cur_len = len(wrapper_open) + len(wrapper_close)

    for ln in lines:
        ln_with_nl = ln + "\n"
        lnlen = len(ln_with_nl)
        if cur_len + lnlen > max_len:
            # flush current chunk
            if cur_lines:
                chunks.append("".join(cur_lines).rstrip('\n'))
            cur_lines = [ln_with_nl]
            cur_len = len(wrapper_open) + len(wrapper_close) + lnlen
        else:
            cur_lines.append(ln_with_nl)
            cur_len += lnlen

    if cur_lines:
        chunks.append("".join(cur_lines).rstrip('\n'))

    # Send chunks as followups
    for chunk in chunks:
        content = wrapper_open + chunk + wrapper_close
        try:
            await interaction.followup.send(content)
        except Exception:
            # fallback: send sanitized text without ANSI wrapper
            try:
                await interaction.followup.send(sanitize_output(chunk))
            except Exception:
                # give up silently
                pass

def load_tracked_users():
    return get_tracked_users()


def load_tracked_streaks() -> dict:
    try:
        return get_all_tracked_streaks()
    except Exception:
        pass
    return {}


def save_tracked_streaks(data: dict):
    try:
        for username, streak_data in data.items():
            update_tracked_streaks(username, streak_data)
    except Exception as e:
        print(f"[STREAK] Failed to save streaks to database: {e}")


def load_user_colors() -> dict:
    """Load user colors and metadata from database"""
    try:
        result = {}
        usernames = get_all_usernames()
        for username in usernames:
            meta = get_user_meta(username)
            if meta:
                # Convert database format to expected format
                result[username.lower()] = {
                    'color': meta.get('ign_color'),
                    'guild_tag': meta.get('guild_tag'),
                    'guild_color': meta.get('guild_hex')
                }
        return result
    except Exception:
        pass
    return {}


def _get_lifetime_value(stats: dict, key: str) -> int:
    try:
        return int(stats.get(key, {}).get("lifetime", 0))
    except Exception:
        return 0


def update_streaks_from_stats(username: str, processed_stats: dict) -> bool:
    streaks = load_tracked_streaks()
    entry = streaks.get(username)
    if not entry:
        return False

    wins = _get_lifetime_value(processed_stats, "wins")
    losses = _get_lifetime_value(processed_stats, "losses")
    kills = _get_lifetime_value(processed_stats, "kills")
    deaths = _get_lifetime_value(processed_stats, "deaths")

    last_wins = int(entry.get("last_wins", wins))
    last_losses = int(entry.get("last_losses", losses))
    last_kills = int(entry.get("last_kills", kills))
    last_deaths = int(entry.get("last_deaths", deaths))

    winstreak = int(entry.get("winstreak", 0))
    killstreak = int(entry.get("killstreak", 0))

    win_delta = wins - last_wins
    loss_delta = losses - last_losses
    kill_delta = kills - last_kills
    death_delta = deaths - last_deaths

    if loss_delta > 0:
        winstreak = 0
    elif win_delta > 0 and loss_delta <= 0:
        winstreak = max(0, winstreak) + win_delta

    if death_delta > 0:
        killstreak = 0
    elif kill_delta > 0 and death_delta <= 0:
        killstreak = max(0, killstreak) + kill_delta

    entry.update({
        "winstreak": winstreak,
        "killstreak": killstreak,
        "last_wins": wins,
        "last_losses": losses,
        "last_kills": kills,
        "last_deaths": deaths,
    })
    streaks[username] = entry
    save_tracked_streaks(streaks)
    return True


def initialize_streak_entry(username: str, processed_stats: dict):
    streaks = load_tracked_streaks()
    wins = _get_lifetime_value(processed_stats, "wins")
    losses = _get_lifetime_value(processed_stats, "losses")
    kills = _get_lifetime_value(processed_stats, "kills")
    deaths = _get_lifetime_value(processed_stats, "deaths")

    streaks[username] = {
        "winstreak": 0,
        "killstreak": 0,
        "last_wins": wins,
        "last_losses": losses,
        "last_kills": kills,
        "last_deaths": deaths,
    }
    save_tracked_streaks(streaks)

def load_user_links():
    """Load username -> Discord user ID mappings from database"""
    try:
        return get_all_user_links()
    except Exception:
        return {}

def save_user_links(links: dict):
    """Save username -> Discord user ID mappings to database"""
    for username, discord_id in links.items():
        set_discord_link(username, discord_id)

def link_user_to_ign(discord_user_id: int, ign: str):
    """Link a Discord user ID to a Minecraft username (case-insensitive)"""
    links = load_user_links()
    # Store with original case but search case-insensitively
    links[ign.casefold()] = str(discord_user_id)
    save_user_links(links)

def is_user_authorized(discord_user_id: int, ign: str) -> bool:
    """Check if a Discord user is authorized to manage a username"""
    links = load_user_links()
    key = ign.casefold()
    return links.get(key) == str(discord_user_id)

def is_admin(user: Union[discord.User, discord.Member]) -> bool:
    """Check if user is a bot admin."""
    if str(user.id) in ADMIN_IDS:
        return True
    if user.name.casefold() in [n.casefold() for n in ADMIN_NAMES]:
        return True
    return False

def unlink_user_from_ign(ign: str) -> bool:
    """Remove username -> Discord user ID link"""
    links = load_user_links()
    key = ign.casefold()
    if key in links:
        del links[key]
        save_user_links(links)
        return True
    return False

def remove_user_color(ign: str) -> bool:
    """Remove username color from database"""
    try:
        meta = get_user_meta(ign)
        if meta and meta.get('ign_color'):
            update_user_meta(ign, ign_color="")
            return True
        return False
    except Exception as e:
        print(f"[ERROR] Failed to remove user color for {ign}: {e}")
        return False

def delete_user_sheet(ign: str) -> bool:
    """Delete user's data from database."""
    try:
        from db_helper import delete_user, user_exists
        
        if not user_exists(ign):
            print(f"[INFO] User {ign} not found in database")
            return False
        
        delete_user(ign)
        print(f"[INFO] Deleted user {ign} from database")
        return True
        
    except Exception as e:
        print(f"[ERROR] Failed to delete user {ign}: {e}")
        return False

def render_modern_card(label, value, width, height, color=(255, 255, 255), is_header=False):
    img = Image.new('RGBA', (int(width), int(height)), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    card_bg = (35, 30, 45, 240) 
    draw.rounded_rectangle([0, 0, width-1, height-1], radius=15, fill=card_bg)
    font_label = _load_font("DejaVuSans-Bold.ttf", 14)
    font_value = _load_font("DejaVuSans-Bold.ttf", 28 if is_header else 24)
    l_text = f"{label.upper()}:"
    l_bbox = draw.textbbox((0, 0), l_text, font=font_label)
    draw.text(((width - (l_bbox[2]-l_bbox[0])) // 2, height * 0.2), l_text, font=font_label, fill=(180, 180, 200))
    v_text = str(value)
    draw.text((width // 2, int(height * 0.6)), v_text, font=font_value, fill=color, anchor="mm")
    return img

_UUID_CACHE = {}

def get_uuid(ign: str) -> Optional[str]:
    ign_lower = ign.lower()
    if ign_lower in _UUID_CACHE:
        return _UUID_CACHE[ign_lower]
    
    headers = {"User-Agent": "SheepWarsBot/1.0"}
    
    # Try Mojang
    try:
        r = requests.get(f"https://api.mojang.com/users/profiles/minecraft/{ign}", headers=headers, timeout=2)
        if r.status_code == 200:
            data = r.json()
            uuid = data.get('id')
            if uuid:
                _UUID_CACHE[ign_lower] = uuid
                return uuid
    except:
        pass
        
    # Try PlayerDB
    try:
        r = requests.get(f"https://playerdb.co/api/player/minecraft/{ign}", headers=headers, timeout=2)
        if r.status_code == 200:
            data = r.json()
            if data.get('success'):
                uuid = data.get('data', {}).get('player', {}).get('raw_id')
                if uuid:
                    _UUID_CACHE[ign_lower] = uuid
                    return uuid
    except:
        pass
        
    return None

def get_player_body(ign):
    # Resolve UUID for better API support
    uuid = get_uuid(ign)
    identifier = uuid if uuid else ign

    # Try multiple providers to find one that works/updates
    # Using random param to bypass edge caching where possible
    ts = random.randint(0, 10000)
    providers = [
        f"https://api.mineatar.io/body/full/{identifier}?scale=10&ts={ts}"
    ]
    for url in providers:
        try:
            r = requests.get(url, headers={"User-Agent": "SheepWarsBot/1.0"}, timeout=5)
            if r.status_code == 200:
                return Image.open(io.BytesIO(r.content)).convert("RGBA")
        except Exception:
            continue
    return None

def get_api_key():
    try:
        with open(os.path.join(BOT_DIR, "API_KEY.txt"), "r") as f:
            return f.read().strip()
    except:
        return None

def verify_api_key():
    """Verify Hypixel API key validity on startup."""
    key = get_api_key()
    if not key:
        print("[STARTUP] [ERROR] API_KEY.txt not found or empty!")
        return

    print("[STARTUP] Verifying Hypixel API key...")
    try:
        r = requests.get("https://api.hypixel.net/key", headers={"API-Key": key}, timeout=10)
        if r.status_code == 200:
            data = r.json()
            if data.get('success'):
                record = data.get('record', {})
                owner = record.get('owner')
                limit = record.get('limit')
                print(f"[STARTUP] [OK] API Key verified! Owner UUID: {owner}, Limit: {limit}")
                return
        
        print(f"[STARTUP] [WARNING] API Key verification failed. Status: {r.status_code}")
    except Exception as e:
        print(f"[STARTUP] [ERROR] Failed to connect to Hypixel API: {e}")

async def check_legacy_migration():
    """Check if legacy files exist and database is empty, then migrate."""
    # Check if we need to migrate stats.xlsx
    db_stats = get_database_stats()
    if db_stats['users'] == 0:
        excel_file = BOT_DIR / "stats.xlsx"
        if excel_file.exists():
            print("[MIGRATION] Database empty but stats.xlsx found. Running conversion...")
            try:
                # Run convert_to_db.py with --force to skip prompt
                subprocess.run([sys.executable, "convert_to_db.py", "--force"], cwd=str(BOT_DIR), check=True)
                print("[MIGRATION] Conversion script finished.")
                # Force cache refresh after migration
                await STATS_CACHE.refresh()
            except Exception as e:
                print(f"[ERROR] Migration failed: {e}")

    # Check for JSON files migration (user_links, default_users, tracked_streaks)
    json_files = ["user_links.json", "default_users.json", "tracked_streaks.json"]
    should_migrate_json = False
    
    if (BOT_DIR / "user_links.json").exists() and not get_all_user_links():
        should_migrate_json = True
    elif (BOT_DIR / "default_users.json").exists() and not get_all_default_users():
        should_migrate_json = True
    elif (BOT_DIR / "tracked_streaks.json").exists() and not get_all_tracked_streaks():
        should_migrate_json = True
        
    if should_migrate_json:
        print("[MIGRATION] Legacy JSON files found and tables empty. Running conversion...")
        try:
            subprocess.run([sys.executable, "convert_to_db.py", "--force"], cwd=str(BOT_DIR), check=True)
            print("[MIGRATION] JSON conversion finished.")
        except Exception as e:
            print(f"[ERROR] JSON migration failed: {e}")

    # Check if we need to migrate tracked_users.txt (if not done by convert_to_db)
    tracked_users = get_tracked_users()
    if not tracked_users:
        tracked_file = BOT_DIR / "tracked_users.txt"
        if tracked_file.exists():
            print("[MIGRATION] tracked_users table empty but tracked_users.txt found. Migrating...")
            try:
                # We can re-run convert_to_db.py as it handles tracked_users now, or do it manually.
                # Running convert_to_db.py is safer as it centralizes logic.
                subprocess.run([sys.executable, "convert_to_db.py", "--force"], cwd=str(BOT_DIR), check=True)
            except Exception as e:
                print(f"[ERROR] Tracked users migration failed: {e}")

def get_player_status(ign):
    """Fetch player online status from Hypixel API."""
    api_key = get_api_key()
    if not api_key:
        return "Unknown", (170, 170, 170) # Gray
    
    # Get UUID
    uuid = get_uuid(ign)
    if not uuid:
        return "Unknown", (170, 170, 170)

    try:
        headers = {"API-Key": api_key, "User-Agent": "SheepWarsBot/1.0"}
        r = requests.get("https://api.hypixel.net/status", params={"uuid": uuid}, headers=headers, timeout=3)
        if r.status_code == 200:
            data = r.json()
            session = data.get('session')
            if data.get('success') and session and isinstance(session, dict) and session.get('online'):
                return "Online", (85, 255, 85) # Green
            else:
                return "Offline", (255, 85, 85) # Red
        else:
            print(f"[WARNING] Hypixel status check failed: {r.status_code}")
    except Exception as e:
        print(f"[WARNING] Status check error: {e}")
    
    return "Unknown", (170, 170, 170)

# ---- Default IGN helpers ----
def load_default_users() -> dict:
    try:
        return get_all_default_users()
    except Exception:
        return {}

def save_default_users(defaults: dict):
    for discord_id, username in defaults.items():
        set_default_username(discord_id, username)

def set_default_user(discord_user_id: int, ign: str):
    defaults = load_default_users()
    defaults[str(discord_user_id)] = ign
    save_default_users(defaults)

def remove_default_user(discord_user_id: int) -> bool:
    defaults = load_default_users()
    key = str(discord_user_id)
    if key in defaults:
        del defaults[key]
        save_default_users(defaults)
        return True
    return False

def get_default_user(discord_user_id: int) -> Optional[str]:
    defaults = load_default_users()
    return defaults.get(str(discord_user_id))

async def cleanup_untracked_user_delayed(ign: str, delay_seconds: int = 60):
    """Schedule cleanup of untracked user data after a delay.
    
    Waits for delay_seconds, then checks if the user is still untracked.
    If they're still untracked, removes their color data and sheet.
    """
    try:
        print(f"[CLEANUP] Scheduled cleanup for '{ign}' in {delay_seconds} seconds")
        await asyncio.sleep(delay_seconds)
        
        # Check if user is now tracked
        tracked_users = load_tracked_users()
        print(f"[CLEANUP] After {delay_seconds}s delay, checking if '{ign}' is tracked")
        print(f"[CLEANUP] Tracked users list: {tracked_users}")
        
        key = ign.casefold()
        for tracked_user in tracked_users:
            if tracked_user.casefold() == key:
                # User is now tracked, don't clean up
                print(f"[CLEANUP] SKIPPING cleanup for '{ign}' - found in tracked users database as '{tracked_user}'")
                return
        
        # User is still untracked, proceed with cleanup
        print(f"[CLEANUP] User '{ign}' NOT FOUND in tracked users database")
        print(f"[CLEANUP] Reason: User was queried via /sheepwars but is not in tracked list")
        print(f"[CLEANUP] Proceeding with cleanup: removing color data and deleting sheet")
        
        color_removed = remove_user_color(ign)
        sheet_deleted = delete_user_sheet(ign)
        
        print(f"[CLEANUP] Cleanup complete for '{ign}' - color_removed={color_removed}, sheet_deleted={sheet_deleted}")
    except asyncio.CancelledError:
        print(f"[CLEANUP] Cleanup task cancelled for '{ign}'")
    except Exception as e:
        print(f"[CLEANUP] ERROR during cleanup for '{ign}': {e}")
        import traceback
        traceback.print_exc()

async def send_fetch_message(message: str):
    # DM the creator (prefer explicit ID if set)
    user = None
    if CREATOR_ID is not None:
        try:
            uid = int(CREATOR_ID)
            user = bot.get_user(uid) or await bot.fetch_user(uid)
        except Exception:
            user = None
    if user is None:
        # fallback to name/display name search across guilds
        for guild in bot.guilds:
            for member in guild.members:
                if member.bot:
                    continue
                name_match = member.name.casefold() == CREATOR_NAME.casefold()
                display_match = member.display_name.casefold() == CREATOR_NAME.casefold()
                if name_match or display_match:
                    user = member
                    break
            if user:
                break
    if user:
        try:
            await user.send(message)
            return
        except Exception as e:
            # Common cause: user has DMs disabled (Discord error 50007). Fall back to channel.
            print(f"[WARNING] Could not DM creator: {e}")
    # fallback: send to system channel or first writable channel
    for guild in bot.guilds:
        channel = None
        if guild.system_channel and guild.system_channel.permissions_for(guild.me).send_messages:
            channel = guild.system_channel
        else:
            for ch in guild.text_channels:
                if ch.permissions_for(guild.me).send_messages:
                    channel = ch
                    break
        if channel:
            try:
                await channel.send(message)
                break
            except Exception:
                continue

async def _delayed_refresh_user(username: str, delay: float):
    """Sleep for `delay` seconds then run api_get.py for the given username."""
    try:
        await asyncio.sleep(delay)
        result = await asyncio.to_thread(run_script, "api_get.py", ["-ign", username])

        # Try to update cache/streaks from stdout JSON
        if result and result.stdout:
            try:
                json_data = None
                for line in reversed(result.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            json_data = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                if json_data and "processed_stats" in json_data and "username" in json_data:
                    await STATS_CACHE.update_cache_entry(json_data["username"], json_data["processed_stats"])
            except Exception as parse_err:
                print(f"[REFRESH] Failed to parse refresh output for {username}: {parse_err}")
    except asyncio.CancelledError:
        return
    except Exception as e:
        print(f"[REFRESH] Error refreshing {username}: {e}")


async def staggered_stats_refresher(interval_minutes: int = 10):
    """Background task that refreshes every tracked user's stats once per `interval_minutes`.

    Each user's refresh is scheduled at a random point during the interval to spread load.
    """
    interval = interval_minutes * 60
    buffer = 5  # seconds buffer to avoid scheduling at the very end
    while True:
        try:
            users = load_tracked_users()
            if not users:
                await asyncio.sleep(interval)
                continue

            # assign a random delay in [0, interval-buffer) to each user, then schedule
            tasks = []
            for u in users:
                d = random.uniform(0, max(0, interval - buffer))
                tasks.append(asyncio.create_task(_delayed_refresh_user(u, d)))

            # wait for the interval to elapse; leave any straggling tasks to finish in background
            await asyncio.sleep(interval)

            # optionally gather any finished tasks and suppress exceptions
            for t in tasks:
                if t.done():
                    try:
                        t.result()
                    except Exception:
                        pass

        except Exception as e:
            print(f"[REFRESH] Staggered refresher error: {e}")
            await asyncio.sleep(interval)


# Track last known player count for Sheep Wars to calculate delta
_sheep_wars_last_players = None


async def _get_wool_games_status() -> str | None:
    """Fetch Sheep/Wool Wars player status via HyTrack's socket.io feed.

    Returns a string like "Players: 12 (+1)" or None on failure/timeout.
    """
    global _sheep_wars_last_players
    
    try:
        import socketio  # python-socketio
    except Exception as e:
        print(f"[PRESENCE] socketio import failed: {e}")
        return None

    target_key = "WOOL_GAMES__sheep_wars_two_six"
    status_box = {"value": None}
    status_event = asyncio.Event()

    def _set_status(entry):
        global _sheep_wars_last_players
        if not entry:
            return
        if isinstance(entry, list):
            for item in entry:
                _set_status(item)
            return
        if not isinstance(entry, dict):
            return
        info = entry.get("info", {})
        if info.get("key") != target_key:
            return
        players = entry.get("players")
        if players is None:
            return
        
        # Calculate delta by comparing to last known value
        if _sheep_wars_last_players is not None:
            delta = players - _sheep_wars_last_players
        else:
            delta = 0
        
        # Update last known value
        _sheep_wars_last_players = players
        
        status_box["value"] = f"Sheepers: {players} ({delta:+d})"
        status_event.set()

    sio = socketio.AsyncClient(reconnection=False, logger=False, engineio_logger=False)

    @sio.event
    async def connect():
        try:
            await sio.emit("requestListing", "WOOL_GAMES")
        except Exception as e:
            print(f"[PRESENCE] emit requestListing failed: {e}")

    @sio.on("add")
    async def on_add(entries):
        try:
            _set_status(entries)
        except Exception as e:
            pass

    @sio.on("update")
    async def on_update(update):
        try:
            _set_status(update)
        except Exception as e:
            pass

    try:
        await sio.connect("https://hytrack.me", transports=["websocket", "polling"], wait_timeout=5)
        try:
            await asyncio.wait_for(status_event.wait(), timeout=8)
        except asyncio.TimeoutError:
            pass
    except Exception as e:
        print(f"[PRESENCE] socket connect failed: {e}")
    finally:
        try:
            await sio.disconnect()
        except Exception:
            pass

    return status_box["value"]


async def presence_updater_loop(interval_seconds: int = 5):
    """Background loop: poll site and update bot presence to show current players."""
    last = None
    while True:
        try:
            status = await _get_wool_games_status()
            if status:
                # If status changed, update presence
                if status != last:
                    try:
                        await bot.change_presence(activity=discord.Game(name=status))
                        #print(f"[PRESENCE] Updated presence to: {status}")
                        last = status
                    except Exception as e:
                        print(f"[PRESENCE] Failed to change presence: {e}")
            else:
                # If no status, optionally clear presence
                pass
        except Exception as e:
            print(f"[PRESENCE] Loop error: {e}")
        await asyncio.sleep(interval_seconds)


def inline_backup_fallback():
    """Inline backup fallback when backup_hourly.py script fails."""
    import shutil
    from datetime import datetime
    from db_helper import backup_database
    
    try:
        db_file = DB_FILE
        backup_dir = BOT_DIR / "backups"
        
        # Try primary backup directory
        if not backup_dir.exists():
            try:
                backup_dir.mkdir(exist_ok=True, mode=0o755)
            except:
                # Fallback to home directory
                from pathlib import Path
                backup_dir = Path.home() / "backup_api_backups"
                backup_dir.mkdir(exist_ok=True, mode=0o755)
                print(f"[FALLBACK] Using alternate directory: {backup_dir}")
        
        if not db_file.exists():
            print(f"[FALLBACK] Database file not found: {db_file}")
            return False
        
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-00-00")
        backup_path = backup_dir / f"stats_{timestamp}.db"
        
        if backup_path.exists():
            print(f"[FALLBACK] Backup already exists: {backup_path.name}")
            return True
        
        # Use database helper to backup
        success = backup_database(backup_path)
        
        if success and backup_path.exists():
            size = backup_path.stat().st_size
            print(f"[FALLBACK] Backup created: {backup_path.name} ({size:,} bytes)")
            return True
        else:
            print(f"[FALLBACK] Backup was not created")
            return False
            
    except Exception as e:
        print(f"[FALLBACK] Inline backup error: {e}")
        import traceback
        traceback.print_exc()
        return False


async def scheduler_loop():
    """Automatic scheduler for daily and monthly snapshots, plus hourly backups"""
    last_snapshot_run = None
    last_backup_hour = None
    
    while True:
        now = datetime.datetime.now(tz=CREATOR_TZ)
        
        # Hourly backup - runs at :00 minutes every hour
        if now.minute == 0:
            current_hour = (now.date(), now.hour)
            if last_backup_hour != current_hour:
                try:
                    print(f"[SCHEDULER] Running hourly backup at {now.strftime('%I:%M %p')}")
                    
                    def run_hourly_backup():
                        import subprocess
                        script_path = BOT_DIR / "backup_hourly.py"
                        # Ensure script is executable on Linux
                        if not script_path.exists():
                            raise FileNotFoundError(f"Backup script not found: {script_path}")
                        
                        print(f"[SCHEDULER] Backup script path: {script_path}")
                        print(f"[SCHEDULER] Python executable: {sys.executable}")
                        print(f"[SCHEDULER] Working directory: {BOT_DIR}")
                        
                        return subprocess.run(
                            [sys.executable, str(script_path)],
                            cwd=str(BOT_DIR),
                            capture_output=True,
                            text=True,
                            timeout=120
                        )
                    
                    backup_result = await asyncio.to_thread(run_hourly_backup)
                    if backup_result.returncode == 0:
                        print(f"[SCHEDULER] Hourly backup completed successfully")
                        # Show output even on success for debugging
                        if backup_result.stdout:
                            print(f"[SCHEDULER] Backup output:\n{backup_result.stdout[:500]}")
                    else:
                        print(f"[SCHEDULER] Hourly backup failed with exit code {backup_result.returncode}")
                        if backup_result.stdout:
                            print(f"[SCHEDULER] Backup stdout:\n{backup_result.stdout[:500]}")
                        if backup_result.stderr:
                            print(f"[SCHEDULER] Backup stderr:\n{backup_result.stderr[:500]}")
                        
                        # FALLBACK: Try inline backup
                        print(f"[FALLBACK] Attempting inline backup...")
                        try:
                            await asyncio.to_thread(inline_backup_fallback)
                            print(f"[FALLBACK] Inline backup completed")
                        except Exception as fallback_error:
                            print(f"[FALLBACK] Inline backup also failed: {fallback_error}")
                except Exception as e:
                    print(f"[SCHEDULER] Hourly backup error: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    
                    # FALLBACK: Try inline backup
                    print(f"[FALLBACK] Attempting inline backup after exception...")
                    try:
                        await asyncio.to_thread(inline_backup_fallback)
                        print(f"[FALLBACK] Inline backup completed")
                    except Exception as fallback_error:
                        print(f"[FALLBACK] Inline backup also failed: {fallback_error}")
                
                last_backup_hour = current_hour
        
        # Run snapshot updates at 9:30 AM
        if now.hour == 9 and now.minute == 30:
            today = now.date()
            if last_snapshot_run != today:
                try:
                    # Step 1: Run yesterday snapshot (before daily overwrites it)
                    def run_yesterday():
                        return run_script_batch("batch_update.py", ["-schedule", "yesterday"])
                    
                    yesterday_result = await asyncio.to_thread(run_yesterday)
                    if yesterday_result.returncode != 0:
                        error_detail = yesterday_result.stderr or yesterday_result.stdout or "Unknown error"
                        print(f"[SCHEDULER] Yesterday snapshot FAILED - returncode: {yesterday_result.returncode}")
                        print(f"[SCHEDULER] Full stdout:\n{yesterday_result.stdout}")
                        print(f"[SCHEDULER] Full stderr:\n{yesterday_result.stderr}")
                        await send_fetch_message(f"Warning: Yesterday snapshot failed at {now.strftime('%I:%M %p')}\nError: {error_detail[:500]}")
                    
                    # Step 2: Determine which snapshots to take
                    # Daily: always
                    # Monthly: only on 1st of month
                    if now.day == 1:
                        schedule = "all"  # daily + monthly
                    else:
                        schedule = "daily"
                    
                    # Step 3: Run batch_update.py for daily (and monthly if 1st)
                    def run_batch():
                        return run_script_batch("batch_update.py", ["-schedule", schedule])
                    
                    result = await asyncio.to_thread(run_batch)
                    if result.returncode == 0:
                        msg = f"Daily snapshot completed at {now.strftime('%I:%M %p')}"
                        if now.day == 1:
                            msg += " (including monthly snapshots)"
                        await send_fetch_message(msg)
                    else:
                        error_msg = result.stderr or result.stdout or "Unknown error"
                        print(f"[SCHEDULER] Daily snapshot FAILED - returncode: {result.returncode}")
                        print(f"[SCHEDULER] Full stdout:\n{result.stdout}")
                        print(f"[SCHEDULER] Full stderr:\n{result.stderr}")
                        await send_fetch_message(f"Daily snapshot failed: {error_msg[:500]}")
                except Exception as e:
                    print(f"[SCHEDULER] Snapshot update exception: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    await send_fetch_message(f"Snapshot update error: {str(e)}")
                
                last_snapshot_run = today
        
        await asyncio.sleep(20)

# Helper class for stats tab view
class StatsTabView(discord.ui.View):
    def __init__(self, data_dict, ign, level_value: int, prestige_icon: str, 
                 ign_color: str = None, guild_tag: str = None, guild_hex: str = None,
                 status_text="Online", status_color=(85, 255, 85), skin_image=None):
        super().__init__()
        self.data = data_dict 
        self.ign = ign
        self.level_value = level_value
        self.prestige_icon = prestige_icon
        self.status_text = status_text
        self.status_color = status_color
        self.skin_image = skin_image
        self.current_tab = "all-time"
        
        self.ign_color = ign_color
        self.guild_tag = guild_tag
        self.guild_hex = guild_hex
        
        if self.ign_color is None or self.guild_tag is None:
            self._load_color()
            
        self.update_button_styles()

    def _load_color(self):
        try:
            meta = get_user_meta(self.ign)
            if meta:
                self.ign_color = meta.get('ign_color')
                self.guild_tag = meta.get('guild_tag')
                self.guild_hex = meta.get('guild_hex') or "#AAAAAA"
        except: pass

    def update_button_styles(self):
        """Setzt den aktiven Button auf Blau (Primary) und andere auf Grau (Secondary)."""
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                if child.custom_id == self.current_tab:
                    child.style = discord.ButtonStyle.primary
                else:
                    child.style = discord.ButtonStyle.secondary

    def generate_composite_image(self, tab_name):
        tab_data = self.data[tab_name]
        
        # Rendert das Bild mit den gespeicherten Daten
        img_io = create_stats_composite_image(
            self.level_value, self.prestige_icon, self.ign, tab_name,
            tab_data['wins'], tab_data['losses'], tab_data['wlr'], 
            tab_data['kills'], tab_data['deaths'], tab_data['kdr'],
            self.ign_color, self.guild_tag, self.guild_hex, 
            playtime_seconds=tab_data['playtime'],
            status_text=self.status_text, status_color=self.status_color,
            skin_image=self.skin_image
        )
        return discord.File(img_io, filename=f"{self.ign}_{tab_name}.png")

    async def handle_tab_click(self, interaction: discord.Interaction, tab_name: str):
        self.current_tab = tab_name
        self.update_button_styles()
        file = self.generate_composite_image(tab_name)
        # Wichtig: View=self mitgeben, damit die Styles aktualisiert werden
        await interaction.response.edit_message(attachments=[file], view=self)

    @discord.ui.button(label="All-time", custom_id="all-time")
    async def all_time(self, interaction, button):
        await self.handle_tab_click(interaction, "all-time")

    @discord.ui.button(label="Session", custom_id="session")
    async def session(self, interaction, button):
        await self.handle_tab_click(interaction, "session")

    @discord.ui.button(label="Daily", custom_id="daily")
    async def daily(self, interaction, button):
        await self.handle_tab_click(interaction, "daily")

    @discord.ui.button(label="Yesterday", custom_id="yesterday")
    async def yesterday(self, interaction, button):
        await self.handle_tab_click(interaction, "yesterday")

    @discord.ui.button(label="Monthly", custom_id="monthly")
    async def monthly(self, interaction, button):
        await self.handle_tab_click(interaction, "monthly")

# Extended stats view (Template.xlsx layout)
class StatsFullView(discord.ui.View):
    def __init__(self, user_data, ign: str):
        super().__init__()
        self.ign = ign
        self.user_data = user_data
        self.meta = user_data.get("meta", {})
        self.current_tab = "all-time"
        self._load_color()
        
        self.update_buttons()

    def _load_color(self):
        """Load or reload the color and guild info for this username from database"""
        self.ign_color = None
        self.guild_tag = None
        self.guild_color = None
        try:
            meta = get_user_meta(self.ign)
            if meta:
                self.ign_color = meta.get('ign_color')
                self.guild_tag = meta.get('guild_tag')
                self.guild_color = meta.get('guild_hex')
                print(f"[DEBUG] Loaded color for {self.ign}: {self.ign_color}, guild: [{self.guild_tag}] ({self.guild_color})")
        except Exception as e:
            print(f"[WARNING] Failed to load color for {self.ign}: {e}")

    def _get_value(self, stat_key: str, tab_name: str) -> float:
        # Map tab names to cache keys
        key_map = {"all-time": "lifetime"}
        cache_key = key_map.get(tab_name, tab_name)
        return self.user_data.get("stats", {}).get(stat_key.lower(), {}).get(cache_key, 0)

    def _collect_stats(self, tab_name: str) -> dict:
        def safe_div(n, d):
            return n / d if d else 0
        def fmt_int(v):
            return f"{int(round(v)):,}"
        def fmt_ratio(v):
            return f"{v:.2f}"

        # Base values
        experience = self._get_value('experience', tab_name)
        playtime_seconds = self._get_value('playtime', tab_name)
        games = self._get_value('games_played', tab_name)
        wins = self._get_value('wins', tab_name)
        losses = self._get_value('losses', tab_name)
        kills = self._get_value('kills', tab_name)
        deaths = self._get_value('deaths', tab_name)
        coins = self._get_value('coins', tab_name)
        layers = self._get_value('available_layers', tab_name)
        damage = self._get_value('damage_dealt', tab_name)
        kills_void = self._get_value('kills_void', tab_name)
        deaths_void = self._get_value('deaths_void', tab_name)
        magic_wools = self._get_value('magic_wool_hit', tab_name)
        kills_explosive = self._get_value('kills_explosive', tab_name)
        deaths_explosive = self._get_value('deaths_explosive', tab_name)
        sheep_thrown = self._get_value('sheep_thrown', tab_name)
        kills_bow = self._get_value('kills_bow', tab_name)
        deaths_bow = self._get_value('deaths_bow', tab_name)
        kills_melee = self._get_value('kills_melee', tab_name)
        deaths_melee = self._get_value('deaths_melee', tab_name)

        # Derived values
        playtime_hours = playtime_seconds / 3600 if playtime_seconds else 0
        exp_per_hour = safe_div(experience, playtime_hours)
        exp_per_game = safe_div(experience, games)
        wins_per_hour = safe_div(wins, playtime_hours)
        kills_per_hour = safe_div(kills, playtime_hours)
        kdr = safe_div(kills, deaths) if deaths else kills
        wlr = safe_div(wins, losses) if losses else wins
        kills_per_game = safe_div(kills, games)
        kills_per_win = safe_div(kills, wins)
        damage_per_game = safe_div(damage, games)
        damage_per_sheep = safe_div(damage, sheep_thrown)
        void_kdr = safe_div(kills_void, deaths_void) if deaths_void else kills_void
        wools_per_game = safe_div(magic_wools, games)
        explosive_kdr = safe_div(kills_explosive, deaths_explosive) if deaths_explosive else kills_explosive
        sheeps_per_game = safe_div(sheep_thrown, games)
        bow_kdr = safe_div(kills_bow, deaths_bow) if deaths_bow else kills_bow
        melee_kdr = safe_div(kills_melee, deaths_melee) if deaths_melee else kills_melee

        stats = {
            "username": self.ign,
            "guild": f"[{self.guild_tag}]" if self.guild_tag else "N/A",
            "playtime": format_playtime(int(playtime_seconds)) if playtime_seconds else "0s",
            "level": fmt_int(self._get_value('level', tab_name)),
            "exp_per_hour": fmt_ratio(exp_per_hour),
            "exp_per_game": fmt_ratio(exp_per_game),
            "wins_per_hour": fmt_ratio(wins_per_hour),
            "kills_per_hour": fmt_ratio(kills_per_hour),
            "sheepwars_label": "",
            "wins": fmt_int(wins),
            "losses": fmt_int(losses),
            "wlr": fmt_ratio(wlr),
            "layers": fmt_int(layers),
            "coins": fmt_int(coins),
            "kills": fmt_int(kills),
            "deaths": fmt_int(deaths),
            "kdr": fmt_ratio(kdr),
            "kills_per_game": fmt_ratio(kills_per_game),
            "kills_per_win": fmt_ratio(kills_per_win),
            "damage": fmt_int(damage),
            "damage_per_game": fmt_ratio(damage_per_game),
            "damage_per_sheep": fmt_ratio(damage_per_sheep),
            "void_kills": fmt_int(kills_void),
            "void_deaths": fmt_int(deaths_void),
            "void_kdr": fmt_ratio(void_kdr),
            "magic_wools": fmt_int(magic_wools),
            "wools_per_game": fmt_ratio(wools_per_game),
            "explosive_kills": fmt_int(kills_explosive),
            "explosive_deaths": fmt_int(deaths_explosive),
            "explosive_kdr": fmt_ratio(explosive_kdr),
            "sheeps_thrown": fmt_int(sheep_thrown),
            "sheeps_per_game": fmt_ratio(sheeps_per_game),
            "bow_kills": fmt_int(kills_bow),
            "bow_deaths": fmt_int(deaths_bow),
            "bow_kdr": fmt_ratio(bow_kdr),
            "games_played": fmt_int(games),
            "melee_kills": fmt_int(kills_melee),
            "melee_deaths": fmt_int(deaths_melee),
            "melee_kdr": fmt_ratio(melee_kdr),
        }

        ordered_fields = [
            ("Wins", stats["wins"]), ("Losses", stats["losses"]), ("WLR", stats["wlr"]), ("Layers", stats["layers"]), ("Coins", stats["coins"]),
            ("Kills", stats["kills"]), ("Deaths", stats["deaths"]), ("KDR", stats["kdr"]), ("Kill/Game", stats["kills_per_game"]), ("Kill/Win", stats["kills_per_win"]),
            ("Damage dealt", stats["damage"]), ("Damage/Game", stats["damage_per_game"]), ("Void kills", stats["void_kills"]), ("Void deaths", stats["void_deaths"]), ("Void KDR", stats["void_kdr"]),
            ("Magic wools", stats["magic_wools"]), ("Wools/Game", stats["wools_per_game"]), ("Explosive kills", stats["explosive_kills"]), ("Explosive deaths", stats["explosive_deaths"]), ("Explosive KDR", stats["explosive_kdr"]),
            ("Sheeps thrown", stats["sheeps_thrown"]), ("Sheeps thrown/Game", stats["sheeps_per_game"]), ("Bow kills", stats["bow_kills"]), ("Bow deaths", stats["bow_deaths"]), ("Bow KDR", stats["bow_kdr"]),
            ("Games Played", stats["games_played"]), ("Damage/Sheep", stats["damage_per_sheep"]), ("Meelee kills", stats["melee_kills"]), ("Meelee Deaths", stats["melee_deaths"]), ("Meelee KDR", stats["melee_kdr"]),
        ]
        stats["ordered_fields"] = ordered_fields
        return stats

    def update_buttons(self):
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                child.style = discord.ButtonStyle.primary if child.custom_id == self.current_tab else discord.ButtonStyle.secondary

    def generate_full_image(self, tab_name: str):
        stats = self._collect_stats(tab_name)
        if Image is not None:
            try:
                img_io = create_full_stats_image(self.ign, tab_name, self.meta.get("level", 0), self.meta.get("icon", ""), stats, self.ign_color, self.guild_tag, self.guild_color)
                filename = f"{self.ign}_{tab_name}_stats_full.png"
                return None, discord.File(img_io, filename=filename)
            except Exception as e:
                print(f"[WARNING] Full stats image generation failed: {e}")

        embed = discord.Embed(title=f"{self.ign} - {tab_name.title()} stats")
        for label, value in stats.get("ordered_fields", [])[:25]:
            embed.add_field(name=label, value=f"```{value}```", inline=True)
        return embed, None

    @discord.ui.button(label="All-time", custom_id="all-time", style=discord.ButtonStyle.primary)
    async def full_all_time_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "all-time"
        self.update_buttons()
        embed, file = self.generate_full_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Session", custom_id="session", style=discord.ButtonStyle.secondary)
    async def full_session_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "session"
        self.update_buttons()
        embed, file = self.generate_full_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Daily", custom_id="daily", style=discord.ButtonStyle.secondary)
    async def full_daily_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "daily"
        self.update_buttons()
        embed, file = self.generate_full_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Yesterday", custom_id="yesterday", style=discord.ButtonStyle.secondary)
    async def full_yesterday_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "yesterday"
        self.update_buttons()
        embed, file = self.generate_full_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Monthly", custom_id="monthly", style=discord.ButtonStyle.secondary)
    async def full_monthly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "monthly"
        self.update_buttons()
        embed, file = self.generate_full_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)


class DistributionView(discord.ui.View):
    def __init__(self, user_data, ign: str, mode: str):
        super().__init__()
        self.ign = ign
        self.user_data = user_data
        self.mode = mode  # 'kill' or 'death'
        self.current_tab = "all-time"
        
        # Colors for legend slices
        self.slice_colors = {
            "void": (90, 155, 255),        # blue
            "explosive": (255, 119, 84),   # orange-red
            "bow": (255, 214, 102),        # golden
            "melee": (126, 217, 126),      # green
        }
        self.update_buttons()

    def update_buttons(self):
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                child.style = discord.ButtonStyle.primary if child.custom_id.endswith(self.current_tab) else discord.ButtonStyle.secondary

    def _get_counts(self, tab_name: str):
        if self.mode == "kill":
            keys = [
                ("Melee Kills", "kills_melee", "melee"),
                ("Bow Kills", "kills_bow", "bow"),
                ("Explosive Kills", "kills_explosive", "explosive"),
                ("Void Kills", "kills_void", "void"),
            ]
        else:
            keys = [
                ("Melee Deaths", "deaths_melee", "melee"),
                ("Bow Deaths", "deaths_bow", "bow"),
                ("Explosive Deaths", "deaths_explosive", "explosive"),
                ("Void Deaths", "deaths_void", "void"),
            ]

        key_map = {"all-time": "lifetime"}
        cache_key = key_map.get(tab_name, tab_name)
        counts = []
        stats = self.user_data.get("stats", {})
        for label, key, color_key in keys:
            val = stats.get(key, {}).get(cache_key, 0)
            counts.append((label, max(0, float(val)), color_key))
        return counts

    def generate_distribution(self, tab_name: str):
        counts = self._get_counts(tab_name)
        total = sum(v for _, v, _ in counts)
        metric_label = "Kill" if self.mode == "kill" else "Death"

        if total <= 0:
            embed = discord.Embed(
                title=f"{self.ign} - {tab_name.title()} {metric_label} Distribution",
                description="No data for this period.",
                color=discord.Color.from_rgb(54, 57, 63),
            )
            return embed, None

        slice_payload = []
        for label, value, color_key in counts:
            color = self.slice_colors.get(color_key, (180, 180, 180))
            slice_payload.append((label, value, color))

        if Image is not None:
            try:
                title = f"{self.ign} - {tab_name.title()} {metric_label} Distribution"
                img_io = create_distribution_pie(title, slice_payload)
                filename = f"{self.ign}_{self.mode}_{tab_name}_distribution.png"
                return None, discord.File(img_io, filename=filename)
            except Exception as e:
                print(f"[WARNING] Distribution image generation failed: {e}")

        # Fallback to embed if Pillow is missing or image failed
        embed = discord.Embed(
            title=f"{self.ign} - {tab_name.title()} {metric_label} Distribution",
            color=discord.Color.from_rgb(54, 57, 63),
        )
        lines = []
        for label, value, _ in counts:
            percent = (value / total * 100) if total else 0
            lines.append(f"{label}: {value} ({percent:.1f}%)")
        embed.description = "\n".join(lines)
        return embed, None

    @discord.ui.button(label="All-time", custom_id="dist-all-time", style=discord.ButtonStyle.primary)
    async def dist_all_time_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "all-time"
        self.update_buttons()
        embed, file = self.generate_distribution(self.current_tab)
        if file:
            await interaction.response.edit_message(embed=None, view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self, attachments=[])

    @discord.ui.button(label="Session", custom_id="dist-session", style=discord.ButtonStyle.secondary)
    async def dist_session_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "session"
        self.update_buttons()
        embed, file = self.generate_distribution(self.current_tab)
        if file:
            await interaction.response.edit_message(embed=None, view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self, attachments=[])

    @discord.ui.button(label="Daily", custom_id="dist-daily", style=discord.ButtonStyle.secondary)
    async def dist_daily_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "daily"
        self.update_buttons()
        embed, file = self.generate_distribution(self.current_tab)
        if file:
            await interaction.response.edit_message(embed=None, view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self, attachments=[])

    @discord.ui.button(label="Yesterday", custom_id="dist-yesterday", style=discord.ButtonStyle.secondary)
    async def dist_yesterday_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "yesterday"
        self.update_buttons()
        embed, file = self.generate_distribution(self.current_tab)
        if file:
            await interaction.response.edit_message(embed=None, view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self, attachments=[])

    @discord.ui.button(label="Monthly", custom_id="dist-monthly", style=discord.ButtonStyle.secondary)
    async def dist_monthly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "monthly"
        self.update_buttons()
        embed, file = self.generate_distribution(self.current_tab)
        if file:
            await interaction.response.edit_message(embed=None, view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self, attachments=[])


class LeaderboardView(discord.ui.View):
    def __init__(self, metric: str, data_cache: dict):
        super().__init__()
        self.metric = metric
        self.data_cache = data_cache
        self.current_period = "lifetime"
        self.page = 0
        self.page_size = 10
        
        # Column mappings for each period
        self.column_map = {
            "lifetime": "B",      # All-time values
            # Use the DELTA columns for period comparisons
            "session": "C",       # Session Delta
            "daily": "E",         # Daily Delta
            "yesterday": "G",     # Yesterday Delta
            "monthly": "I",       # Monthly Delta
        }
        
        self.metric_labels = {
            "kills": "Kills",
            "kills_void": "Void Kills",
            "kills_explosive": "Explosive Kills",
            "kills_melee": "Melee Kills",
            "kills_bow": "Bow Kills",
            "deaths": "Deaths",
            "deaths_void": "Void Deaths",
            "deaths_explosive": "Explosive Deaths",
            "deaths_melee": "Melee Deaths",
            "deaths_bow": "Bow Deaths",
            "kdr": "K/D Ratio",
            "wins": "Wins",
            "losses": "Losses",
            "wlr": "W/L Ratio",
            "experience": "Experience",
            "level": "Level",
            "coins": "Coins",
            "damage_dealt": "Damage Dealt",
            "games_played": "Games Played",
            "sheep_thrown": "Sheep Thrown",
            "magic_wool_hit": "Magic Wool Hit",
            "playtime": "Playtime",
        }
        
        # Load user colors
        self.user_colors = {}
        try:
            usernames = get_all_usernames()
            for username in usernames:
                meta = get_user_meta(username)
                if meta and meta.get('ign_color'):
                    self.user_colors[username.lower()] = meta.get('ign_color')
        except Exception as e:
            print(f"[WARNING] Failed to load user colors: {e}")

        # Period selector dropdown
        self.period_select = LeaderboardPeriodSelect(self)
        self.add_item(self.period_select)
        
    def _get_leaderboard(self, period: str):
        return self.metric_labels[self.metric], self.data_cache.get(period, [])

    def _paginate(self, leaderboard: list, page: int):
        total_pages = max(1, (len(leaderboard) + self.page_size - 1) // self.page_size)
        clamped_page = max(0, min(page, total_pages - 1))
        start_index = clamped_page * self.page_size
        return leaderboard[start_index:start_index + self.page_size], total_pages, clamped_page, start_index

    def generate_leaderboard_image(self, period: str, page: int):
        metric_label, leaderboard = self._get_leaderboard(period)

        if not leaderboard:
            empty_embed = self.get_leaderboard_embed(period, page=0, total_pages=1, leaderboard=leaderboard)
            return empty_embed, None, 1

        sliced, total_pages, clamped_page, start_index = self._paginate(leaderboard, page)
        self.page = clamped_page

        image_data = []
        for idx, entry in enumerate(sliced):
            player, _, value, is_playtime, level, icon, p_hex, g_tag, g_hex = entry
            rank = start_index + idx + 1
            image_data.append((rank, player, level, icon, p_hex, g_tag, g_hex, value, is_playtime))

        if Image is not None:
            try:
                img_io = create_leaderboard_image(period.title(), metric_label, image_data, page=clamped_page, total_pages=total_pages)
                filename = f"leaderboard_{self.metric}_{period}_p{clamped_page + 1}.png"
                return None, discord.File(img_io, filename=filename), total_pages
            except Exception as e:
                print(f"[WARNING] Leaderboard image generation failed: {e}")
                return self.get_leaderboard_embed(period, clamped_page, total_pages, leaderboard), None, total_pages
        else:
            return self.get_leaderboard_embed(period, clamped_page, total_pages, leaderboard), None, total_pages

    def get_leaderboard_embed(self, period: str, page: int = 0, total_pages: int = 1, leaderboard: Optional[list] = None):
        metric_label, leaderboard_data = self._get_leaderboard(period) if leaderboard is None else (self.metric_labels[self.metric], leaderboard)

        if not leaderboard_data:
            embed = discord.Embed(
                title=f"{period.title()} {metric_label} Leaderboard",
                description="No data available",
                color=discord.Color.from_rgb(54, 57, 63)
            )
            return embed

        sliced, total_pages, clamped_page, start_index = self._paginate(leaderboard_data, page)
        self.page = clamped_page

        embed = discord.Embed(
            title=f"{period.title()} {metric_label} Leaderboard",
            color=discord.Color.from_rgb(54, 57, 63)
        )

        description_lines = []
        for idx, entry in enumerate(sliced):
            player = entry[0]
            value = entry[2]
            is_playtime = entry[3]
            level_value = entry[4]
            icon = entry[5]

            medal = {1: "1.", 2: "2.", 3: "3."}.get(start_index + idx + 1, f"{start_index + idx + 1}.")
            prestige_display = format_prestige_ansi(level_value, icon)

            if is_playtime:
                formatted_value = format_playtime(int(value))
            elif "Ratio" in metric_label or "/" in metric_label or "Per" in metric_label:
                formatted_value = f"{float(value):,.2f}"
            else:
                formatted_value = f"{int(value):,}"

            description_lines.append(f"{medal} {prestige_display} {player}: {formatted_value}")

        embed.description = f"```ansi\n" + "\n".join(description_lines) + "\n```"
        embed.set_footer(text=f"Page {clamped_page + 1} of {total_pages}")
        return embed

    async def _refresh(self, interaction: discord.Interaction, *, new_period: Optional[str] = None, page_delta: int = 0):
        if new_period is not None:
            self.current_period = new_period
            self.page = 0
            # sync dropdown defaults
            for option in self.period_select.options:
                option.default = option.value == new_period
        else:
            self.page += page_delta

        embed, file, _ = self.generate_leaderboard_image(self.current_period, self.page)

        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Prev Page", custom_id="page_prev", style=discord.ButtonStyle.secondary)
    async def prev_page(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._refresh(interaction, page_delta=-1)

    @discord.ui.button(label="Next Page", custom_id="page_next", style=discord.ButtonStyle.secondary)
    async def next_page(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._refresh(interaction, page_delta=1)


class LeaderboardPeriodSelect(discord.ui.Select):
    def __init__(self, view: LeaderboardView):
        options = [
            discord.SelectOption(label="Lifetime", value="lifetime", default=True),
            discord.SelectOption(label="Session", value="session"),
            discord.SelectOption(label="Daily", value="daily"),
            discord.SelectOption(label="Yesterday", value="yesterday"),
            discord.SelectOption(label="Monthly", value="monthly"),
        ]
        super().__init__(
            placeholder="Select leaderboard period",
            min_values=1,
            max_values=1,
            options=options,
            custom_id="leaderboard_period_select",
        )
        self.view_ref = view

    async def callback(self, interaction: discord.Interaction):
        selected = self.values[0]
        for opt in self.options:
            opt.default = opt.value == selected
        await self.view_ref._refresh(interaction, new_period=selected)


def _load_leaderboard_data_from_excel(metric: str):
    """Load leaderboard data from database.
    
    Returns dict with period -> list of (username, value, display_value, is_playtime, level, icon, color, guild_tag, guild_color)
    """
    periods = ["lifetime", "session", "daily", "yesterday", "monthly"]
    result = {p: [] for p in periods}
    
    try:
        # Get all users from database
        usernames = get_all_usernames()
        
        # Load user colors
        user_colors = load_user_colors()
        
        for username in usernames:
            try:
                # Get stats with deltas
                stats_dict = get_user_stats_with_deltas(username)
                
                if not stats_dict:
                    continue
                
                # Get metadata
                user_meta = user_colors.get(username.lower(), {})
                
                # Try to get level from stats_dict first
                level = stats_dict.get("level", {}).get("lifetime", 0)
                if not level:
                    level = stats_dict.get("prestige level", {}).get("lifetime", 0)
                if not level:
                    level = 0
                else:
                    level = int(level) if level else 0
                
                icon = ""
                ign_color = user_meta.get("color", "#FFFFFF")
                guild_tag = user_meta.get("guild_tag")
                raw_g = str(user_meta.get('guild_color', 'GRAY')).upper()
                guild_color = raw_g if raw_g.startswith('#') else MINECRAFT_NAME_TO_HEX.get(raw_g, "#AAAAAA")
                
                # Process each period
                for period in periods:
                    val = 0
                    
                    # Map metric names to database columns
                    if metric == "kills":
                        val = stats_dict.get("kills", {}).get(period, 0)
                    elif metric == "deaths":
                        val = stats_dict.get("deaths", {}).get(period, 0)
                    elif metric == "kdr":
                        k = stats_dict.get("kills", {}).get(period, 0)
                        d = stats_dict.get("deaths", {}).get(period, 0)
                        val = k / d if d > 0 else k
                    elif metric == "wins":
                        val = stats_dict.get("wins", {}).get(period, 0)
                    elif metric == "losses":
                        val = stats_dict.get("losses", {}).get(period, 0)
                    elif metric == "wlr":
                        w = stats_dict.get("wins", {}).get(period, 0)
                        l = stats_dict.get("losses", {}).get(period, 0)
                        val = w / l if l > 0 else w
                    elif metric == "experience":
                        val = stats_dict.get("experience", {}).get(period, 0)
                    elif metric == "level":
                        val = stats_dict.get("level", {}).get(period, 0)
                    elif metric == "coins":
                        val = stats_dict.get("coins", {}).get(period, 0)
                    elif metric == "damage_dealt":
                        val = stats_dict.get("damage_dealt", {}).get(period, 0)
                    elif metric == "games_played":
                        val = stats_dict.get("games_played", {}).get(period, 0)
                    elif metric == "sheep_thrown":
                        val = stats_dict.get("sheep_thrown", {}).get(period, 0)
                    elif metric == "magic_wool_hit":
                        val = stats_dict.get("magic_wool_hit", {}).get(period, 0)
                    elif metric == "playtime":
                        val = stats_dict.get("playtime", {}).get(period, 0)
                    else:
                        val = stats_dict.get(metric, {}).get(period, 0)
                    
                    is_playtime = (metric == "playtime")
                    result[period].append((
                        username, float(val), val, is_playtime, level, icon, ign_color, guild_tag, guild_color
                    ))
            except Exception as e:
                print(f"[LEADERBOARD] Error processing {username}: {e}")
                continue
            
        # Sort each period by value descending
        for p in result:
            result[p].sort(key=lambda x: x[1], reverse=True)
        
        return result
    except Exception as e:
        print(f"[LEADERBOARD] Error loading from Excel: {e}")
        return result

def _process_leaderboard_data(cache_data, metric):
    periods = ["lifetime", "session", "daily", "yesterday", "monthly"]
    result = {p: [] for p in periods}

    for username, data in cache_data.items():
        stats = data.get("stats", {})
        meta = data.get("meta", {})
        
        for period in periods:
            val = 0
            if metric == "kdr":
                k = stats.get("kills", {}).get(period, 0)
                d = stats.get("deaths", {}).get(period, 0)
                val = k / d if d > 0 else k
            elif metric == "wlr":
                w = stats.get("wins", {}).get(period, 0)
                l = stats.get("losses", {}).get(period, 0)
                val = w / l if l > 0 else w
            else:
                val = stats.get(metric, {}).get(period, 0)
            
            is_playtime = (metric == "playtime")
            result[period].append((
                meta.get("username", username), float(val), val, is_playtime,
                meta.get("level", 0), meta.get("icon", ""), meta.get("ign_color"), meta.get("guild_tag"), meta.get("guild_hex")
            ))

    for p in result:
        result[p].sort(key=lambda x: x[1], reverse=True)
    return result

def _calculate_ratio_value_from_excel(stats_dict: dict, period: str, metric: str):
    """Calculate ratio values from Excel data."""
    try:
        if metric == "wl_ratio":
            wins = stats_dict.get("wins", {}).get(period, 0)
            losses = stats_dict.get("losses", {}).get(period, 0)
            return round(wins / losses, 2) if losses > 0 else wins
        elif metric == "kd_ratio":
            kills = stats_dict.get("kills", {}).get(period, 0)
            deaths = stats_dict.get("deaths", {}).get(period, 0)
            return round(kills / deaths, 2) if deaths > 0 else kills
        elif metric == "kills_per_game":
            kills = stats_dict.get("kills", {}).get(period, 0)
            games = stats_dict.get("games_played", {}).get(period, 0)
            return round(kills / games, 2) if games > 0 else 0
        elif metric == "kills_per_win":
            kills = stats_dict.get("kills", {}).get(period, 0)
            wins = stats_dict.get("wins", {}).get(period, 0)
            return round(kills / wins, 2) if wins > 0 else 0
        elif metric == "damage_per_game":
            damage = stats_dict.get("damage_dealt", {}).get(period, 0)
            games = stats_dict.get("games_played", {}).get(period, 0)
            return round(damage / games, 2) if games > 0 else 0
        elif metric == "damage_per_sheep":
            damage = stats_dict.get("damage_dealt", {}).get(period, 0)
            sheep = stats_dict.get("sheep_thrown", {}).get(period, 0)
            return round(damage / sheep, 2) if sheep > 0 else 0
        elif metric == "wools_per_game":
            wools = stats_dict.get("magic_wool_hit", {}).get(period, 0)
            games = stats_dict.get("games_played", {}).get(period, 0)
            return round(wools / games, 2) if games > 0 else 0
        elif metric == "void_kd_ratio":
            void_kills = stats_dict.get("kills_void", {}).get(period, 0)
            void_deaths = stats_dict.get("deaths_void", {}).get(period, 0)
            return round(void_kills / void_deaths, 2) if void_deaths > 0 else void_kills
        elif metric == "explosive_kd_ratio":
            exp_kills = stats_dict.get("kills_explosive", {}).get(period, 0)
            exp_deaths = stats_dict.get("deaths_explosive", {}).get(period, 0)
            return round(exp_kills / exp_deaths, 2) if exp_deaths > 0 else exp_kills
        elif metric == "bow_kd_ratio":
            bow_kills = stats_dict.get("kills_bow", {}).get(period, 0)
            bow_deaths = stats_dict.get("deaths_bow", {}).get(period, 0)
            return round(bow_kills / bow_deaths, 2) if bow_deaths > 0 else bow_kills
        elif metric == "melee_kd_ratio":
            melee_kills = stats_dict.get("kills_melee", {}).get(period, 0)
            melee_deaths = stats_dict.get("deaths_melee", {}).get(period, 0)
            return round(melee_kills / melee_deaths, 2) if melee_deaths > 0 else melee_kills
        elif metric == "exp_per_hour":
            exp = stats_dict.get("experience", {}).get(period, 0)
            playtime = stats_dict.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(exp / hours, 2) if hours > 0 else 0
        elif metric == "exp_per_game":
            exp = stats_dict.get("experience", {}).get(period, 0)
            games = stats_dict.get("games_played", {}).get(period, 0)
            return round(exp / games, 2) if games > 0 else 0
        elif metric == "wins_per_hour":
            wins = stats_dict.get("wins", {}).get(period, 0)
            playtime = stats_dict.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(wins / hours, 2) if hours > 0 else 0
        elif metric == "kills_per_hour":
            kills = stats_dict.get("kills", {}).get(period, 0)
            playtime = stats_dict.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(kills / hours, 2) if hours > 0 else 0
        elif metric == "sheeps_per_game":
            sheep = stats_dict.get("sheep_thrown", {}).get(period, 0)
            games = stats_dict.get("games_pulled", {}).get(period, 0)
            return round(sheep / games, 2) if games > 0 else 0
    except:
        return None
    return None

def _load_ratio_leaderboard_data_from_excel(metric: str):
    """Load ratio leaderboard data from database."""
    periods = ["lifetime", "session", "daily", "yesterday", "monthly"]
    result = {p: [] for p in periods}
    
    try:
        # Get all users from database
        usernames = get_all_usernames()
        
        # Load user colors
        user_colors = load_user_colors()
        
        for username in usernames:
            try:
                # Get stats with deltas
                stats_dict = get_user_stats_with_deltas(username)
                
                if not stats_dict:
                    continue
                
                # Get metadata
                user_meta = user_colors.get(username.lower(), {})
                
                # Try to get level from stats_dict first
                level = stats_dict.get("level", {}).get("lifetime", 0)
                if not level:
                    level = stats_dict.get("prestige level", {}).get("lifetime", 0)
                if not level:
                    level = 0
                else:
                    level = int(level) if level else 0
                
                icon = ""
                ign_color = user_meta.get("color", "#FFFFFF")
                guild_tag = user_meta.get("guild_tag")
                raw_g = str(user_meta.get('guild_color', 'GRAY')).upper()
                guild_color = raw_g if raw_g.startswith('#') else MINECRAFT_NAME_TO_HEX.get(raw_g, "#AAAAAA")
                
                # Process each period
                for period in periods:
                    val = _calculate_ratio_value_from_excel(stats_dict, period, metric)
                    if val is not None:
                        result[period].append((
                            username, float(val), val, level, icon, ign_color, guild_tag, guild_color
                        ))
            except Exception as e:
                print(f"[LEADERBOARD] Error processing {username}: {e}")
                continue
            
        # Sort each period by value descending
        for p in result:
            result[p].sort(key=lambda x: x[1], reverse=True)
        
        return result
    except Exception as e:
        print(f"[LEADERBOARD] Error loading ratio data from Excel: {e}")
        return result

def _calculate_ratio_value_from_cache(stats, period, metric):
    try:
        if metric == "wl_ratio":
            wins = stats.get("wins", {}).get(period, 0)
            losses = stats.get("losses", {}).get(period, 0)
            return round(wins / losses, 2) if losses > 0 else wins
        elif metric == "kd_ratio":
            kills = stats.get("kills", {}).get(period, 0)
            deaths = stats.get("deaths", {}).get(period, 0)
            return round(kills / deaths, 2) if deaths > 0 else kills
        elif metric == "kills_per_game":
            kills = stats.get("kills", {}).get(period, 0)
            games = stats.get("games_played", {}).get(period, 0)
            return round(kills / games, 2) if games > 0 else 0
        elif metric == "kills_per_win":
            kills = stats.get("kills", {}).get(period, 0)
            wins = stats.get("wins", {}).get(period, 0)
            return round(kills / wins, 2) if wins > 0 else 0
        elif metric == "damage_per_game":
            damage = stats.get("damage_dealt", {}).get(period, 0)
            games = stats.get("games_played", {}).get(period, 0)
            return round(damage / games, 2) if games > 0 else 0
        elif metric == "damage_per_sheep":
            damage = stats.get("damage_dealt", {}).get(period, 0)
            sheep = stats.get("sheep_thrown", {}).get(period, 0)
            return round(damage / sheep, 2) if sheep > 0 else 0
        elif metric == "wools_per_game":
            wools = stats.get("magic_wool_hit", {}).get(period, 0)
            games = stats.get("games_played", {}).get(period, 0)
            return round(wools / games, 2) if games > 0 else 0
        elif metric == "void_kd_ratio":
            void_kills = stats.get("kills_void", {}).get(period, 0)
            void_deaths = stats.get("deaths_void", {}).get(period, 0)
            return round(void_kills / void_deaths, 2) if void_deaths > 0 else void_kills
        elif metric == "explosive_kd_ratio":
            exp_kills = stats.get("kills_explosive", {}).get(period, 0)
            exp_deaths = stats.get("deaths_explosive", {}).get(period, 0)
            return round(exp_kills / exp_deaths, 2) if exp_deaths > 0 else exp_kills
        elif metric == "bow_kd_ratio":
            bow_kills = stats.get("kills_bow", {}).get(period, 0)
            bow_deaths = stats.get("deaths_bow", {}).get(period, 0)
            return round(bow_kills / bow_deaths, 2) if bow_deaths > 0 else bow_kills
        elif metric == "melee_kd_ratio":
            melee_kills = stats.get("kills_melee", {}).get(period, 0)
            melee_deaths = stats.get("deaths_melee", {}).get(period, 0)
            return round(melee_kills / melee_deaths, 2) if melee_deaths > 0 else melee_kills
        elif metric == "exp_per_hour":
            exp = stats.get("experience", {}).get(period, 0)
            playtime = stats.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(exp / hours, 2) if hours > 0 else 0
        elif metric == "exp_per_game":
            exp = stats.get("experience", {}).get(period, 0)
            games = stats.get("games_played", {}).get(period, 0)
            return round(exp / games, 2) if games > 0 else 0
        elif metric == "wins_per_hour":
            wins = stats.get("wins", {}).get(period, 0)
            playtime = stats.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(wins / hours, 2) if hours > 0 else 0
        elif metric == "kills_per_hour":
            kills = stats.get("kills", {}).get(period, 0)
            playtime = stats.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(kills / hours, 2) if hours > 0 else 0
        elif metric == "sheeps_per_game":
            sheep = stats.get("sheep_thrown", {}).get(period, 0)
            games = stats.get("games_played", {}).get(period, 0)
            return round(sheep / games, 2) if games > 0 else 0
    except:
        return None
    return None

def _process_ratio_data(cache_data, metric):
    periods = ["lifetime", "session", "daily", "yesterday", "monthly"]
    result = {p: [] for p in periods}

    for username, data in cache_data.items():
        stats = data.get("stats", {})
        meta = data.get("meta", {})
        
        for period in periods:
            val = _calculate_ratio_value_from_cache(stats, period, metric)
            if val is not None:
                result[period].append((
                    meta.get("username", username), float(val), val, 
                    meta.get("level", 0), meta.get("icon", ""), meta.get("ign_color"), meta.get("guild_tag"), meta.get("guild_hex")
                ))

    for p in result:
        result[p].sort(key=lambda x: x[1], reverse=True)
    return result

class RatioLeaderboardView(discord.ui.View):
    def __init__(self, metric: str, data_cache: dict):
        super().__init__()
        self.metric = metric
        self.data_cache = data_cache
        self.current_period = "lifetime"
        self.page = 0
        self.page_size = 10
        
        # Column mappings for each period
        self.column_map = {
            "lifetime": "B",      # All-time values
            # Use the DELTA columns for period comparisons
            "session": "C",       # Session Delta
            "daily": "E",         # Daily Delta
            "yesterday": "G",     # Yesterday Delta
            "monthly": "I",       # Monthly Delta
        }
        
        self.metric_labels = {
            "wl_ratio": "W/L Ratio",
            "kd_ratio": "K/D Ratio",
            "kills_per_game": "Kills/Game",
            "kills_per_win": "Kills/Win",
            "kills_per_hour": "Kills/Hour",
            "damage_per_game": "Damage/Game",
            "damage_per_sheep": "Damage/Sheep",
            "wools_per_game": "Wools/Game",
            "sheeps_per_game": "Sheeps/Game",
            "void_kd_ratio": "Void K/D Ratio",
            "explosive_kd_ratio": "Explosive K/D Ratio",
            "bow_kd_ratio": "Bow K/D Ratio",
            "melee_kd_ratio": "Melee K/D Ratio",
            "wins_per_hour": "Wins/Hour",
            "exp_per_hour": "EXP/Hour",
            "exp_per_game": "EXP/Game",
        }
        
        # Period selector dropdown
        self.period_select = RatioPeriodSelect(self)
        self.add_item(self.period_select)
        
    def _get_leaderboard(self, period: str):
        return self.metric_labels[self.metric], self.data_cache.get(period, [])

    def _paginate(self, leaderboard: list, page: int):
        total_pages = max(1, (len(leaderboard) + self.page_size - 1) // self.page_size)
        clamped_page = max(0, min(page, total_pages - 1))
        start_index = clamped_page * self.page_size
        return leaderboard[start_index:start_index + self.page_size], total_pages, clamped_page, start_index

    def generate_leaderboard_image(self, period: str, page: int):
        metric_label, leaderboard = self._get_leaderboard(period)

        if not leaderboard:
            empty_embed = self.get_leaderboard_embed(period, page=0, total_pages=1, leaderboard=leaderboard)
            return empty_embed, None, 1

        sliced, total_pages, clamped_page, start_index = self._paginate(leaderboard, page)
        self.page = clamped_page

        image_data = []
        for idx, entry in enumerate(sliced):
            player, _, value, level_value, icon, ign_color, g_tag, g_hex = entry
            rank = start_index + idx + 1
            image_data.append((rank, player, level_value, icon, ign_color, g_tag, g_hex, value, False))

        if Image is not None:
            try:
                img_io = create_leaderboard_image(period.title(), metric_label, image_data, page=clamped_page, total_pages=total_pages)
                filename = f"ratio_leaderboard_{self.metric}_{period}_p{clamped_page + 1}.png"
                return None, discord.File(img_io, filename=filename), total_pages
            except Exception as e:
                print(f"[WARNING] Ratio leaderboard image generation failed: {e}")
                return self.get_leaderboard_embed(period, clamped_page, total_pages, leaderboard), None, total_pages
        else:
            return self.get_leaderboard_embed(period, clamped_page, total_pages, leaderboard), None, total_pages

    def get_leaderboard_embed(self, period: str, page: int = 0, total_pages: int = 1, leaderboard: Optional[list] = None):
        metric_label, leaderboard_data = self._get_leaderboard(period) if leaderboard is None else (self.metric_labels[self.metric], leaderboard)

        if not leaderboard_data:
            embed = discord.Embed(
                title=f"{period.title()} {metric_label} Leaderboard",
                description="No data available",
                color=discord.Color.from_rgb(54, 57, 63)
            )
            return embed

        sliced, total_pages, clamped_page, start_index = self._paginate(leaderboard_data, page)
        self.page = clamped_page

        embed = discord.Embed(
            title=f"{period.title()} {metric_label} Leaderboard",
            color=discord.Color.from_rgb(54, 57, 63)
        )

        description_lines = []
        for idx, entry in enumerate(sliced):
            player = entry[0]
            value = entry[2]
            level_value = entry[3]
            icon = entry[4]

            medal = {1: "1.", 2: "2.", 3: "3."}.get(start_index + idx + 1, f"{start_index + idx + 1}.")
            prestige_display = format_prestige_ansi(level_value, icon)

            if "Ratio" in metric_label or "/" in metric_label or "Per" in metric_label:
                formatted_value = f"{float(value):,.2f}"
            else:
                formatted_value = f"{value}"

            description_lines.append(f"{medal} {prestige_display} {player}: {formatted_value}")

        embed.description = f"```ansi\n" + "\n".join(description_lines) + "\n```"
        embed.set_footer(text=f"Page {clamped_page + 1} of {total_pages}")
        return embed

    async def _refresh(self, interaction: discord.Interaction, *, new_period: Optional[str] = None, page_delta: int = 0):
        if new_period is not None:
            self.current_period = new_period
            self.page = 0
            # sync dropdown defaults
            for option in self.period_select.options:
                option.default = option.value == new_period
        else:
            self.page += page_delta

        embed, file, _ = self.generate_leaderboard_image(self.current_period, self.page)

        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Prev Page", custom_id="page_prev_ratio", style=discord.ButtonStyle.secondary)
    async def prev_page(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._refresh(interaction, page_delta=-1)

    @discord.ui.button(label="Next Page", custom_id="page_next_ratio", style=discord.ButtonStyle.secondary)
    async def next_page(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._refresh(interaction, page_delta=1)


class RatioPeriodSelect(discord.ui.Select):
    def __init__(self, view: RatioLeaderboardView):
        options = [
            discord.SelectOption(label="Lifetime", value="lifetime", default=True),
            discord.SelectOption(label="Session", value="session"),
            discord.SelectOption(label="Daily", value="daily"),
            discord.SelectOption(label="Yesterday", value="yesterday"),
            discord.SelectOption(label="Monthly", value="monthly"),
        ]
        super().__init__(
            placeholder="Select leaderboard period",
            min_values=1,
            max_values=1,
            options=options,
            custom_id="ratio_leaderboard_period_select",
        )
        self.view_ref = view

    async def callback(self, interaction: discord.Interaction):
        selected = self.values[0]
        for opt in self.options:
            opt.default = opt.value == selected
        await self.view_ref._refresh(interaction, new_period=selected)


# Create bot with command tree for slash commands
intents = discord.Intents.default()
# Enabled intents: members and presences required for member/presence features;
# message_content allows reading message content if needed (user enabled in Dev Portal).
intents.members = True
intents.presences = True
intents.message_content = True
bot = commands.Bot(command_prefix="!", intents=intents)

# In-memory pending claim registry so approvals can be handled via slash command even after buttons expire
PENDING_CLAIMS: dict[int, dict] = {}
PENDING_STREAKS: dict[int, dict] = {}

# Approval system for claim command
class ApprovalView(discord.ui.View):
    def __init__(self, ign: str, requester: str, requester_id: int, original_interaction: discord.Interaction):
        super().__init__(timeout=None)
        self.ign = ign
        self.requester = requester
        self.requester_id = requester_id
        self.original_interaction = original_interaction
        self.approved = None
        self.done_event = asyncio.Event()
        self.processed_by_admin_command = False
    
    @discord.ui.button(label="Accept", style=discord.ButtonStyle.success)
    async def accept_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.approved = True
        PENDING_CLAIMS.pop(self.requester_id, None)
        self.done_event.set()
        await interaction.response.edit_message(content=f"You accepted claim for {self.ign}.", view=None)
    
    @discord.ui.button(label="Deny", style=discord.ButtonStyle.danger)
    async def deny_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.approved = False
        PENDING_CLAIMS.pop(self.requester_id, None)
        self.done_event.set()
        await interaction.response.edit_message(content=f"You denied claim for {self.ign}.", view=None)


class StreakApprovalView(discord.ui.View):
    def __init__(self, ign: str, requester: str, requester_id: int, stats_snapshot: dict):
        super().__init__(timeout=None)
        self.ign = ign
        self.requester = requester
        self.requester_id = requester_id
        self.stats_snapshot = stats_snapshot or {}
        self.approved = None
        self.done_event = asyncio.Event()
        self.processed_by_admin_command = False

    @discord.ui.button(label="Accept", style=discord.ButtonStyle.success)
    async def accept_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        try:
            initialize_streak_entry(self.ign, self.stats_snapshot)
            PENDING_STREAKS.pop(self.requester_id, None)
            self.approved = True
            self.done_event.set()
            try:
                requester_user = interaction.client.get_user(self.requester_id) or await interaction.client.fetch_user(self.requester_id)
                if requester_user:
                    await requester_user.send(f"✅ Your streak tracking request for {self.ign} was approved.")
            except Exception:
                pass
            await interaction.response.edit_message(content=f"You approved streak tracking for {self.ign}.", view=None)
        except Exception as e:
            await interaction.response.edit_message(content=f"[ERROR] Failed to initialize streaks: {e}", view=None)

    @discord.ui.button(label="Deny", style=discord.ButtonStyle.danger)
    async def deny_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        PENDING_STREAKS.pop(self.requester_id, None)
        self.approved = False
        self.done_event.set()
        try:
            requester_user = interaction.client.get_user(self.requester_id) or await interaction.client.fetch_user(self.requester_id)
            if requester_user:
                await requester_user.send(f"❌ Your streak tracking request for {self.ign} was denied.")
        except Exception:
            pass
        await interaction.response.edit_message(content=f"You denied streak tracking for {self.ign}.", view=None)


class StreakRequestView(discord.ui.View):
    def __init__(self, ign: str, requester: discord.User, stats_snapshot: dict):
        super().__init__()
        self.ign = ign
        self.requester = requester
        self.requester_id = requester.id
        self.stats_snapshot = stats_snapshot or {}

    @discord.ui.button(label="Request tracking", style=discord.ButtonStyle.primary)
    async def request_tracking(self, interaction: discord.Interaction, button: discord.ui.Button):
        if interaction.user.id != self.requester_id:
            await interaction.response.send_message("This request button is only for the original requester.", ephemeral=True)
            return

        if self.requester_id in PENDING_STREAKS:
            await interaction.response.send_message("You already have a pending streak tracking request.", ephemeral=True)
            return

        admins = []
        for admin_id in ADMIN_IDS:
            try:
                uid = int(admin_id)
                user = interaction.client.get_user(uid) or await interaction.client.fetch_user(uid)
                if user:
                    admins.append(user)
            except Exception:
                pass

        if not admins:
            await interaction.response.send_message("[ERROR] Cannot reach administrators for approval.", ephemeral=True)
            return

        view = StreakApprovalView(self.ign, self.requester.name, self.requester_id, self.stats_snapshot)
        _register_pending_streak(self.requester_id, self.ign, self.stats_snapshot, view)

        sent_count = 0
        for admin in admins:
            try:
                await admin.send(
                    f"{self.requester.name} ({self.requester_id}) requests streak tracking for {self.ign}.\n"
                    f"Click Accept/Deny below or run /verification-streak user:{self.requester_id} option: accept/deny.",
                    view=view,
                )
                sent_count += 1
            except Exception:
                pass

        if sent_count == 0:
            _pop_pending_streak(self.requester_id)
            await interaction.response.send_message(f"[ERROR] Could not send streak approval request to administrators.", ephemeral=True)
            return

        await interaction.response.send_message("✅ Sent streak tracking request for approval.", ephemeral=True)


def _register_pending_claim(user_id: int, ign: str, view: ApprovalView):
    PENDING_CLAIMS[user_id] = {"ign": ign, "view": view}


def _pop_pending_claim(user_id: int):
    return PENDING_CLAIMS.pop(user_id, None)


def _register_pending_streak(user_id: int, ign: str, stats_snapshot: dict, view):
    PENDING_STREAKS[user_id] = {"ign": ign, "stats": stats_snapshot, "view": view}


def _pop_pending_streak(user_id: int):
    return PENDING_STREAKS.pop(user_id, None)

# Bot token
# Read from BOT_TOKEN.txt in the same directory
TOKEN_FILE = os.path.join(os.path.dirname(__file__), "BOT_TOKEN.txt")
try:
    with open(TOKEN_FILE, "r", encoding="utf-8") as f:
        DISCORD_TOKEN = f.read().strip()
except Exception as e:
    DISCORD_TOKEN = None
    print(f"[ERROR] Failed to read BOT_TOKEN.txt: {e}")
if not DISCORD_TOKEN:
    raise ValueError("BOT_TOKEN.txt is missing or empty")

@bot.event
async def on_ready():
    import time
    bot_instance_id = int(time.time() * 1000) % 100000
    print(f"[OK] Bot logged in as {bot.user} - Instance ID: {bot_instance_id}")
    
    # Initialize database schema to ensure all tables exist
    await asyncio.to_thread(init_database)
    
    # Check for legacy data migration
    await check_legacy_migration()
    
    # Verify API key
    await asyncio.to_thread(verify_api_key)
    
    try:
        synced = await bot.tree.sync()
        print(f"[OK] Synced {len(synced)} command(s) - Instance ID: {bot_instance_id}")

    except Exception as e:
        print(f"[ERROR] Failed to sync commands: {e}")
    # start background tasks once
    if not getattr(bot, "background_tasks_started", False):
        # Store task references for graceful shutdown
        bot.background_tasks = [
            bot.loop.create_task(scheduler_loop()),
            bot.loop.create_task(staggered_stats_refresher(interval_minutes=10)),
            bot.loop.create_task(presence_updater_loop(interval_seconds=5))
        ]
        bot.background_tasks_started = True
        print(f"[OK] Background tasks started - Instance ID: {bot_instance_id}")

@bot.tree.command(name="track", description="Create a stats sheet for a player (no authorization required)")
@discord.app_commands.describe(ign="Minecraft IGN")
async def track(interaction: discord.Interaction, ign: str):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        # Check if user is already in tracked users database
        tracked_users = load_tracked_users()
        key = ign.casefold()
        for tracked_user in tracked_users:
            if tracked_user.casefold() == key:
                await interaction.followup.send(f"{tracked_user} is already being tracked.")
                return
        
        # Create sheet using api_get.py
        # Initialize session, daily, and monthly snapshots (yesterday will be populated from daily rotation)
        result = run_script("api_get.py", ["-ign", ign, "-session", "-yesterday", "-daily", "-monthly"])

        if result.returncode == 0:
            print(f"[OK] api_get.py succeeded for {ign}")
            
            # Parse output to get proper username
            actual_ign = ign
            if result.stdout:
                try:
                    for line in reversed(result.stdout.splitlines()):
                        if line.strip().startswith('{'):
                            data = json.loads(line.strip())
                            if 'username' in data:
                                actual_ign = data['username']
                                break
                except:
                    pass
            
            # Verify user exists in DB
            if not user_exists(actual_ign):
                await interaction.followup.send(f"[ERROR] Database entry for {ign} was not created.")
                return
            
            # Add to tracked users list using the properly-cased username
            added = add_tracked_user(actual_ign)
            
            if added:
                await interaction.followup.send(f"{actual_ign} is now being tracked. Use `/claim ign:{actual_ign}` to link this username to your Discord account.")
            else:
                await interaction.followup.send(f"{actual_ign} is already being tracked.")
        else:
            err = (result.stderr or result.stdout) or "Unknown error"
            print(f"[ERROR] api_get.py failed for {ign}:")
            print(f"  stdout: {result.stdout}")
            print(f"  stderr: {result.stderr}")
            await interaction.followup.send(f"Error creating sheet for {ign}:\n```{sanitize_output(err[:500])}```")
            
    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)")
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

@bot.tree.command(name="claim", description="Link a Minecraft username to your Discord account")
@discord.app_commands.describe(ign="Minecraft IGN")
async def claim(interaction: discord.Interaction, ign: str):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        # Check if username is tracked
        users = load_tracked_users()
        found = False
        for u in users:
            if u.casefold() == ign.casefold():
                found = True
                break
        
        if not found:
            await interaction.followup.send(f"[ERROR] {ign} is not being tracked. Use `/track ign:{ign}` first.")
            return
        
        # Check if already claimed
        links = load_user_links()
        if ign.casefold() in links:
            claimed_by = links[ign.casefold()]
            if claimed_by == str(interaction.user.id):
                await interaction.followup.send(f"[ERROR] You have already claimed {ign}.")
            else:
                await interaction.followup.send(f"[ERROR] {ign} is already claimed by another user.")
            return
        
        # Get admin users
        admins = []
        for admin_id in ADMIN_IDS:
            try:
                uid = int(admin_id)
                user = bot.get_user(uid) or await bot.fetch_user(uid)
                if user:
                    admins.append(user)
            except Exception:
                pass
        
        if not admins:
            await interaction.followup.send("[ERROR] Cannot reach administrators for approval.")
            return
        
        # Send waiting message to requester
        requester_name = interaction.user.name
        await interaction.followup.send(f"Asked administrators for approval to claim {ign}. Please wait for confirmation.")
        
        # Create approval view and send to admins
        view = ApprovalView(ign, requester_name, interaction.user.id, interaction)
        _register_pending_claim(interaction.user.id, ign, view)
        
        sent_count = 0
        for admin in admins:
            try:
                await admin.send(
                    f"{requester_name} ({interaction.user.id}) wants to claim {ign}.\n"
                    f"Click Accept/Deny below or run /verification user:{interaction.user.id} option: accept/deny.",
                    view=view,
                )
                sent_count += 1
            except Exception:
                pass
        
        if sent_count == 0:
            _pop_pending_claim(interaction.user.id)
            await interaction.followup.send(f"[ERROR] Could not send approval request to administrators.")
            return
        
        # Wait for approval (no timeout)
        await view.done_event.wait()
        
        # Process based on approval
        if getattr(view, "processed_by_admin_command", False):
            # Manual /verification already handled linking/notification
            if view.approved:
                await interaction.followup.send(f"Chuckegg has approved your claim. {ign} is now linked to your Discord account.")
            else:
                await interaction.followup.send(f"Chuckegg has denied your claim for {ign}.")
            return

        _pop_pending_claim(interaction.user.id)

        if view.approved:
            link_user_to_ign(interaction.user.id, ign)
            await interaction.followup.send(f"An administrator has approved your claim. {ign} is now linked to your Discord account.")
        else:
            await interaction.followup.send(f"An administrator has denied your claim for {ign}.")
            
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

@bot.tree.command(name="unclaim", description="Unlink a Minecraft username from your Discord account")
@discord.app_commands.describe(ign="Minecraft IGN to unclaim")
async def unclaim(interaction: discord.Interaction, ign: str):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    # Check if user is authorized to unclaim this username
    if not is_user_authorized(interaction.user.id, ign):
        await interaction.followup.send(f"[ERROR] You are not authorized to unclaim {ign}. Only the user who claimed this username can unclaim it.")
        return
    
    try:
        # Remove from user links
        removed_link = unlink_user_from_ign(ign)
        
        if removed_link:
            await interaction.followup.send(f"Successfully unclaimed {ign}. You are no longer linked to this username.")
        else:
            await interaction.followup.send(f"[WARNING] No claim found for {ign}.")
            
    except Exception as e:
        await interaction.followup.send(f"[ERROR] Failed to unclaim: {str(e)}")


@bot.tree.command(name="verification", description="Manually approve or deny a claim (admin only)")
@discord.app_commands.describe(option="Accept or deny the claim", user="Discord user ID of the requester")
@discord.app_commands.choices(option=[
    discord.app_commands.Choice(name="Accept", value="accept"),
    discord.app_commands.Choice(name="Deny", value="deny"),
])
async def verification(interaction: discord.Interaction, option: discord.app_commands.Choice[str], user: str):
    if not is_admin(interaction.user):
        await interaction.response.send_message("❌ This command is only available to bot administrators.", ephemeral=True)
        return

    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    try:
        user_id = int(user)
    except ValueError:
        await interaction.followup.send("[ERROR] Invalid user ID.", ephemeral=True)
        return

    pending = _pop_pending_claim(user_id)
    if not pending:
        await interaction.followup.send("[ERROR] No pending claim found for that user.", ephemeral=True)
        return

    ign = pending.get("ign")
    view = pending.get("view")
    approved = option.value == "accept"

    # If we have the original view, mark it and unblock the waiting task
    if view:
        view.approved = approved
        view.processed_by_admin_command = True
        view.done_event.set()

    requester_user = None
    try:
        requester_user = bot.get_user(user_id) or await bot.fetch_user(user_id)
    except Exception:
        requester_user = None

    if approved:
        link_user_to_ign(user_id, ign)
        if requester_user:
            try:
                await requester_user.send(f"✅ Your claim for {ign} was approved by an admin.")
            except Exception:
                pass
        await interaction.followup.send(f"Approved claim: {ign} linked to <@{user_id}>.", ephemeral=True)
    else:
        if requester_user:
            try:
                await requester_user.send(f"❌ Your claim for {ign} was denied by an admin.")
            except Exception:
                pass
        await interaction.followup.send(f"Denied claim for {ign} (requester <@{user_id}>).", ephemeral=True)


@bot.tree.command(name="verification-streak", description="Approve or deny a streak tracking request (admin only)")
@discord.app_commands.describe(option="Accept or deny the streak request", user="Discord user ID of the requester")
@discord.app_commands.choices(option=[
    discord.app_commands.Choice(name="Accept", value="accept"),
    discord.app_commands.Choice(name="Deny", value="deny"),
])
async def verification_streak(interaction: discord.Interaction, option: discord.app_commands.Choice[str], user: str):
    if not is_admin(interaction.user):
        await interaction.response.send_message("❌ This command is only available to bot administrators.", ephemeral=True)
        return

    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    try:
        user_id = int(user)
    except ValueError:
        await interaction.followup.send("[ERROR] Invalid user ID.", ephemeral=True)
        return

    pending = _pop_pending_streak(user_id)
    if not pending:
        await interaction.followup.send("[ERROR] No pending streak request found for that user.", ephemeral=True)
        return

    ign = pending.get("ign")
    stats_snapshot = pending.get("stats", {})
    view = pending.get("view")
    approved = option.value == "accept"

    requester_user = None
    try:
        requester_user = bot.get_user(user_id) or await bot.fetch_user(user_id)
    except Exception:
        requester_user = None

    if approved:
        initialize_streak_entry(ign, stats_snapshot)
        if view:
            view.approved = True
            view.processed_by_admin_command = True
            view.done_event.set()
        if requester_user:
            try:
                await requester_user.send(f"✅ Your streak tracking request for {ign} was approved by an admin.")
            except Exception:
                pass
        await interaction.followup.send(f"Approved streak tracking for {ign} (requester <@{user_id}>).", ephemeral=True)
    else:
        if view:
            view.approved = False
            view.processed_by_admin_command = True
            view.done_event.set()
        if requester_user:
            try:
                await requester_user.send(f"❌ Your streak tracking request for {ign} was denied by an admin.")
            except Exception:
                pass
        await interaction.followup.send(f"Denied streak tracking for {ign} (requester <@{user_id}>).", ephemeral=True)

@bot.tree.command(name="untrack", description="Remove all tracking data for a Minecraft username")
@discord.app_commands.describe(ign="Minecraft IGN to untrack")
async def untrack(interaction: discord.Interaction, ign: str):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    # Allow creator override; otherwise require claim authorization
    if not (is_admin(interaction.user) or is_user_authorized(interaction.user.id, ign)):
        await interaction.followup.send(f"[ERROR] You are not authorized to untrack {ign}. Only the user who claimed this username or the creator can untrack it.")
        return
    
    try:
        actual_ign = ign
        
        # Remove from all locations
        removed_tracked = remove_tracked_user(actual_ign)
        removed_link = unlink_user_from_ign(actual_ign)
        removed_color = remove_user_color(actual_ign)
        removed_sheet = delete_user_sheet(actual_ign)
        
        if removed_tracked or removed_link or removed_color or removed_sheet:
            results = []
            if removed_tracked:
                results.append("tracked users list")
            if removed_link:
                results.append("user links")
            if removed_color:
                results.append("user colors")
            if removed_sheet:
                results.append("stats sheet")
            
            await interaction.followup.send(f"Successfully untracked {actual_ign}. Removed from: {', '.join(results)}.")
        else:
            await interaction.followup.send(f"[WARNING] {ign} was not found in any tracking data.")
            
    except Exception as e:
        await interaction.followup.send(f"[ERROR] Failed to untrack: {str(e)}")

# Create color choices from MINECRAFT_CODE_TO_HEX
COLOR_CHOICES = [
    discord.app_commands.Choice(name="Black", value="0"),
    discord.app_commands.Choice(name="Dark Blue", value="1"),
    discord.app_commands.Choice(name="Dark Green", value="2"),
    discord.app_commands.Choice(name="Dark Aqua", value="3"),
    discord.app_commands.Choice(name="Dark Red", value="4"),
    discord.app_commands.Choice(name="Dark Purple", value="5"),
    discord.app_commands.Choice(name="Gold", value="6"),
    discord.app_commands.Choice(name="Gray", value="7"),
    discord.app_commands.Choice(name="Dark Gray", value="8"),
    discord.app_commands.Choice(name="Blue", value="9"),
    discord.app_commands.Choice(name="Green", value="a"),
    discord.app_commands.Choice(name="Aqua", value="b"),
    discord.app_commands.Choice(name="Red", value="c"),
    discord.app_commands.Choice(name="Light Purple/Pink", value="d"),
    discord.app_commands.Choice(name="Yellow", value="e"),
    discord.app_commands.Choice(name="White", value="f"),
]

@bot.tree.command(name="color", description="Set a custom color for your username in stats displays")
@discord.app_commands.describe(
    ign="Minecraft IGN (optional if you set /default)",
    color="Color for your username"
)
@discord.app_commands.choices(color=COLOR_CHOICES)
async def color(interaction: discord.Interaction, ign: str = None, color: discord.app_commands.Choice[str] = None):
    # Resolve default IGN if not provided, and validate before any heavy work
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    # Validate username via Mojang API and simple format
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            pass
    
    # Check if user is authorized to change color for this username
    if not is_user_authorized(interaction.user.id, ign):
        await interaction.followup.send(f"[ERROR] You are not authorized to change the color for {ign}. Only the user who claimed this username can change its color.", ephemeral=True)
        return
    
    try:
        # Get hex color from code
        color_code = color.value
        hex_color = MINECRAFT_CODE_TO_HEX.get(color_code, '#FFFFFF')
        
        # Update color in database
        update_user_meta(ign, ign_color=hex_color)
        
        await interaction.followup.send(f"Successfully set {ign}'s username color to {color.name}!", ephemeral=True)
        
    except Exception as e:
        await interaction.followup.send(f"[ERROR] Failed to set color: {str(e)}", ephemeral=True)

@bot.tree.command(name="reset", description="Reset session snapshot for a player")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def reset(interaction: discord.Interaction, ign: str = None):
    # Resolve default IGN and validate before any heavy work
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            pass
    
    # Check if user is authorized to reset session for this username
    if not is_user_authorized(interaction.user.id, ign):
        await interaction.followup.send(f"[ERROR] You are not authorized to reset session for {ign}. Only the user who claimed this username can reset its session.", ephemeral=True)
        return
    
    try:
        result = run_script("api_get.py", ["-ign", ign, "-session"])

        if result.returncode == 0:
            await interaction.followup.send(f"Session snapshot reset for {ign}.", ephemeral=True)
        else:
            err = (result.stderr or result.stdout) or "Unknown error"
            await interaction.followup.send(f"[ERROR] {sanitize_output(err)}", ephemeral=True)
    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)", ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}", ephemeral=True)

@bot.tree.command(name="dmme", description="Send yourself a test DM from the bot")
async def dmme(interaction: discord.Interaction):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    try:
        await interaction.user.send("Hello! This is a private message from the bot.")
        await interaction.followup.send("Sent you a DM.", ephemeral=True)
    except Exception as e:
        await interaction.followup.send("Couldn't DM you. Check your privacy settings (Allow DMs from server members).", ephemeral=True)


@bot.tree.command(name="default", description="Set your default Minecraft username")
@discord.app_commands.describe(ign="Minecraft IGN to use by default")
async def default(interaction: discord.Interaction, ign: str):
    # Validate username before persisting
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return

    # Quick response, no heavy work
    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            pass
    try:
        # Optionally validate tracked status to help users
        excel_file = BOT_DIR / "stats.xlsx"
        is_known = False
        if excel_file.exists():
            wb = None
            try:
                wb = load_workbook(str(excel_file), read_only=True, data_only=True)
                key = proper_ign.casefold()
                for sheet_name in wb.sheetnames:
                    if sheet_name.casefold() == key:
                        is_known = True
                        break
            except Exception:
                pass
            finally:
                if wb is not None:
                    try:
                        wb.close()
                    except Exception:
                        pass

        set_default_user(interaction.user.id, proper_ign)
        if is_known:
            await interaction.followup.send(f"Default username set to {proper_ign}.", ephemeral=True)
        else:
            await interaction.followup.send(f"Default username set to {proper_ign}. Note: {proper_ign} is not tracked yet; some commands may fail until you run /track.", ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"[ERROR] Failed to set default: {str(e)}", ephemeral=True)

@bot.tree.command(name="removedefault", description="Remove your default Minecraft username")
async def removedefault(interaction: discord.Interaction):
    if not interaction.response.is_done():
        await interaction.response.defer(ephemeral=True)
    
    removed = remove_default_user(interaction.user.id)
    if removed:
        await interaction.followup.send("Your default username has been removed.", ephemeral=True)
    else:
        await interaction.followup.send("You don't have a default username set.", ephemeral=True)

@bot.tree.command(name="prestige", description="Display a prestige prefix for any level")
@discord.app_commands.describe(
    level="The prestige level (e.g., 1964)",
    ign="Optional: Username to display after the prefix"
)
async def prestige(interaction: discord.Interaction, level: int, ign: str = None):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        # Validate level range
        if level < 0 or level > 10000:
            await interaction.followup.send("[ERROR] Level must be between 0 and 10000")
            return
        
        # Get prestige icon for this level
        icon = get_prestige_icon(level)
        
        # Build the colored prefix
        colored_prefix = format_prestige_ansi(level, icon)
        
        # Add IGN if provided
        if ign:
            display_text = f"{colored_prefix} {ign}"
        else:
            display_text = colored_prefix
        
        # Try to render as image if Pillow is available
        if Image is not None:
            try:
                base = (level // 100) * 100
                raw = PRESTIGE_RAW_PATTERNS.get(base)
                
                if raw:
                    # Parse the pattern and replace the level number
                    parts = _parse_raw_pattern(raw)
                    
                    # Build segments with the actual level
                    concat = ''.join(t for (_, t) in parts)
                    m = re.search(r"\d+", concat)
                    
                    segments = []
                    if m:
                        num_start, num_end = m.start(), m.end()
                        pos = 0
                        replaced = False
                        
                        for code, text in parts:
                            part_start = pos
                            part_end = pos + len(text)
                            pos = part_end
                            hexcol = MINECRAFT_CODE_TO_HEX.get(code.lower(), '#FFFFFF')
                            
                            if part_end <= num_start or part_start >= num_end:
                                segments.append((hexcol, text))
                                continue
                            
                            # Prefix before number
                            prefix_len = max(0, num_start - part_start)
                            if prefix_len > 0:
                                segments.append((hexcol, text[:prefix_len]))
                            
                            # Replace with actual level
                            if not replaced:
                                # Check if this is a rainbow prestige
                                rainbow_bases = {k for k, v in PRESTIGE_COLORS.items() if v is None}
                                if base in rainbow_bases:
                                    # Build rainbow colors
                                    colors_in_span = []
                                    pos2 = 0
                                    for c_code, c_text in parts:
                                        part_s = pos2
                                        part_e = pos2 + len(c_text)
                                        pos2 = part_e
                                        overlap_s = max(part_s, num_start)
                                        overlap_e = min(part_e, num_end)
                                        if overlap_e > overlap_s:
                                            hexcol_span = MINECRAFT_CODE_TO_HEX.get(c_code.lower(), '#FFFFFF')
                                            for _ in range(overlap_e - overlap_s):
                                                colors_in_span.append(hexcol_span)
                                    
                                    if not colors_in_span:
                                        RAINBOW_CODES = ['c', '6', 'e', 'a', 'b', 'd', '9', '3']
                                        colors_in_span = [MINECRAFT_CODE_TO_HEX.get(c, '#FFFFFF') for c in RAINBOW_CODES]
                                    
                                    # Apply colors to level digits
                                    for i, ch in enumerate(str(level)):
                                        col = colors_in_span[i % len(colors_in_span)]
                                        segments.append((col, ch))
                                else:
                                    segments.append((hexcol, str(level)))
                                replaced = True
                            
                            # Suffix after number
                            suffix_start_in_part = max(0, num_end - part_start)
                            if suffix_start_in_part < len(text):
                                segments.append((hexcol, text[suffix_start_in_part:]))
                    else:
                        # No number found, just use the pattern as-is with level prepended
                        segments = [(MINECRAFT_CODE_TO_HEX.get(parts[0][0], '#FFFFFF'), f"[{level}")]
                        segments.extend([(MINECRAFT_CODE_TO_HEX.get(code, '#FFFFFF'), text) for code, text in parts[1:]])
                    
                    # Add IGN if provided
                    if ign:
                        segments.append((MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF'), f" {ign}"))
                    
                    # Render to image
                    img_io = _render_text_segments_to_image(segments)
                    filename = f"prestige_{level}" + (f"_{ign}" if ign else "") + ".png"
                    await interaction.followup.send(file=discord.File(img_io, filename=filename))
                    return
            except Exception as e:
                # Fall back to ANSI if image rendering fails
                print(f"[WARNING] Image rendering failed: {e}")
        
        # Fallback: send as ANSI text
        await interaction.followup.send(f"```ansi\n{display_text}\n```")
        
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="instructions", description="Display bot usage instructions")
async def instructions(interaction: discord.Interaction):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        instructions_file = BOT_DIR / "instructions.txt"
        if not instructions_file.exists():
            await interaction.followup.send("[ERROR] Instructions file not found")
            return
        
        with open(instructions_file, "r", encoding="utf-8") as f:
            content = f.read()
        
        # Discord has a 2000 character limit for messages
        if len(content) > 1900:
            # Split into chunks if needed
            chunks = [content[i:i+1900] for i in range(0, len(content), 1900)]
            for chunk in chunks:
                await interaction.followup.send(chunk)
        else:
            await interaction.followup.send(content)
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="whatamirunningon", description="Creator-only: show public IP and bot file path")
async def whatamirunningon(interaction: discord.Interaction):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    # Only allow the creator (by ID) to run this
    allowed = is_admin(interaction.user)

    if not allowed:
        await interaction.followup.send("[ERROR] You are not authorized to run this command.", ephemeral=True)
        return

    # Try to get public IP, fall back to local hostname IP
    ip = None
    try:
        import urllib.request, json, ssl
        ctx = ssl.create_default_context()
        with urllib.request.urlopen("https://api.ipify.org?format=json", timeout=5, context=ctx) as resp:
            data = json.load(resp)
            ip = data.get("ip")
    except Exception:
        try:
            import socket
            ip = socket.gethostbyname(socket.gethostname())
        except Exception:
            ip = "unknown"

    path = str(BOT_DIR / "discord_bot.py")
    await interaction.followup.send(f"IP: {ip}\nPath: {path}", ephemeral=True)

@bot.tree.command(name="version", description="Show bot uptime and verify code version")
async def version(interaction: discord.Interaction):
    if not interaction.response.is_done():
        await interaction.response.defer(ephemeral=True)
    
    uptime = int(time.time() - START_TIME)
    hours, rem = divmod(uptime, 3600)
    minutes, seconds = divmod(rem, 60)
    
    await interaction.followup.send(f"Bot Uptime: {hours}h {minutes}m {seconds}s\nRunning from: {BOT_DIR}", ephemeral=True)


@bot.tree.command(name="refresh", description="Manually run batch snapshot update for all tracked users")
@discord.app_commands.describe(mode="One of: session, daily, yesterday, monthly, all, or all+session", ign="Optional: Minecraft IGN to refresh")
@discord.app_commands.choices(mode=[
    discord.app_commands.Choice(name="session", value="session"),
    discord.app_commands.Choice(name="daily", value="daily"),
    discord.app_commands.Choice(name="yesterday", value="yesterday"),
    discord.app_commands.Choice(name="monthly", value="monthly"),
    discord.app_commands.Choice(name="all (daily + yesterday + monthly)", value="all"),
    discord.app_commands.Choice(name="all+session (session + daily + yesterday + monthly)", value="all-session"),
])
async def refresh(interaction: discord.Interaction, mode: discord.app_commands.Choice[str], ign: str = None):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException) as e:
            print(f"[REFRESH] Failed to defer interaction: {e}")
            return
    try:
        # If an IGN was supplied, run per-user api_get.py with appropriate flags
        if ign:
            # Only allow creator or the Discord user who claimed the IGN
            allowed = is_admin(interaction.user)
            if not allowed and not is_user_authorized(interaction.user.id, ign):
                await interaction.followup.send(f"[ERROR] You are not authorized to refresh {ign}.", ephemeral=True)
                return

            # Map mode to api_get flags
            mode_map = {
                'session': ['-session'],
                'daily': ['-daily'],
                'yesterday': ['-yesterday'],
                'monthly': ['-monthly'],
                'all': ['-daily', '-yesterday', '-monthly'],
                'all-session': ['-session', '-daily', '-yesterday', '-monthly'],
            }
            flags = mode_map.get(mode.value, [])

            args = ['-ign', ign, *flags]

            # Use batch runner for potentially longer single-user operations
            result = await asyncio.to_thread(run_script_batch, "api_get.py", args)

            if result.returncode == 0:
                msg = f"Refresh completed for {ign} (schedule: {mode.name})"
            else:
                error_msg = result.stderr or result.stdout or "Unknown error"
                print(f"[REFRESH] Single-user refresh FAILED for {ign} - returncode: {result.returncode}")
                print(f"[REFRESH] stdout: {result.stdout[:500] if result.stdout else '(empty)'}")
                print(f"[REFRESH] stderr: {result.stderr[:500] if result.stderr else '(empty)'}")
                msg = f"Refresh failed for {ign}: {error_msg[:300]}"
        else:
            # Run batch_update.py with selected schedule (use extended timeout)
            def run_batch():
                return run_script_batch("batch_update.py", ["-schedule", mode.value])

            result = await asyncio.to_thread(run_batch)

            if result.returncode == 0:
                msg = f"Batch snapshot update completed for schedule: {mode.name}"
            else:
                error_msg = result.stderr or result.stdout or "Unknown error"
                print(f"[REFRESH] Batch update FAILED for schedule {mode.value} - returncode: {result.returncode}")
                print(f"[REFRESH] Full stdout:")
                print(result.stdout if result.stdout else "(empty)")
                print(f"[REFRESH] Full stderr:")
                print(result.stderr if result.stderr else "(empty)")
                msg = f"Batch update failed: {error_msg[:800]}"
        
        # Try to DM the invoking user directly
        try:
            await interaction.user.send(msg)
            try:
                await interaction.followup.send("Sent you a DM with the results.", ephemeral=True)
            except (discord.errors.NotFound, discord.errors.HTTPException):
                print(f"[REFRESH] Interaction expired, but DM was sent to {interaction.user.name}")
        except Exception as dm_error:
            # Fallback to ephemeral if DMs are closed
            print(f"[REFRESH] Failed to send DM: {dm_error}")
            try:
                await interaction.followup.send(msg, ephemeral=True)
            except (discord.errors.NotFound, discord.errors.HTTPException):
                print(f"[REFRESH] Interaction expired, couldn't send results to {interaction.user.name}")
    except subprocess.TimeoutExpired:
        try:
            await interaction.followup.send(f"[ERROR] Batch update timed out after 5 minutes. Try a smaller schedule (e.g., just 'daily' or 'session').", ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            print(f"[REFRESH] Timeout error but interaction expired")
    except Exception as e:
        print(f"[REFRESH] Exception: {e}")
        import traceback
        traceback.print_exc()
        try:
            await interaction.followup.send(f"[ERROR] {str(e)}", ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            print(f"[REFRESH] Exception but interaction expired")

@bot.tree.command(name="fixguilds", description="Admin: Force refresh all guild tags and colors")
async def fixguilds(interaction: discord.Interaction):
    if not is_admin(interaction.user):
        await interaction.response.send_message("❌ This command is only available to bot administrators.", ephemeral=True)
        return

    if not interaction.response.is_done():
        await interaction.response.defer(ephemeral=True)

    await interaction.followup.send("Starting guild data repair... This may take a while.", ephemeral=True)
    
    try:
        # Run the fix script
        result = await asyncio.to_thread(run_script, "fix_guilds.py", [])
        
        if result.returncode == 0:
            # Force cache refresh so the bot sees the new tags immediately
            await STATS_CACHE.refresh()
            await interaction.followup.send("✅ Guild data repair completed successfully!", ephemeral=True)
        else:
            err = result.stderr or result.stdout or "Unknown error"
            await interaction.followup.send(f"❌ Guild repair failed:\n```{sanitize_output(err[:1000])}```", ephemeral=True)
            
    except Exception as e:
        await interaction.followup.send(f"❌ Exception during guild repair: {str(e)}", ephemeral=True)

@bot.tree.command(name="stats", description="Get full player stats (Template.xlsx layout) with deltas")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def stats(interaction: discord.Interaction, ign: str = None):
    print(f"[DEBUG] /stats triggered for IGN: {ign} by user: {interaction.user.name} in guild: {interaction.guild.name if interaction.guild else 'DM'}")
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign
    
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException) as e:
            print(f"[DEBUG] Defer failed for {ign} in /stats: {e}")
            return

    try:
        # Fetch fresh stats
        print(f"[DEBUG] Running api_get.py for IGN: {ign} (/stats)")
        result = run_script("api_get.py", ["-ign", ign])
        print(f"[DEBUG] api_get.py returncode (/stats): {result.returncode}")
        print(f"[DEBUG] api_get.py stdout (/stats): {result.stdout if result.stdout else 'None'}")
        print(f"[DEBUG] api_get.py stderr (/stats): {result.stderr if result.stderr else 'None'}")

        if result.returncode != 0:
            if result.stderr and "429" in result.stderr:
                print(f"[DEBUG] Rate limited for {ign} (/stats), attempting to use existing data")
            else:
                error_msg = result.stderr or result.stdout or "Unknown error"
                await interaction.followup.send(f"[ERROR] Failed to fetch stats:\n```{error_msg[:500]}```")
                return

        # Optimistically update cache if we have JSON output
        try:
            if result.stdout:
                # Try to find the JSON object in the output (ignoring debug logs)
                json_data = None
                for line in reversed(result.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            json_data = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                
                if json_data and "processed_stats" in json_data and "username" in json_data:
                    await STATS_CACHE.update_cache_entry(json_data["username"], json_data["processed_stats"])
                    # Update ign to the proper case returned by API
                    ign = json_data["username"]
        except Exception as e:
            print(f"[WARNING] Failed to update cache from output: {e}")

        EXCEL_FILE = BOT_DIR / "stats.xlsx"
        if not EXCEL_FILE.exists():
            await interaction.followup.send("[ERROR] Excel file not found")
            return

        cache_data = await STATS_CACHE.get_data()
        
        # Find user in cache case-insensitively
        key = ign.casefold()
        user_data = None
        actual_ign = ign
        for name, data in cache_data.items():
            if name.casefold() == key:
                user_data = data
                actual_ign = name
                break
        
        if not user_data:
            await interaction.followup.send(f"[ERROR] Player sheet '{ign}' not found")
            return

        tracked_users = load_tracked_users()
        is_tracked = any(u.casefold() == actual_ign.casefold() for u in tracked_users)

        view = StatsFullView(user_data, actual_ign)
        embed, file = view.generate_full_image("all-time")

        if file:
            if is_tracked:
                await interaction.followup.send(view=view, file=file)
            else:
                message = f"{actual_ign} is not currently tracked. Only all-time stats are available.\nUse `/track ign:{actual_ign}` to start tracking and enable session/daily/monthly stats."
                await interaction.followup.send(content=message, file=file)
                bot.loop.create_task(cleanup_untracked_user_delayed(actual_ign, delay_seconds=60))
        else:
            if is_tracked:
                await interaction.followup.send(embed=embed, view=view)
            else:
                message = f"{actual_ign} is not currently tracked. Only all-time stats are available.\nUse `/track ign:{actual_ign}` to start tracking and enable session/daily/monthly stats."
                await interaction.followup.send(content=message, embed=embed)
                bot.loop.create_task(cleanup_untracked_user_delayed(actual_ign, delay_seconds=60))

    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)")
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="streak", description="View current win/kill streaks (approved users)")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def streak(interaction: discord.Interaction, ign: str = None):
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign

    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    try:
        # Ensure user is cached (fetches if needed)
        cached, actual_ign = await ensure_user_cached(ign)
        if not cached:
            await interaction.followup.send(f"[ERROR] Could not find or fetch data for {ign}")
            return

        cache_data = await STATS_CACHE.get_data()
        key = actual_ign.casefold()
        user_data = None
        for name, data in cache_data.items():
            if name.casefold() == key:
                user_data = data
                actual_ign = name
                break

        if not user_data:
            await interaction.followup.send(f"[ERROR] Player sheet '{actual_ign}' not found")
            return

        # Require user to be tracked in the main tracked list
        tracked_users = load_tracked_users()
        if not any(u.casefold() == actual_ign.casefold() for u in tracked_users):
            await interaction.followup.send(f"{actual_ign} is not currently tracked. Use `/track ign:{actual_ign}` first.")
            return

        streaks = load_tracked_streaks()
        entry_key = None
        for k in streaks.keys():
            if k.casefold() == actual_ign.casefold():
                entry_key = k
                break

        if entry_key:
            entry = streaks.get(entry_key, {})
            winstreak = int(entry.get("winstreak", 0))
            killstreak = int(entry.get("killstreak", 0))

            meta = user_data.get("meta", {})
            level = meta.get("level", 0)
            icon = meta.get("icon", "")
            ign_color = meta.get("ign_color")
            guild_tag = meta.get("guild_tag")
            guild_color = meta.get("guild_hex")

            if Image is not None:
                try:
                    img_io = create_streaks_image(actual_ign, level, icon, ign_color, guild_tag, guild_color, winstreak, killstreak)
                    filename = f"{actual_ign}_streaks.png"
                    await interaction.followup.send(file=discord.File(img_io, filename=filename))
                    return
                except Exception as e:
                    print(f"[STREAK] Failed to render streaks image: {e}")

            embed = discord.Embed(title=f"{actual_ign} Streaks")
            embed.add_field(name="Current Winstreak", value=f"```{winstreak}```", inline=True)
            embed.add_field(name="Current Killstreak", value=f"```{killstreak}```", inline=True)
            await interaction.followup.send(embed=embed)
        else:
            message = (
                f"{actual_ign} does not have streaks tracked. "
                f"Click the button to request to track {actual_ign}'s streaks."
            )
            view = StreakRequestView(actual_ign, interaction.user, user_data.get("stats", {}))
            await interaction.followup.send(content=message, view=view)

    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="killdistribution", description="View kill-type distribution as a pie chart")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def killdistribution(interaction: discord.Interaction, ign: str = None):
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    user_data = None
    try:
        result = run_script("api_get.py", ["-ign", ign])
        if result.returncode != 0 and not (result.stderr and "429" in result.stderr):
            error_msg = result.stderr or result.stdout or "Unknown error"
            await interaction.followup.send(f"[ERROR] Failed to fetch stats:\n```{error_msg[:500]}```")
            return

        # Optimistically update cache
        try:
            if result.stdout:
                data = json.loads(result.stdout)
                if "processed_stats" in data and "username" in data:
                    user_data = await STATS_CACHE.update_cache_entry(data["username"], data["processed_stats"])
                    ign = data["username"]
        except Exception as e:
            print(f"[WARNING] Failed to update cache from output: {e}")

        EXCEL_FILE = BOT_DIR / "stats.xlsx"
        if not EXCEL_FILE.exists():
            await interaction.followup.send("[ERROR] Excel file not found")
            return

        if not user_data:
            cache_data = await STATS_CACHE.get_data()
            key = ign.casefold()
            actual_ign = ign
            for name, data in cache_data.items():
                if name.casefold() == key:
                    user_data = data
                    actual_ign = name
                    break
        else:
            actual_ign = ign
        
        if not user_data:
            await interaction.followup.send(f"[ERROR] Player sheet '{ign}' not found")
            return

        view = DistributionView(user_data, actual_ign, mode="kill")
        embed, file = view.generate_distribution("all-time")

        if file:
            await interaction.followup.send(file=file, view=view)
        else:
            await interaction.followup.send(embed=embed, view=view)
    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)")
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="deathdistribution", description="View death-type distribution as a pie chart")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def deathdistribution(interaction: discord.Interaction, ign: str = None):
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    user_data = None
    try:
        result = run_script("api_get.py", ["-ign", ign])
        if result.returncode != 0 and not (result.stderr and "429" in result.stderr):
            error_msg = result.stderr or result.stdout or "Unknown error"
            await interaction.followup.send(f"[ERROR] Failed to fetch stats:\n```{error_msg[:500]}```")
            return

        # Optimistically update cache
        try:
            if result.stdout:
                data = json.loads(result.stdout)
                if "processed_stats" in data and "username" in data:
                    user_data = await STATS_CACHE.update_cache_entry(data["username"], data["processed_stats"])
                    ign = data["username"]
        except Exception as e:
            print(f"[WARNING] Failed to update cache from output: {e}")

        EXCEL_FILE = BOT_DIR / "stats.xlsx"
        if not EXCEL_FILE.exists():
            await interaction.followup.send("[ERROR] Excel file not found")
            return

        if not user_data:
            cache_data = await STATS_CACHE.get_data()
            key = ign.casefold()
            actual_ign = ign
            for name, data in cache_data.items():
                if name.casefold() == key:
                    user_data = data
                    actual_ign = name
                    break
        else:
            actual_ign = ign
        
        if not user_data:
            await interaction.followup.send(f"[ERROR] Player sheet '{ign}' not found")
            return

        view = DistributionView(user_data, actual_ign, mode="death")
        embed, file = view.generate_distribution("all-time")

        if file:
            await interaction.followup.send(file=file, view=view)
        else:
            await interaction.followup.send(embed=embed, view=view)
    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)")
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

@bot.tree.command(name="sheepwars", description="Get player stats with deltas")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def sheepwars(interaction: discord.Interaction, ign: str = None):
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign

    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    EXCEL_FILE = BOT_DIR / "stats.xlsx"
    if not EXCEL_FILE.exists():
        await interaction.followup.send("Stats file not found.")
        return

    # Check if user is tracked
    tracked_users = load_tracked_users()
    
    # Try to find in cache first
    cache_data = await STATS_CACHE.get_data()
    user_data = None
    key = ign.casefold()
    for name, data in cache_data.items():
        if name.casefold() == key:
            user_data = data
            ign = name
            break

    # If not in cache, try to fetch fresh stats (for untracked users)
    if not user_data:
        result = run_script("api_get.py", ["-ign", ign])
        if result.returncode == 0 and result.stdout:
            try:
                json_data = None
                for line in reversed(result.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            json_data = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                
                if json_data and "processed_stats" in json_data and "username" in json_data:
                    user_data = await STATS_CACHE.update_cache_entry(json_data["username"], json_data["processed_stats"])
                    ign = json_data["username"]
            except Exception as e:
                print(f"[WARNING] Failed to parse api_get output in sheepwars: {e}")
        elif result.returncode != 0:
            # If api_get failed, log it and potentially inform user
            print(f"[ERROR] api_get failed for {ign} in sheepwars: {result.stderr}")

    if not user_data:
        await interaction.followup.send("Player not found in database or API error.")
        return

    # Build all_data from cache
    all_data = {}
    stats = user_data.get("stats", {})
    
    # Map tab names to cache keys
    tab_map = {
        "all-time": "lifetime",
        "session": "session",
        "daily": "daily",
        "yesterday": "yesterday",
        "monthly": "monthly"
    }
    
    for tab, cache_key in tab_map.items():
        w = stats.get("wins", {}).get(cache_key, 0)
        l = stats.get("losses", {}).get(cache_key, 0)
        k = stats.get("kills", {}).get(cache_key, 0)
        d = stats.get("deaths", {}).get(cache_key, 0)
        p = stats.get("playtime", {}).get(cache_key, 0)
        
        all_data[tab] = {
            'wins': w, 'losses': l, 'kills': k, 'deaths': d, 'playtime': p,
            'wlr': w/l if l > 0 else w,
            'kdr': k/d if d > 0 else k
        }

    meta = user_data.get("meta", {})
    level = meta.get("level", 0)
    icon = meta.get("icon", "")
    ign_color = meta.get("ign_color")
    guild_tag = meta.get("guild_tag")
    guild_hex = meta.get("guild_hex")
    
    # Get real-time status
    # Parallelize status and skin fetching to reduce load time
    status_task = asyncio.to_thread(get_player_status, ign)
    skin_task = asyncio.to_thread(get_player_body, ign)
    (status_text, status_color), skin_image = await asyncio.gather(status_task, skin_task)
    
    view = StatsTabView(all_data, ign, int(level), icon, 
                        ign_color=ign_color, guild_tag=guild_tag, guild_hex=guild_hex,
                        status_text=status_text, status_color=status_color, skin_image=skin_image)
    
    # Check if tracked to determine response
    is_tracked = any(u.casefold() == ign.casefold() for u in tracked_users)
    
    file = view.generate_composite_image("all-time")
    if is_tracked:
        await interaction.followup.send(file=file, view=view)
    else:
        msg = f"{ign} is not currently tracked. Only all-time stats are available.\nUse `/track ign:{ign}` to start tracking and enable session/daily/monthly stats."
        await interaction.followup.send(content=msg, file=file)
        bot.loop.create_task(cleanup_untracked_user_delayed(ign, delay_seconds=60))

# Standalone leaderboard commands
@bot.tree.command(name="leaderboard", description="View player leaderboards")
@discord.app_commands.describe(metric="Choose a stat to rank players by")
@discord.app_commands.choices(metric=[
    discord.app_commands.Choice(name="Kills", value="kills"),
    discord.app_commands.Choice(name="Deaths", value="deaths"),
    discord.app_commands.Choice(name="K/D Ratio", value="kdr"),
    discord.app_commands.Choice(name="Wins", value="wins"),
    discord.app_commands.Choice(name="Losses", value="losses"),
    discord.app_commands.Choice(name="W/L Ratio", value="wlr"),
    discord.app_commands.Choice(name="Experience", value="experience"),
    discord.app_commands.Choice(name="Level", value="level"),
    discord.app_commands.Choice(name="Coins", value="coins"),
    discord.app_commands.Choice(name="Damage Dealt", value="damage_dealt"),
    discord.app_commands.Choice(name="Games Played", value="games_played"),
    discord.app_commands.Choice(name="Sheep Thrown", value="sheep_thrown"),
    discord.app_commands.Choice(name="Magic Wool Hit", value="magic_wool_hit"),
    discord.app_commands.Choice(name="Playtime", value="playtime"),
])
async def leaderboard(interaction: discord.Interaction, metric: discord.app_commands.Choice[str]):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    wb = None
    try:
        EXCEL_FILE = "stats.xlsx"
        if not os.path.exists(EXCEL_FILE):
            await interaction.followup.send("[ERROR] Excel file not found")
            return
        
        # Load data directly from Excel instead of cache
        processed_data = await asyncio.to_thread(_load_leaderboard_data_from_excel, metric.value)
        view = LeaderboardView(metric.value, processed_data)
        embed, file, _ = await asyncio.to_thread(view.generate_leaderboard_image, "lifetime", 0)
        if file:
            await interaction.followup.send(view=view, file=file)
        else:
            await interaction.followup.send(embed=embed, view=view)
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="kill-leaderboard", description="View kills leaderboard by type")
@discord.app_commands.describe(metric="Choose which kill type to rank by")
@discord.app_commands.choices(metric=[
    discord.app_commands.Choice(name="Total Kills", value="kills"),
    discord.app_commands.Choice(name="Void Kills", value="kills_void"),
    discord.app_commands.Choice(name="Explosive Kills", value="kills_explosive"),
    discord.app_commands.Choice(name="Melee Kills", value="kills_melee"),
    discord.app_commands.Choice(name="Bow Kills", value="kills_bow"),
])
async def kill_leaderboard(interaction: discord.Interaction, metric: discord.app_commands.Choice[str]):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    wb = None
    try:
        EXCEL_FILE = "stats.xlsx"
        if not os.path.exists(EXCEL_FILE):
            await interaction.followup.send("[ERROR] Excel file not found")
            return
        
        # Load data directly from Excel instead of cache
        processed_data = await asyncio.to_thread(_load_leaderboard_data_from_excel, metric.value)
        view = LeaderboardView(metric.value, processed_data)
        embed, file, _ = await asyncio.to_thread(view.generate_leaderboard_image, "lifetime", 0)
        if file:
            await interaction.followup.send(file=file, view=view)
        else:
            await interaction.followup.send(embed=embed, view=view)
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="death-leaderboard", description="View deaths leaderboard by type")
@discord.app_commands.describe(metric="Choose which death type to rank by")
@discord.app_commands.choices(metric=[
    discord.app_commands.Choice(name="Total Deaths", value="deaths"),
    discord.app_commands.Choice(name="Void Deaths", value="deaths_void"),
    discord.app_commands.Choice(name="Explosive Deaths", value="deaths_explosive"),
    discord.app_commands.Choice(name="Melee Deaths", value="deaths_melee"),
    discord.app_commands.Choice(name="Bow Deaths", value="deaths_bow"),
])
async def death_leaderboard(interaction: discord.Interaction, metric: discord.app_commands.Choice[str]):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    wb = None
    try:
        EXCEL_FILE = "stats.xlsx"
        if not os.path.exists(EXCEL_FILE):
            await interaction.followup.send("[ERROR] Excel file not found")
            return
        
        # Load data directly from Excel instead of cache
        processed_data = await asyncio.to_thread(_load_leaderboard_data_from_excel, metric.value)
        view = LeaderboardView(metric.value, processed_data)
        embed, file, _ = await asyncio.to_thread(view.generate_leaderboard_image, "lifetime", 0)
        if file:
            await interaction.followup.send(file=file, view=view)
        else:
            await interaction.followup.send(embed=embed, view=view)
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="ratio-leaderboard", description="View ratio-based leaderboard")
@discord.app_commands.describe(metric="Choose which ratio to rank by")
@discord.app_commands.choices(metric=[
    discord.app_commands.Choice(name="Win/Loss", value="wl_ratio"),
    discord.app_commands.Choice(name="Kill/Death", value="kd_ratio"),
    discord.app_commands.Choice(name="Kill/Game", value="kills_per_game"),
    discord.app_commands.Choice(name="Kill/Win", value="kills_per_win"),
    discord.app_commands.Choice(name="Kill/Hour", value="kills_per_hour"),
    discord.app_commands.Choice(name="Damage/Game", value="damage_per_game"),
    discord.app_commands.Choice(name="Damage/Sheep", value="damage_per_sheep"),
    discord.app_commands.Choice(name="Wools/Game", value="wools_per_game"),
    discord.app_commands.Choice(name="Sheeps/Game", value="sheeps_per_game"),
    discord.app_commands.Choice(name="Void Kill/Death", value="void_kd_ratio"),
    discord.app_commands.Choice(name="Explosive Kill/Death", value="explosive_kd_ratio"),
    discord.app_commands.Choice(name="Bow Kill/Death", value="bow_kd_ratio"),
    discord.app_commands.Choice(name="Melee Kill/Death", value="melee_kd_ratio"),
    discord.app_commands.Choice(name="Wins/Hour", value="wins_per_hour"),
    discord.app_commands.Choice(name="EXP/Hour", value="exp_per_hour"),
    discord.app_commands.Choice(name="EXP/Game", value="exp_per_game"),
])
async def ratio_leaderboard(interaction: discord.Interaction, metric: discord.app_commands.Choice[str]):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    wb = None
    try:
        EXCEL_FILE = "stats.xlsx"
        if not os.path.exists(EXCEL_FILE):
            await interaction.followup.send("[ERROR] Excel file not found")
            return
        
        # Load data directly from Excel instead of cache
        processed_data = await asyncio.to_thread(_load_ratio_leaderboard_data_from_excel, metric.value)
        view = RatioLeaderboardView(metric.value, processed_data)
        embed, file, _ = await asyncio.to_thread(view.generate_leaderboard_image, "lifetime", 0)
        if file:
            await interaction.followup.send(file=file, view=view)
        else:
            await interaction.followup.send(embed=embed, view=view)
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="prestiges", description="List all prestige prefixes with their colors")
async def prestiges(interaction: discord.Interaction):
    # Defer in case composing takes a moment
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    try:
        # Use image rendering if Pillow is available
        if Image is not None:
            try:
                combined = render_all_prestiges_combined()
                await interaction.followup.send(file=discord.File(combined, filename="Wool Games prestiges 0-4000.png"))
            except Exception:
                # If combining fails, fall back to sending individual images
                for base in sorted(PRESTIGE_RAW_PATTERNS.keys()):
                    end_display = base + 99
                    try:
                        imgio = render_prestige_range_image(base, end_display)
                        fname = f"prestige_{base}.png"
                        await interaction.followup.send(file=discord.File(imgio, filename=fname))
                    except Exception:
                        start_str = format_prestige_ansi(base, '')
                        end_str = format_prestige_ansi(end_display, '')
                        await interaction.followup.send(f"{start_str} - {end_str}")
        else:
            # Pillow not installed; fallback to ANSI list
            lines = []
            for base in sorted(PRESTIGE_RAW_PATTERNS.keys()):
                start_str = format_prestige_ansi(base, '')
                end_str = format_prestige_ansi(base + 99, '')
                lines.append(f"{start_str} - {end_str}")
            await _send_paged_ansi_followups(interaction, lines, block='ansi')
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

@bot.tree.command(name="stopbot", description="Gracefully shutdown the bot (admin only)")
async def stopbot(interaction: discord.Interaction):
    if not is_admin(interaction.user):
        await interaction.response.send_message("❌ This command is only available to bot administrators.", ephemeral=True)
        return
    
    await interaction.response.send_message("🛑 Shutting down bot gracefully... Please wait for all tasks to complete.", ephemeral=True)
    print(f"[SHUTDOWN] Bot shutdown initiated by {interaction.user.name} ({interaction.user.id})")
    print(f"[SHUTDOWN] Waiting for background tasks to complete...")
    
    # Cancel background tasks gracefully
    if hasattr(bot, 'background_tasks'):
        for task in bot.background_tasks:
            if not task.done():
                task.cancel()
        
        # Wait for tasks to complete their cleanup with timeout
        results = await asyncio.gather(*bot.background_tasks, return_exceptions=True)
        for i, result in enumerate(results):
            if isinstance(result, asyncio.CancelledError):
                print(f"[SHUTDOWN] Background task {i+1} cancelled successfully")
            elif isinstance(result, Exception):
                print(f"[SHUTDOWN] Background task {i+1} raised exception: {result}")
            else:
                print(f"[SHUTDOWN] Background task {i+1} completed normally")
    
    # Wait for any pending Discord operations
    print(f"[SHUTDOWN] Waiting for pending operations...")
    await asyncio.sleep(2)  # Give time for any pending messages/edits
    
    print(f"[SHUTDOWN] All cleanup complete, closing bot...")
    await bot.close()

# Run bot
if __name__ == "__main__":
    bot.run(DISCORD_TOKEN) 