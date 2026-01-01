#!/usr/bin/env python3
"""
Complete migration script for VM deployment.
Migrates ALL data from Excel and JSON files to SQLite database.

Run this ONCE on the VM to migrate from the old file-based system to the new database.
"""

import json
import sys
from pathlib import Path
import openpyxl
import db_helper

SCRIPT_DIR = Path(__file__).parent


def migrate_excel_stats(excel_path):
    """Migrate stats.xlsx to database"""
    print("\n[1/5] Migrating stats.xlsx...")
    
    if not excel_path.exists():
        print(f"  ⚠️  SKIP: {excel_path} not found")
        return 0
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    total_stats = 0
    user_count = 0
    
    # Get database connection for batch insert
    with db_helper.get_db_connection() as conn:
        cursor = conn.cursor()
        
        for sheet_name in wb.sheetnames:
            if sheet_name == "Template":
                continue
            
            sheet = wb[sheet_name]
            username = sheet_name
            user_count += 1
            
            # Extract metadata (rows 2-6)
            level = sheet.cell(2, 2).value or 0
            icon = sheet.cell(3, 2).value or ''
            ign_color = sheet.cell(4, 2).value
            guild_tag = sheet.cell(5, 2).value
            guild_hex = sheet.cell(6, 2).value
            
            # Insert user metadata
            cursor.execute('''
                INSERT OR REPLACE INTO user_meta (username, level, icon, ign_color, guild_tag, guild_hex)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                username,
                int(level) if isinstance(level, (int, float)) else 0,
                str(icon) if icon else '',
                str(ign_color) if ign_color else None,
                str(guild_tag) if guild_tag else None,
                str(guild_hex) if guild_hex else None
            ))
            
            # Extract stats (starting from row 2)
            for row_idx in range(2, sheet.max_row + 1):
                stat_name = sheet.cell(row_idx, 1).value
                if not stat_name:
                    continue
                
                stat_name = str(stat_name).strip().lower()
                
                lifetime = float(sheet.cell(row_idx, 2).value or 0)
                session = float(sheet.cell(row_idx, 4).value or 0)
                daily = float(sheet.cell(row_idx, 6).value or 0)
                yesterday = float(sheet.cell(row_idx, 8).value or 0)
                monthly = float(sheet.cell(row_idx, 10).value or 0)
                
                # Insert stat directly
                cursor.execute('''
                    INSERT OR REPLACE INTO user_stats 
                    (username, stat_name, lifetime, session, daily, yesterday, monthly)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (username, stat_name, lifetime, session, daily, yesterday, monthly))
                total_stats += 1
        
        conn.commit()
    
    wb.close()
    print(f"  ✅ Migrated stats for {user_count} users, {total_stats} total stat records")
    return total_stats


def migrate_user_links():
    """Migrate user_links.json"""
    print("\n[2/5] Migrating user_links.json...")
    
    file_path = SCRIPT_DIR / "user_links.json"
    if not file_path.exists():
        print(f"  ⚠️  SKIP: {file_path.name} not found")
        return 0
    
    with open(file_path, 'r') as f:
        data = json.load(f)
    
    for username, discord_id in data.items():
        db_helper.set_discord_link(username, discord_id)
    
    print(f"  ✅ Migrated {len(data)} user links")
    return len(data)


def migrate_default_users():
    """Migrate default_users.json"""
    print("\n[3/5] Migrating default_users.json...")
    
    file_path = SCRIPT_DIR / "default_users.json"
    if not file_path.exists():
        print(f"  ⚠️  SKIP: {file_path.name} not found")
        return 0
    
    with open(file_path, 'r') as f:
        data = json.load(f)
    
    for discord_id, username in data.items():
        db_helper.set_default_username(discord_id, username)
    
    print(f"  ✅ Migrated {len(data)} default users")
    return len(data)


def migrate_tracked_streaks():
    """Migrate tracked_streaks.json"""
    print("\n[4/5] Migrating tracked_streaks.json...")
    
    file_path = SCRIPT_DIR / "tracked_streaks.json"
    if not file_path.exists():
        print(f"  ⚠️  SKIP: {file_path.name} not found")
        return 0
    
    with open(file_path, 'r') as f:
        data = json.load(f)
    
    for username, streak_data in data.items():
        db_helper.update_tracked_streaks(username, streak_data)
    
    print(f"  ✅ Migrated {len(data)} tracked streaks")
    return len(data)


def migrate_tracked_users():
    """Migrate tracked_users.txt"""
    print("\n[5/5] Migrating tracked_users.txt...")
    
    file_path = SCRIPT_DIR / "tracked_users.txt"
    if not file_path.exists():
        print(f"  ⚠️  SKIP: {file_path.name} not found")
        return 0
    
    with open(file_path, 'r', encoding='utf-8') as f:
        usernames = [line.strip() for line in f if line.strip()]
    
    for username in usernames:
        db_helper.add_tracked_user(username)
    
    print(f"  ✅ Migrated {len(usernames)} tracked users")
    return len(usernames)


def verify_migration():
    """Verify the migration was successful"""
    print("\n" + "=" * 60)
    print("VERIFICATION")
    print("=" * 60)
    
    users = len(db_helper.get_all_usernames())
    links = len(db_helper.get_all_user_links())
    defaults = len(db_helper.get_all_default_users())
    streaks = len(db_helper.get_all_tracked_streaks())
    tracked = len(db_helper.get_tracked_users())
    
    print(f"✅ Database contains:")
    print(f"   - {users} users with stats")
    print(f"   - {links} user links (username → Discord ID)")
    print(f"   - {defaults} default users (Discord ID → username)")
    print(f"   - {streaks} tracked streaks")
    print(f"   - {tracked} tracked users")
    
    # Test a sample query
    if users > 0:
        sample_username = db_helper.get_all_usernames()[0]
        stats = db_helper.get_user_stats_with_deltas(sample_username)
        if stats and len(stats) > 0:
            sample_stat = list(stats.keys())[0]
            print(f"\n✅ Sample query successful:")
            print(f"   User: {sample_username}")
            print(f"   Stat: {sample_stat}")
            print(f"   Lifetime: {stats[sample_stat]['lifetime']}")


def main():
    print("=" * 60)
    print("COMPLETE MIGRATION TO DATABASE")
    print("=" * 60)
    print("\nThis will migrate ALL data to stats.db:")
    print("  • stats.xlsx → user_stats & user_meta tables")
    print("  • user_links.json → user_links table")
    print("  • default_users.json → default_users table")
    print("  • tracked_streaks.json → tracked_streaks table")
    print("  • tracked_users.txt → tracked_users table")
    
    db_path = SCRIPT_DIR / "stats.db"
    if db_path.exists():
        response = input(f"\n⚠️  WARNING: {db_path} already exists. Overwrite? (yes/no): ")
        if response.lower() != 'yes':
            print("❌ Migration cancelled")
            sys.exit(0)
        db_path.unlink()
        print(f"🗑️  Deleted existing {db_path}")
    
    print("\n[INIT] Creating database schema...")
    db_helper.init_database()
    print("  ✅ Database schema created")
    
    # Run all migrations
    excel_path = SCRIPT_DIR / "stats.xlsx"
    total = 0
    total += migrate_excel_stats(excel_path)
    total += migrate_user_links()
    total += migrate_default_users()
    total += migrate_tracked_streaks()
    total += migrate_tracked_users()
    
    verify_migration()
    
    print("\n" + "=" * 60)
    print(f"✅ SUCCESS: Migrated {total} total records")
    print("=" * 60)
    print("\nNext steps:")
    print("  1. Backup your old files (stats.xlsx, *.json, tracked_users.txt)")
    print("  2. Test the bot to ensure everything works")
    print("  3. Once verified, you can delete the old files")
    print(f"  4. Keep {db_path} as your single source of truth")
    print()


if __name__ == "__main__":
    main()
