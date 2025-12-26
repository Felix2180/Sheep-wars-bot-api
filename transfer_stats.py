#!/usr/bin/env python3
"""
Transfer stats data from sheep_wars_stats.xlsx to stats.xlsx.
Matches sheets by name (case-insensitive) and transfers specific cell values.
"""

from openpyxl import load_workbook
import os

# Cell mappings: source cell -> destination cell
CELL_MAPPINGS = {
    'E3': 'D15',
    'E4': 'D7',
    'E6': 'D14',
    'E7': 'D10',
    'E12': 'F15',
    'E13': 'F7',
    'E15': 'F14',
    'E16': 'F10',
    'E30': 'J15',
    'E31': 'J7',
    'E33': 'J14',
    'E34': 'J10'
}

def transfer_stats(source_file='sheep_wars_stats.xlsx', dest_file='stats.xlsx'):
    """
    Transfer stats from source spreadsheet to destination spreadsheet.
    
    Args:
        source_file: Path to sheep_wars_stats.xlsx
        dest_file: Path to stats.xlsx
    """
    
    # Check if files exist
    if not os.path.exists(source_file):
        print(f"Error: Source file '{source_file}' not found!")
        return False
    
    if not os.path.exists(dest_file):
        print(f"Error: Destination file '{dest_file}' not found!")
        return False
    
    try:
        # Load both workbooks
        print(f"Loading {source_file}...")
        source_wb = load_workbook(source_file, data_only=True)
        
        print(f"Loading {dest_file}...")
        dest_wb = load_workbook(dest_file)
        
        # Create a mapping of lowercase sheet names to actual sheet objects
        dest_sheets = {sheet.title.lower(): sheet for sheet in dest_wb.worksheets}
        
        # Track processed sheets
        processed = 0
        skipped = 0
        
        # Iterate through source sheets
        for source_sheet in source_wb.worksheets:
            source_name = source_sheet.title
            source_name_lower = source_name.lower()
            
            # Find matching destination sheet (case-insensitive)
            if source_name_lower in dest_sheets:
                dest_sheet = dest_sheets[source_name_lower]
                
                print(f"\nProcessing sheet: '{source_name}' -> '{dest_sheet.title}'")
                
                # Transfer each mapped cell
                transfers = 0
                for source_cell, dest_cell in CELL_MAPPINGS.items():
                    source_value = source_sheet[source_cell].value
                    old_value = dest_sheet[dest_cell].value
                    
                    dest_sheet[dest_cell].value = source_value
                    
                    print(f"  {source_cell} ({source_value}) -> {dest_cell} (was: {old_value})")
                    transfers += 1
                
                processed += 1
                print(f"  ✓ Transferred {transfers} cells")
            else:
                print(f"\nSkipping sheet '{source_name}' - no matching sheet in destination")
                skipped += 1
        
        # Save the destination workbook
        print(f"\nSaving changes to {dest_file}...")
        dest_wb.save(dest_file)
        
        print(f"\n{'='*60}")
        print(f"Transfer complete!")
        print(f"Processed: {processed} sheet(s)")
        print(f"Skipped: {skipped} sheet(s)")
        print(f"{'='*60}")
        
        return True
        
    except Exception as e:
        print(f"Error during transfer: {e}")
        return False
    
    finally:
        # Close workbooks
        if 'source_wb' in locals():
            source_wb.close()
        if 'dest_wb' in locals():
            dest_wb.close()

if __name__ == "__main__":
    transfer_stats()
