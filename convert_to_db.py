#!/usr/bin/env python3
"""
Convert stats.xlsx to stats.db SQLite database.
This script extracts all data from the Excel file and migrates it to SQLite.
"""

import sqlite3
import openpyxl
from pathlib import Path
import sys
import argparse
import json

# Import schema initialization from db_helper
from db_helper import init_database

EXCEL_FILE = Path(__file__).parent / "stats.xlsx"
DB_FILE = Path(__file__).parent / "stats.db"

def extract_excel_data(excel_path):
    """Extract all data from stats.xlsx."""
    if not excel_path.exists():
        print(f"[ERROR] Excel file not found: {excel_path}")
        sys.exit(1)
    
    print(f"[EXCEL] Loading {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    all_users_data = {}
    
    for sheet_name in wb.sheetnames:
        if sheet_name == "Template":
            continue
        
        sheet = wb[sheet_name]
        username = sheet_name
        
        # Initialize user data
        user_data = {
            'stats': {},
            'meta': {}
        }
        
        # Extract metadata from first few rows
        # Row 1: Username (A1), Level value (B1)
        # Row 2: Level label (A2), Level value with icon (B2) 
        # Row 3: Icon (A3), Icon value (B3)
        # Row 4: IGN Color (A4), Color value (B4)
        # Row 5: Guild Tag (A5), Tag value (B5)
        # Row 6: Guild Hex (A6), Hex value (B6)
        
        try:
            level_val = sheet.cell(row=2, column=2).value
            if level_val and isinstance(level_val, (int, float)):
                user_data['meta']['level'] = int(level_val)
            else:
                user_data['meta']['level'] = 0
                
            icon_val = sheet.cell(row=3, column=2).value
            user_data['meta']['icon'] = str(icon_val) if icon_val else ''
            
            color_val = sheet.cell(row=4, column=2).value
            user_data['meta']['ign_color'] = str(color_val) if color_val else None
            
            guild_tag = sheet.cell(row=5, column=2).value
            user_data['meta']['guild_tag'] = str(guild_tag) if guild_tag else None
            
            guild_hex = sheet.cell(row=6, column=2).value
            user_data['meta']['guild_hex'] = str(guild_hex) if guild_hex else None
        except Exception as e:
            print(f"[WARN] Error extracting metadata for {username}: {e}")
        
        # Extract stats starting from row 2 (row 1 has headers)
        # Column A: Stat name
        # Column B: Lifetime value
        # Column C: Session Delta
        # Column D: Session value (snapshot)
        # Column E: Daily Delta
        # Column F: Daily value (snapshot)
        # Column G: Yesterday Delta
        # Column H: Yesterday value (snapshot)
        # Column I: Monthly Delta
        # Column J: Monthly value (snapshot)
        
        for row_idx in range(2, sheet.max_row + 1):
            stat_name_cell = sheet.cell(row=row_idx, column=1).value
            
            if not stat_name_cell:
                continue
            
            stat_name = str(stat_name_cell).strip()
            if not stat_name:
                continue
            
            # Extract values for each period
            lifetime_val = sheet.cell(row=row_idx, column=2).value  # Column B
            session_val = sheet.cell(row=row_idx, column=4).value   # Column D
            daily_val = sheet.cell(row=row_idx, column=6).value     # Column F
            yesterday_val = sheet.cell(row=row_idx, column=8).value # Column H
            monthly_val = sheet.cell(row=row_idx, column=10).value  # Column J
            
            # Convert to float, default to 0 if None or invalid
            def safe_float(val):
                try:
                    return float(val) if val is not None else 0.0
                except (ValueError, TypeError):
                    return 0.0
            
            user_data['stats'][stat_name] = {
                'lifetime': safe_float(lifetime_val),
                'session': safe_float(session_val),
                'daily': safe_float(daily_val),
                'yesterday': safe_float(yesterday_val),
                'monthly': safe_float(monthly_val)
            }
        
        all_users_data[username] = user_data
        print(f"[EXCEL] Extracted data for {username}: {len(user_data['stats'])} stats")
    
    wb.close()
    print(f"[EXCEL] Extraction complete. Total users: {len(all_users_data)}")
    return all_users_data

def insert_data_to_db(conn, all_users_data):
    """Insert all extracted data into the database."""
    cursor = conn.cursor()
    
    stats_inserted = 0
    users_inserted = 0
    
    for username, user_data in all_users_data.items():
        # Insert metadata
        meta = user_data['meta']
        cursor.execute('''
            INSERT OR REPLACE INTO user_meta (username, level, icon, ign_color, guild_tag, guild_hex)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            username,
            meta.get('level', 0),
            meta.get('icon', ''),
            meta.get('ign_color'),
            meta.get('guild_tag'),
            meta.get('guild_hex')
        ))
        users_inserted += 1
        
        # Insert stats
        for stat_name, periods in user_data['stats'].items():
            # Make stat_name lowercase for consistency
            stat_name_lower = stat_name.lower()
            cursor.execute('''
                INSERT OR REPLACE INTO user_stats (username, stat_name, lifetime, session, daily, yesterday, monthly)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                username,
                stat_name_lower,
                periods['lifetime'],
                periods['session'],
                periods['daily'],
                periods['yesterday'],
                periods['monthly']
            ))
            stats_inserted += 1
    
    conn.commit()
    print(f"[DB] Inserted {users_inserted} users and {stats_inserted} stat records")

def migrate_tracked_users(conn):
    """Migrate tracked users from text file to database."""
    tracked_file = Path(__file__).parent / "tracked_users.txt"
    if tracked_file.exists():
        print(f"[MIGRATE] Found {tracked_file}, migrating to database...")
        try:
            with open(tracked_file, 'r', encoding='utf-8') as f:
                users = [line.strip() for line in f if line.strip()]
            
            cursor = conn.cursor()
            count = 0
            for user in users:
                cursor.execute('INSERT OR IGNORE INTO tracked_users (username) VALUES (?)', (user,))
                if cursor.rowcount > 0:
                    count += 1
            conn.commit()
            print(f"[MIGRATE] Migrated {count} tracked users.")
        except Exception as e:
            print(f"[ERROR] Failed to migrate tracked users: {e}")

def migrate_json_data(conn):
    """Migrate data from legacy JSON files (user_links, default_users, tracked_streaks)."""
    base_dir = Path(__file__).parent
    
    # Migrate user_links.json
    links_file = base_dir / "user_links.json"
    if links_file.exists():
        print(f"[MIGRATE] Found {links_file}, migrating...")
        try:
            with open(links_file, 'r', encoding='utf-8') as f:
                links = json.load(f)
            
            cursor = conn.cursor()
            count = 0
            for username, discord_id in links.items():
                cursor.execute('''
                    INSERT OR REPLACE INTO user_links (username, discord_id)
                    VALUES (?, ?)
                ''', (username, str(discord_id)))
                count += 1
            conn.commit()
            print(f"[MIGRATE] Migrated {count} user links.")
        except Exception as e:
            print(f"[ERROR] Failed to migrate user links: {e}")

    # Migrate default_users.json
    defaults_file = base_dir / "default_users.json"
    if defaults_file.exists():
        print(f"[MIGRATE] Found {defaults_file}, migrating...")
        try:
            with open(defaults_file, 'r', encoding='utf-8') as f:
                defaults = json.load(f)
            
            cursor = conn.cursor()
            count = 0
            for discord_id, username in defaults.items():
                cursor.execute('''
                    INSERT OR REPLACE INTO default_users (discord_id, username)
                    VALUES (?, ?)
                ''', (str(discord_id), username))
                count += 1
            conn.commit()
            print(f"[MIGRATE] Migrated {count} default users.")
        except Exception as e:
            print(f"[ERROR] Failed to migrate default users: {e}")

    # Migrate tracked_streaks.json
    streaks_file = base_dir / "tracked_streaks.json"
    if streaks_file.exists():
        print(f"[MIGRATE] Found {streaks_file}, migrating...")
        try:
            with open(streaks_file, 'r', encoding='utf-8') as f:
                streaks = json.load(f)
            
            cursor = conn.cursor()
            count = 0
            for username, data in streaks.items():
                cursor.execute('''
                    INSERT OR REPLACE INTO tracked_streaks 
                    (username, winstreak, killstreak, last_wins, last_losses, last_kills, last_deaths)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    username,
                    data.get('winstreak', 0),
                    data.get('killstreak', 0),
                    data.get('last_wins', 0),
                    data.get('last_losses', 0),
                    data.get('last_kills', 0),
                    data.get('last_deaths', 0)
                ))
                count += 1
            conn.commit()
            print(f"[MIGRATE] Migrated {count} streak records.")
        except Exception as e:
            print(f"[ERROR] Failed to migrate streaks: {e}")

def verify_conversion(conn, all_users_data):
    """Verify that all data was converted correctly."""
    cursor = conn.cursor()
    
    print("\n[VERIFY] Checking data integrity...")
    
    # Check user count
    cursor.execute('SELECT COUNT(DISTINCT username) FROM user_stats')
    db_user_count = cursor.fetchone()[0]
    excel_user_count = len(all_users_data)
    
    print(f"  Users: Excel={excel_user_count}, DB={db_user_count} {'✓' if db_user_count == excel_user_count else '✗'}")
    
    # Sample check: verify a few random stats
    errors = 0
    for username in list(all_users_data.keys())[:3]:  # Check first 3 users
        for stat_name in list(all_users_data[username]['stats'].keys())[:3]:  # Check first 3 stats
            cursor.execute('''
                SELECT lifetime, session, daily, yesterday, monthly 
                FROM user_stats 
                WHERE username = ? AND stat_name = ?
            ''', (username, stat_name))
            
            result = cursor.fetchone()
            if result:
                excel_data = all_users_data[username]['stats'][stat_name]
                db_data = {'lifetime': result[0], 'session': result[1], 'daily': result[2], 
                          'yesterday': result[3], 'monthly': result[4]}
                
                if excel_data != db_data:
                    print(f"  [WARN] Mismatch for {username}.{stat_name}")
                    errors += 1
    
    if errors == 0:
        print("  Sample verification: All checks passed ✓")
    else:
        print(f"  Sample verification: {errors} errors found ✗")
    
    print("\n[VERIFY] Conversion complete!")

def main():
    """Main conversion process."""
    print("=" * 60)
    print("Excel to SQLite Conversion Tool")
    print("=" * 60)
    
    parser = argparse.ArgumentParser(description="Convert Excel stats to SQLite database")
    parser.add_argument("--force", action="store_true", help="Overwrite existing database without prompt")
    args = parser.parse_args()
    
    # Check if database already exists
    if DB_FILE.exists():
        if not args.force:
            response = input(f"\n[WARN] {DB_FILE} already exists. Update/Merge? (yes/no): ")
            if response.lower() != 'yes':
                print("[ABORT] Conversion cancelled.")
                sys.exit(0)
        print(f"[DB] Using existing {DB_FILE}")
    
    # Create database and schema
    print("[DB] Initializing database schema...")
    init_database(DB_FILE)
    
    conn = sqlite3.connect(DB_FILE)
    
    # Extract data from Excel
    if EXCEL_FILE.exists():
        all_users_data = extract_excel_data(EXCEL_FILE)
        insert_data_to_db(conn, all_users_data)
        verify_conversion(conn, all_users_data)
    else:
        print(f"[WARN] {EXCEL_FILE} not found, skipping Excel import.")
    
    # Migrate tracked users
    migrate_tracked_users(conn)
    
    # Migrate JSON data
    migrate_json_data(conn)
    
    # Close connection
    conn.close()
    
    print("\n" + "=" * 60)
    print(f"SUCCESS: Data converted to {DB_FILE}")
    print("=" * 60)
    print("\nNext steps:")
    print("  1. Keep stats.xlsx as a backup")
    print("  2. Update discord_bot.py and api_get.py to use SQLite")
    print("  3. Test the bot with the new database")

if __name__ == "__main__":
    main()
