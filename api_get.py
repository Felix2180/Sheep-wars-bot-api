import os
import argparse
import time
import json
from pathlib import Path
from typing import Dict, Optional
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent.absolute()
EXCEL_FILE = str(SCRIPT_DIR / "stats.xlsx")
STAT_ORDER = ["Kills", "Deaths", "K/D", "Wins", "Losses", "W/L"]

def read_api_key_file() -> Optional[str]:
    """Read API key from API_KEY.txt next to the script, if present."""
    key_path = SCRIPT_DIR / "API_KEY.txt"
    if key_path.exists():
        try:
            content = key_path.read_text(encoding="utf-8").strip()
            if content:
                return content
        except Exception:
            # ignore read errors and fall back to other sources
            pass
    return None

# -------- Wool Games level calculation --------

def experience_to_level(exp: float) -> int:
    """Calculate Wool Games level from experience with prestige scaling.
    
    At each prestige (0, 100, 200, 300, etc), the XP requirements reset:
    - Level X+0 to X+1: 1000 XP
    - Level X+1 to X+2: 2000 XP
    - Level X+2 to X+3: 3000 XP
    - Level X+3 to X+4: 4000 XP
    - Level X+4 to X+5: 5000 XP
    - Level X+5 to X+100: 5000 XP each (95 levels)
    
    Total XP per prestige (100 levels): 1000+2000+3000+4000+5000 + 95*5000 = 490000
    """
    if exp <= 0:
        return 0
    
    XP_PER_PRESTIGE = 490000
    prestige_count = int(exp / XP_PER_PRESTIGE)
    remaining_xp = exp - (prestige_count * XP_PER_PRESTIGE)
    
    # Calculate level within current prestige (0-99)
    if remaining_xp < 1000:
        level_in_prestige = 0
    elif remaining_xp < 3000:  # 1000 + 2000
        level_in_prestige = 1
    elif remaining_xp < 6000:  # 1000 + 2000 + 3000
        level_in_prestige = 2
    elif remaining_xp < 10000:  # 1000 + 2000 + 3000 + 4000
        level_in_prestige = 3
    elif remaining_xp < 15000:  # 1000 + 2000 + 3000 + 4000 + 5000
        level_in_prestige = 4
    else:
        # Level 5+ in prestige: 5000 XP each
        remaining_after_first_5 = remaining_xp - 15000
        # Use floor division so partial progress doesn't round up to the next level
        level_in_prestige = 5 + int(remaining_after_first_5 // 5000)
    
    # Convert to 1-based display level (Hypixel shows levels starting at 1)
    return prestige_count * 100 + level_in_prestige + 1

# -------- API helpers --------

def get_uuid(username: str) -> str:
    r = requests.get(f"https://api.mojang.com/users/profiles/minecraft/{username}", timeout=15)
    r.raise_for_status()
    data = r.json()
    return data["id"]

def get_hypixel_player(uuid: str, api_key: str) -> Dict:
    r = requests.get(
        "https://api.hypixel.net/v2/player",
        headers={"API-Key": api_key},
        params={"uuid": uuid},
        timeout=20,
    )
    r.raise_for_status()
    return r.json()

def get_hypixel_guild(uuid: str, api_key: str) -> Dict:
    """Fetch guild information for a player from Hypixel API."""
    r = requests.get(
        "https://api.hypixel.net/v2/guild",
        headers={"API-Key": api_key},
        params={"player": uuid},
        timeout=20,
    )
    r.raise_for_status()
    return r.json()

def extract_guild_info(guild_json: Dict) -> tuple[Optional[str], Optional[str]]:
    """Extract guild tag and tag color from Hypixel guild API response.
    
    Returns (tag, tagColor) or (None, None) if not in a guild.
    """
    if not isinstance(guild_json, dict):
        return None, None
    
    guild = guild_json.get("guild")
    if not guild or not isinstance(guild, dict):
        return None, None
    
    tag = guild.get("tag")
    tag_color = guild.get("tagColor")
    
    return tag, tag_color

def extract_player_rank(player_json: Dict) -> Optional[str]:
    """Extract the player's rank from Hypixel API response.
    
    Returns rank in order of priority, skipping "NONE" values:
    rank, monthlyPackageRank, newPackageRank, packageRank
    """
    player = player_json.get("player", {}) if isinstance(player_json, dict) else {}
    if not isinstance(player, dict):
        return None
    
    # Check in order of priority, skip "NONE" values
    rank = player.get("rank")
    if rank and rank.upper() != "NONE":
        return rank
    
    monthly = player.get("monthlyPackageRank")
    if monthly and monthly.upper() != "NONE":
        return monthly
    
    new_package = player.get("newPackageRank")
    if new_package and new_package.upper() != "NONE":
        return new_package
    
    package = player.get("packageRank")
    if package and package.upper() != "NONE":
        return package
    
    return None

def extract_wool_games_flat(player_json: Dict) -> Dict:
    """Extract Wool Games data and flatten into a single stats dict.
    Includes progression.available_layers/experience, coins, sheep_wars.stats*, and playtime.
    """
    player = player_json.get("player", {}) if isinstance(player_json, dict) else {}
    stats_root = player.get("stats", {}) if isinstance(player, dict) else {}
    # Try common wool keys
    wool_keys = ["WOOL_GAMES", "Wool_Games", "WoolGames", "WoolWars"]
    wool = None
    for k in wool_keys:
        if k in stats_root:
            wool = stats_root[k]
            break
    if not isinstance(wool, dict):
        return {}

    flat: Dict[str, float] = {}

    # progression fields
    progression = wool.get("progression")
    if isinstance(progression, dict):
        if "available_layers" in progression:
            flat["available_layers"] = progression.get("available_layers")
        if "experience" in progression:
            exp_val = progression.get("experience") or 0
            flat["experience"] = exp_val
            # Derive wool games level from experience with prestige scaling
            try:
                flat["level"] = experience_to_level(exp_val)
            except Exception:
                flat["level"] = 0

    # coins
    if "coins" in wool:
        flat["coins"] = wool.get("coins")

    # sheep_wars stats
    sheep_stats = (wool.get("sheep_wars", {}) or {}).get("stats")
    if isinstance(sheep_stats, dict):
        for k, v in sheep_stats.items():
            flat[k] = v

    # playtime
    if "playtime" in wool:
        flat["playtime"] = wool.get("playtime")

    return flat

# -------- Excel helpers (horizontal layout) --------

def ensure_workbook(path: str):
    """Load workbook or create new one if file doesn't exist or is corrupted."""
    if os.path.exists(path):
        try:
            return load_workbook(path)
        except Exception as e:
            print(f"[WARNING] Failed to load Excel file (possibly corrupted): {e}")
            print(f"[WARNING] Creating backup and starting with new workbook")
            # Backup corrupted file
            backup_path = f"{path}.corrupted.bak"
            try:
                import shutil
                shutil.copy(path, backup_path)
                print(f"[INFO] Backed up corrupted file to {backup_path}")
            except Exception:
                pass
            # Remove corrupted file and create new workbook
            try:
                os.remove(path)
            except Exception:
                pass
            return Workbook()
    else:
        return Workbook()


def title_and_headers(ws, start_col: int, title: str, stat_list):
    # Determine how many columns the block spans: 4 columns per stat (Stat, Value, Snapshot, Value)
    block_cols = len(stat_list) * 4
    # Title row = 1, merged across the block
    from openpyxl.utils import get_column_letter
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(start_col + block_cols - 1)
    ws.merge_cells(f"{start_letter}1:{end_letter}1")
    ws.cell(row=1, column=start_col).value = title
    ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=start_col).font = Font(bold=True)

    # Row 2: repeated headers for each stat group
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    col = start_col
    for stat in stat_list:
        ws.cell(row=2, column=col, value=stat).font = Font(bold=True)
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col).fill = header_fill
        ws.cell(row=2, column=col).border = border

        ws.cell(row=2, column=col + 1, value="Value").font = Font(bold=True)
        ws.cell(row=2, column=col + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col + 1).fill = header_fill
        ws.cell(row=2, column=col + 1).border = border

        ws.cell(row=2, column=col + 2, value="Snapshot").font = Font(bold=True)
        ws.cell(row=2, column=col + 2).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col + 2).fill = header_fill
        ws.cell(row=2, column=col + 2).border = border

        ws.cell(row=2, column=col + 3, value="Value").font = Font(bold=True)
        ws.cell(row=2, column=col + 3).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col + 3).fill = header_fill
        ws.cell(row=2, column=col + 3).border = border

        col += 4


def write_block(ws, start_col: int, current_stats: Dict[str, float], snapshot_row: int):
    # Writes a single data row (row 3) for the block.
    # If snapshot columns (row 3, col+2/col+3) are empty, initialize snapshot with current stats.
    from openpyxl.utils import get_column_letter
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    align = Alignment(horizontal="center")

    col = start_col
    for stat in STAT_ORDER:
        # Current value (for Session/Daily, this will be filled by caller as the delta)
        curr_val = current_stats.get(stat, 0)
        ws.cell(row=3, column=col, value=stat).alignment = align
        ws.cell(row=3, column=col).border = border

        ws.cell(row=3, column=col + 1, value=curr_val).alignment = align
        ws.cell(row=3, column=col + 1).border = border

        snap_label_cell = ws.cell(row=snapshot_row, column=col + 2)
        snap_value_cell = ws.cell(row=snapshot_row, column=col + 3)
        # initialize snapshot if empty
        if snap_label_cell.value is None and snap_value_cell.value is None:
            snap_label_cell.value = "Snapshot"
            snap_value_cell.value = current_stats.get(stat, 0)
        snap_label_cell.alignment = align
        snap_value_cell.alignment = align
        snap_label_cell.border = border
        snap_value_cell.border = border

        col += 4


def compute_deltas(current_all_time: Dict[str, float], snapshot_values: Dict[str, float]) -> Dict[str, float]:
    # Deltas for counts; ratios based on delta counts
    kills_delta = (current_all_time.get("Kills", 0) or 0) - (snapshot_values.get("Kills", 0) or 0)
    deaths_delta = (current_all_time.get("Deaths", 0) or 0) - (snapshot_values.get("Deaths", 0) or 0)
    wins_delta = (current_all_time.get("Wins", 0) or 0) - (snapshot_values.get("Wins", 0) or 0)
    losses_delta = (current_all_time.get("Losses", 0) or 0) - (snapshot_values.get("Losses", 0) or 0)
    # Ratios from deltas
    kd = round(kills_delta / deaths_delta, 2) if deaths_delta else float(kills_delta)
    wl = round(wins_delta / losses_delta, 2) if losses_delta else float(wins_delta)
    return {
        "Kills": kills_delta,
        "Deaths": deaths_delta,
        "K/D": kd,
        "Wins": wins_delta,
        "Losses": losses_delta,
        "W/L": wl,
    }


def read_snapshot_row(ws, start_col: int) -> Dict[str, float]:
    # Reads the snapshot 'Value' from row 3, col+3 for each stat.
    vals: Dict[str, float] = {}
    col = start_col
    for stat in STAT_ORDER:
        val = ws.cell(row=3, column=col + 3).value
        vals[stat] = float(val or 0)
        col += 4
    return vals


def get_rank_color(rank: Optional[str]) -> str:
    """Get the default color for a rank.
    
    Returns hex color code based on rank priority.
    """
    if not rank:
        return "#FFFFFF"  # White for no rank
    
    rank_upper = rank.upper()
    
    # Rank color mapping
    rank_colors = {
        "ADMIN": "#FF5555",           # Red (c)
        "SUPERSTAR": "#FFAA00",       # Gold (6)
        "MVP_PLUS": "#55FFFF",        # Aqua (b)
        "MVP_PLUS_PLUS": "#FFAA00",   # Gold (6) - MVP++ permanent is gold too
        "MVP": "#55FFFF",             # Aqua (b)
        "VIP_PLUS": "#00AA00",        # Dark Green (2)
        "VIP": "#55FF55",             # Green (a)
    }
    
    return rank_colors.get(rank_upper, "#FFFFFF")  # Default to white


def save_user_color_and_rank(username: str, rank: Optional[str], guild_tag: Optional[str] = None, guild_color: Optional[str] = None):
    """Save or update user's rank and guild info in user_colors.json.
    
    Only assigns color automatically for NEW users based on their rank.
    Existing users keep their custom color.
    
    Structure: {"username_lowercase": {"color": "#FFFFFF", "rank": "MVP_PLUS", "guild_tag": "QUEBEC", "guild_color": "DARK_AQUA"}}
    """
    colors_file = SCRIPT_DIR / "user_colors.json"
    
    # Load existing data
    color_data = {}
    if colors_file.exists():
        try:
            with open(colors_file, 'r') as f:
                color_data = json.load(f)
        except Exception:
            pass
    
    username_key = username.lower()
    
    # Check if user already exists
    if username_key in color_data:
        # User exists - only update rank and guild info, preserve their color
        existing_entry = color_data[username_key]
        print(f"[DEBUG] User {username} already exists with data: {existing_entry}")
        
        if isinstance(existing_entry, str):
            # Old format (just color string), convert to new format
            print(f"[DEBUG] Converting old format to new format, preserving color: {existing_entry}")
            color_data[username_key] = {
                "color": existing_entry, 
                "rank": rank,
                "guild_tag": guild_tag,
                "guild_color": guild_color
            }
        else:
            # New format, update rank and guild info, preserve color
            old_color = existing_entry.get("color")
            print(f"[DEBUG] Preserving existing color {old_color}, updating rank to {rank}, guild: {guild_tag}")
            color_data[username_key]["rank"] = rank
            color_data[username_key]["guild_tag"] = guild_tag
            color_data[username_key]["guild_color"] = guild_color
    else:
        # NEW USER - assign color based on rank automatically
        auto_color = get_rank_color(rank)
        print(f"[DEBUG] NEW USER {username} - assigning auto color {auto_color} for rank {rank}, guild: {guild_tag}")
        color_data[username_key] = {
            "color": auto_color, 
            "rank": rank,
            "guild_tag": guild_tag,
            "guild_color": guild_color
        }
    
    # Save back to file
    print(f"[DEBUG] Saving to file: {username_key} -> {color_data[username_key]}")
    with open(colors_file, 'w') as f:
        json.dump(color_data, f, indent=2)


def api_update_sheet(username: str, api_key: str, snapshot_sections: set[str] | None = None):
    uuid = get_uuid(username)
    data = get_hypixel_player(uuid, api_key)
    current = extract_wool_games_flat(data)
    if not current:
        raise RuntimeError(f"No Wool Games -> Sheep Wars stats for {username}")

    # Fetch guild information
    print(f"[DEBUG] Fetching guild information for {username} (UUID: {uuid})")
    try:
        guild_data = get_hypixel_guild(uuid, api_key)
        # Save guild data to file for inspection
        guild_file = SCRIPT_DIR / "guild_info.json"
        with open(guild_file, 'w') as f:
            json.dump(guild_data, f, indent=2)
        print(f"[DEBUG] Guild data saved to guild_info.json")
        guild_tag, guild_color = extract_guild_info(guild_data)
        print(f"[DEBUG] Extracted guild tag: {guild_tag}, color: {guild_color}")
    except Exception as e:
        print(f"[DEBUG] Failed to fetch guild data: {e}")
        guild_data = None
        guild_tag, guild_color = None, None

    # Extract and save player rank and guild info to user_colors.json
    rank = extract_player_rank(data)
    print(f"[DEBUG] Extracted rank for {username}: {rank}")
    save_user_color_and_rank(username, rank, guild_tag, guild_color)

    # Open workbook (create if missing) and create/select player's sheet
    wb = ensure_workbook(EXCEL_FILE)
    
    try:
        # Detect if existing sheet already uses the single-table layout; if not, rebuild it.
        def is_single_table_layout(ws) -> bool:
            expected = [
                "Stat",
                "All-time",
                "Session Delta",
                "Session Snapshot",
                "Daily Delta",
                "Daily Snapshot",
                "Yesterday Delta",
                "Yesterday Snapshot",
                "Monthly Delta",
                "Monthly Snapshot",
            ]
            for i, title in enumerate(expected, start=1):
                cell_val = ws.cell(row=1, column=i).value
                if (cell_val or "").strip() != title:
                    return False
            return True

        if username in wb.sheetnames:
            ws = wb[username]
            # If not the new layout, clear and rebuild
            if not is_single_table_layout(ws):
                ws.delete_rows(1, ws.max_row)
                ws.delete_cols(1, ws.max_column)
                new_sheet = True
            else:
                new_sheet = False
        else:
            ws = wb.create_sheet(username)
            new_sheet = True

        # Build ordered key list: preferred first, then any remaining keys
        preferred = [
            "available_layers",
            "experience",
            "level",
            "coins",
            "damage_dealt",
            "deaths",
            "deaths_explosive",
            "games_played",
            "losses",
            "sheep_thrown",
            "deaths_bow",
            "deaths_void",
            "wins",
            "kills",
            "kills_void",
            "magic_wool_hit",
            "kills_explosive",
            "kills_melee",
            "deaths_melee",
            "kills_bow",
            "playtime",
        ]
        ordered_keys = []
        seen = set()
        for k in preferred:
            if k in current and k not in seen:
                ordered_keys.append(k)
                seen.add(k)
        for k in sorted(current.keys()):
            if k not in seen:
                ordered_keys.append(k)
                seen.add(k)

        # Single table layout (no gaps):
        # A: Stat label, B: All-time, C: Session Delta, D: Session Snapshot,
        # E: Daily Delta, F: Daily Snapshot, G: Yesterday Delta, H: Yesterday Snapshot,
        # I: Monthly Delta, J: Monthly Snapshot
        label_col = 1
        at_col = 2
        sess_delta_col, sess_snap_col = 3, 4
        daily_delta_col, daily_snap_col = 5, 6
        yest_delta_col, yest_snap_col = 7, 8
        mon_delta_col, mon_snap_col = 9, 10

        # Headers row
        headers = [
            (label_col, "Stat"),
            (at_col, "All-time"),
            (sess_delta_col, "Session Delta"),
            (sess_snap_col, "Session Snapshot"),
            (daily_delta_col, "Daily Delta"),
            (daily_snap_col, "Daily Snapshot"),
            (yest_delta_col, "Yesterday Delta"),
            (yest_snap_col, "Yesterday Snapshot"),
            (mon_delta_col, "Monthly Delta"),
            (mon_snap_col, "Monthly Snapshot"),
        ]
        for col, title in headers:
            c = ws.cell(row=1, column=col)
            c.value = title
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        # Populate rows
        for idx, key in enumerate(ordered_keys):
            row = 2 + idx
            # Label
            ws.cell(row=row, column=label_col, value=key)
            # All-time value
            ws.cell(row=row, column=at_col, value=current.get(key))
            # Reserve delta and snapshot cells. We'll compute numeric deltas after
            # snapshot columns are optionally updated below so deltas reflect
            # the correct (current - snapshot_before_update) values or zero when
            # snapshot is set to current.
            if new_sheet:
                # initialize snapshot cells to None for clarity
                ws.cell(row=row, column=sess_snap_col, value=None)
                ws.cell(row=row, column=daily_snap_col, value=None)
                ws.cell(row=row, column=yest_snap_col, value=None)
                ws.cell(row=row, column=mon_snap_col, value=None)
            # leave delta cells empty for now
            ws.cell(row=row, column=sess_delta_col, value=None)
            ws.cell(row=row, column=daily_delta_col, value=None)
            ws.cell(row=row, column=yest_delta_col, value=None)
            ws.cell(row=row, column=mon_delta_col, value=None)

        # If snapshot flags provided, write snapshot values into appropriate snapshot columns
        snapshot_sections = snapshot_sections or set()
        def write_snapshot_column(col_idx: int):
            for idx, key in enumerate(ordered_keys):
                row = 2 + idx
                ws.cell(row=row, column=col_idx, value=current.get(key))

        if "session" in snapshot_sections:
            write_snapshot_column(sess_snap_col)
        if "daily" in snapshot_sections:
            write_snapshot_column(daily_snap_col)
        if "yesterday" in snapshot_sections:
            write_snapshot_column(yest_snap_col)
        if "monthly" in snapshot_sections:
            write_snapshot_column(mon_snap_col)

        # Compute numeric deltas for each period (current - snapshot). Treat missing
        # snapshot values as 0. This ensures the bot (which reads cell values)
        # sees numeric deltas rather than formula strings.
        for idx, key in enumerate(ordered_keys):
            row = 2 + idx
            cur = current.get(key) or 0
            # Session delta
            snap_val = ws.cell(row=row, column=sess_snap_col).value
            try:
                snap_val = float(snap_val or 0)
            except Exception:
                # remove commas and try
                try:
                    snap_val = float(str(snap_val).replace(",", ""))
                except Exception:
                    snap_val = 0
            ws.cell(row=row, column=sess_delta_col, value=(cur - snap_val))
            # Daily delta
            snap_val = ws.cell(row=row, column=daily_snap_col).value
            try:
                snap_val = float(snap_val or 0)
            except Exception:
                try:
                    snap_val = float(str(snap_val).replace(",", ""))
                except Exception:
                    snap_val = 0
            ws.cell(row=row, column=daily_delta_col, value=(cur - snap_val))
            # Yesterday delta
            snap_val = ws.cell(row=row, column=yest_snap_col).value
            try:
                snap_val = float(snap_val or 0)
            except Exception:
                try:
                    snap_val = float(str(snap_val).replace(",", ""))
                except Exception:
                    snap_val = 0
            ws.cell(row=row, column=yest_delta_col, value=(cur - snap_val))
            # Monthly delta
            snap_val = ws.cell(row=row, column=mon_snap_col).value
            try:
                snap_val = float(snap_val or 0)
            except Exception:
                try:
                    snap_val = float(str(snap_val).replace(",", ""))
                except Exception:
                    snap_val = 0
            ws.cell(row=row, column=mon_delta_col, value=(cur - snap_val))

        # Force Excel to recalculate formulas on load so deltas show correctly
        try:
            wb.calculation_properties.fullCalcOnLoad = True
        except Exception:
            # older openpyxl versions may not have calculation_properties
            pass

        wb.save(EXCEL_FILE)
        return {
            "uuid": uuid,
            "stats": current,
            "excel": EXCEL_FILE,
        }
    
    except Exception as e:
        # Always close the workbook if an error occurs
        print(f"[ERROR] Exception occurred during Excel update: {e}")
        try:
            wb.close()
        except Exception:
            pass
        raise
    finally:
        # Ensure workbook is always closed
        try:
            wb.close()
        except Exception:
            pass


def main():
    parser = argparse.ArgumentParser(description="API-based Wool Games stats to Excel (horizontal)")
    parser.add_argument("-ign", "--username", required=True, help="Minecraft IGN")
    # API key must be provided via API_KEY.txt; no CLI/env flag
    parser.add_argument("-session", action="store_true", help="Write snapshot values into Session block")
    parser.add_argument("-daily", action="store_true", help="Write snapshot values into Daily block")
    parser.add_argument("-yesterday", action="store_true", help="Write snapshot values into Yesterday block")
    parser.add_argument("-monthly", action="store_true", help="Write snapshot values into Monthly block")
    args = parser.parse_args()

    # Only use the API key from API_KEY.txt (no CLI/env/default fallback)
    api_key = read_api_key_file()
    if not api_key:
        raise RuntimeError(
            "Missing API key: create API_KEY.txt next to api_get.py containing your Hypixel API key"
        )
    sections = set()
    if args.session:
        sections.add("session")
    if args.daily:
        sections.add("daily")
    if args.yesterday:
        sections.add("yesterday")
    if args.monthly:
        sections.add("monthly")

    res = api_update_sheet(args.username, api_key, snapshot_sections=sections)
    print(res)

if __name__ == "__main__":
    main()
